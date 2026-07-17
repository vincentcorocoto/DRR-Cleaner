import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False
try:
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    _MPL_AVAILABLE = True
except ImportError:
    _MPL_AVAILABLE = False
import os
import io
import re
import json
import difflib
import sqlite3
import hashlib
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.text import RichText
from openpyxl.chart.plotarea import DataTable
from openpyxl.drawing.text import RichTextProperties, Paragraph, ParagraphProperties, CharacterProperties
import openpyxl

# ══════════════════════════════════════════════════════════════════════════════
# THEME
# ══════════════════════════════════════════════════════════════════════════════
BG        = "#0F1117"
CARD      = "#1A1D27"
BORDER    = "#2A2D3E"
ACCENT    = "#4F8EF7"
ACCENT2   = "#7C3AED"
SUCCESS   = "#22C55E"
DANGER    = "#EF4444"
WARNING   = "#F59E0B"
TEXT      = "#F1F5F9"
SUBTEXT   = "#94A3B8"
HEADER_BG = "#1E2235"
DARK      = "#0A0C14"
GOLD      = "#F59E0B"

# ══════════════════════════════════════════════════════════════════════════════
# DRR CLEANER CONFIG
# ══════════════════════════════════════════════════════════════════════════════
STATUS_COL_NAME    = "Status"
STATUS_COL_INDEX   = 9
DATE_COL_NAME      = "Date"
TIME_COL_NAME      = "Time"
DPD_COL_NAME       = "DPD"
COL_E_INDEX        = 4
DIALED_COL_INDEX   = 5
PTP_AMOUNT_COL     = "PTP Amount"
CLAIM_PAID_COL     = "Claim Paid Amount"
REMARK_COL         = "Remark"
REMOVE_STATUSES    = {"BP", "REACTIVE", "SMS FAILED", "NEW", "ABORTED"}
COLS_TO_DROP_START = 27
COLS_TO_DROP_END   = 50

# ══════════════════════════════════════════════════════════════════════════════
# ANALYTICS — MASTERFILE HEADER TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════
MASTERLIST_HEADERS = [
    "Placement", "ENDO DATE", "CYCLE", "CHCODE", "FILE_NO",
    "LOAN_ACCOUNT_NUMBER", "CUSTOMER_ID", "FULL_NAME", "Email Address",
    "Contact No.", "Payment Stage", "Amount to Collect", "Loan Receivable",
    "BOM OS", "STATE",
]

def _display_date_ddmmyyyy(val):
    """Grid display only: 'mm/dd/yyyy' -> 'dd-mm-yyyy'. Raw value stays mm/dd/yyyy."""
    if not val:
        return val
    try:
        return datetime.strptime(val, "%m/%d/%Y").strftime("%d-%m-%Y")
    except (ValueError, TypeError):
        return val

def _display_time_only(val):
    """Grid display only: 'mm/dd/yyyy  h:mm:ss AM/PM' -> 'h:mm:ss AM/PM'. Raw value keeps the date."""
    if not val:
        return val
    parts = str(val).split("  ", 1)
    return parts[1] if len(parts) == 2 else val

# ══════════════════════════════════════════════════════════════════════════════
# DATABASE — SQLite user store
# ══════════════════════════════════════════════════════════════════════════════
import sys

def _app_data_dir():
    """Persistent, writable folder for app data.
    - Running as a script: same folder as this .py file.
    - Running as a frozen PyInstaller .exe: a SHARED, machine-wide folder
      (%PROGRAMDATA%\\CleanDRR on Windows, or /var/tmp/CleanDRR / a fixed
      folder under the home directory on other OSes) — NOT the per-user
      %APPDATA%, since the user database needs to be visible to every
      Windows account/session on the machine, not just the one that created
      it. PyInstaller onefile's temp extraction folder is wiped after every
      run, which is why this can't just live next to the .exe either.
      Falls back to the per-user APPDATA/home folder if the shared location
      isn't writable (e.g. restricted permissions)."""
    if getattr(sys, "frozen", False):
        shared_base = os.environ.get("PROGRAMDATA") or os.environ.get("ALLUSERSPROFILE")
        candidates = []
        if shared_base:
            candidates.append(os.path.join(shared_base, "CleanDRR"))
        candidates.append(os.path.join(os.environ.get("APPDATA") or os.path.expanduser("~"), "CleanDRR"))
        for folder in candidates:
            try:
                os.makedirs(folder, exist_ok=True)
                test_file = os.path.join(folder, ".write_test")
                with open(test_file, "w") as f:
                    f.write("ok")
                os.remove(test_file)
                return folder
            except OSError:
                continue
        # Last resort: home directory
        folder = os.path.join(os.path.expanduser("~"), "CleanDRR")
        os.makedirs(folder, exist_ok=True)
        return folder
    return os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(_app_data_dir(), "users.db")

def _hash(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def db_init():
    """Create the users table and seed default accounts if empty."""
    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            created_at TEXT
        )
    """)
    con.commit()
    # Geo Reference table — persists across app restarts, lives in the same
    # shared DB file as users (see _app_data_dir), so nothing is lost when
    # the app is closed.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS geo_reference (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            unique_code   TEXT,
            province      TEXT,
            municipality  TEXT,
            final_area    TEXT,
            geocode       TEXT,
            cluster       TEXT,
            area_status   TEXT,
            created_at    TEXT
        )
    """)
    con.commit()
    # Seed defaults only when the table is empty
    if cur.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        seeds = [
            ("admin",      "admin123",    "admin"),
            ("supervisor", "super2024",   "supervisor"),
            ("vincent",    "orico2024",   "user"),
            ("user",       "password",    "user"),
        ]
        for u, p, r in seeds:
            cur.execute("INSERT INTO users VALUES (?,?,?,datetime('now','localtime'))",
                        (u, _hash(p), r))
        con.commit()

    con.close()

def db_check_login(username: str, password: str):
    """Return role string on success, None on failure."""
    con = sqlite3.connect(DB_PATH)
    row = con.execute(
        "SELECT role FROM users WHERE username=? AND password_hash=?",
        (username, _hash(password))
    ).fetchone()
    con.close()
    return row[0] if row else None

def db_list_users():
    """Return list of (username, role, created_at)."""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT username, role, created_at FROM users ORDER BY username"
    ).fetchall()
    con.close()
    return rows

def db_add_user(username: str, password: str, role: str = "user") -> str:
    """Add a new user. Returns '' on success or an error message."""
    if not username or not password:
        return "Username and password are required."
    try:
        con = sqlite3.connect(DB_PATH)
        con.execute("INSERT INTO users (username, password_hash, role, created_at) VALUES (?,?,?,datetime('now','localtime'))",
                    (username, _hash(password), role))
        con.commit()
        con.close()
        return ""
    except sqlite3.IntegrityError:
        return f"Username '{username}' already exists."

def db_delete_user(username: str) -> str:
    """Delete a user. Prevents deleting the last admin."""
    con = sqlite3.connect(DB_PATH)
    admin_count = con.execute(
        "SELECT COUNT(*) FROM users WHERE role='admin'"
    ).fetchone()[0]
    role = con.execute(
        "SELECT role FROM users WHERE username=?", (username,)
    ).fetchone()
    if role and role[0] == "admin" and admin_count <= 1:
        con.close()
        return "Cannot delete the only admin account."
    con.execute("DELETE FROM users WHERE username=?", (username,))
    con.commit()
    con.close()
    return ""

def db_update_user(username: str, new_password: str = "", new_role: str = "") -> str:
    """Update password and/or role for a user."""
    con = sqlite3.connect(DB_PATH)
    if new_password:
        con.execute("UPDATE users SET password_hash=? WHERE username=?",
                    (_hash(new_password), username))
    if new_role:
        con.execute("UPDATE users SET role=? WHERE username=?",
                    (new_role, username))
    con.commit()
    con.close()
    return ""

# ── Geo Reference table helpers ──────────────────────────────────────────────
GEO_REFERENCE_COLUMNS = [
    ("unique_code",  "UNIQUE"),
    ("province",     "PROVINCE"),
    ("municipality", "MUNICIPALITY"),
    ("final_area",   "FINAL AREA"),
    ("geocode",      "GEOCODE"),
    ("cluster",      "CLUSTER"),
    ("area_status",  "AREA STATUS"),
]

def db_list_geo_reference():
    """Return every saved Geo Reference row as
    (id, unique_code, province, municipality, final_area, geocode, cluster, area_status, created_at)."""
    con = sqlite3.connect(DB_PATH)
    rows = con.execute(
        "SELECT id, unique_code, province, municipality, final_area, geocode, "
        "cluster, area_status, created_at FROM geo_reference ORDER BY id"
    ).fetchall()
    con.close()
    return rows

def db_add_geo_reference(unique_code, province, municipality, final_area,
                          geocode, cluster, area_status):
    """Insert one Geo Reference row. Returns '' on success or an error message."""
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "INSERT INTO geo_reference "
        "(unique_code, province, municipality, final_area, geocode, cluster, area_status, created_at) "
        "VALUES (?,?,?,?,?,?,?, datetime('now','localtime'))",
        (unique_code, province, municipality, final_area, geocode, cluster, area_status)
    )
    con.commit()
    con.close()
    return ""

def db_update_geo_reference(row_id, unique_code, province, municipality,
                             final_area, geocode, cluster, area_status):
    """Update one Geo Reference row by id. Returns '' on success."""
    con = sqlite3.connect(DB_PATH)
    con.execute(
        "UPDATE geo_reference SET unique_code=?, province=?, municipality=?, "
        "final_area=?, geocode=?, cluster=?, area_status=? WHERE id=?",
        (unique_code, province, municipality, final_area, geocode, cluster, area_status, row_id)
    )
    con.commit()
    con.close()
    return ""

def db_delete_geo_reference(row_id):
    """Delete one Geo Reference row by id. Returns '' on success."""
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM geo_reference WHERE id=?", (row_id,))
    con.commit()
    con.close()
    return ""

def db_clear_geo_reference():
    """Delete every Geo Reference row. Returns '' on success."""
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM geo_reference")
    con.commit()
    con.close()
    return ""

def db_import_geo_reference_bulk(rows):
    """Bulk-insert Geo Reference rows.
    rows: iterable of (unique_code, province, municipality, final_area, geocode, cluster, area_status).
    Returns the number of rows inserted."""
    con = sqlite3.connect(DB_PATH)
    con.executemany(
        "INSERT INTO geo_reference "
        "(unique_code, province, municipality, final_area, geocode, cluster, area_status, created_at) "
        "VALUES (?,?,?,?,?,?,?, datetime('now','localtime'))",
        list(rows)
    )
    con.commit()
    con.close()
    return len(rows)

# Initialise DB on import
db_init()

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def show_toast(parent, message, color=SUCCESS, duration=3000):
    toast = tk.Toplevel(parent)
    toast.overrideredirect(True)
    toast.attributes("-topmost", True)
    toast.attributes("-alpha", 0.95)
    frame = tk.Frame(toast, bg=color, padx=20, pady=12)
    frame.pack()
    tk.Label(frame, text="✔  " + message, font=("Segoe UI", 10, "bold"),
             bg=color, fg="white").pack()
    parent.update_idletasks()
    pw, ph = parent.winfo_width(), parent.winfo_height()
    px, py = parent.winfo_x(), parent.winfo_y()
    toast.update_idletasks()
    tw, th = toast.winfo_reqwidth(), toast.winfo_reqheight()
    toast.geometry(f"+{px + pw - tw - 20}+{py + ph - th - 40}")

    def fade(alpha=0.95):
        alpha -= 0.05
        if alpha <= 0:
            toast.destroy()
        else:
            toast.attributes("-alpha", alpha)
            toast.after(40, lambda: fade(alpha))

    toast.after(duration, fade)


# ══════════════════════════════════════════════════════════════════════════════
# DRR PROCESSING LOGIC
# ══════════════════════════════════════════════════════════════════════════════
def process_file(filepath):
    df_headers = pd.read_excel(filepath, nrows=0)
    col_e_name      = df_headers.columns[COL_E_INDEX]      if COL_E_INDEX      < len(df_headers.columns) else None
    dialed_col_name = df_headers.columns[DIALED_COL_INDEX] if DIALED_COL_INDEX < len(df_headers.columns) else None
    for col in df_headers.columns:
        if "dialed" in str(col).lower():
            dialed_col_name = col
            break

    # The DPD column must always be retained exactly as it is in the raw file —
    # never treated as the positional "account" column (COL_E_INDEX) or a
    # "dialed" column.
    dpd_col_name = None
    for col in df_headers.columns:
        if str(col).strip().upper() == DPD_COL_NAME:
            dpd_col_name = col
            break
    if dpd_col_name:
        if col_e_name == dpd_col_name:
            col_e_name = None
        if dialed_col_name == dpd_col_name:
            dialed_col_name = None

    # Read every column as raw text. This guarantees the cleaned/removed
    # output matches the raw file exactly (no lost leading zeros, no
    # appended ".0", no auto-parsed dates/numbers) for every column that
    # isn't deliberately transformed below (account/dialed cleanup,
    # Date/Time reformatting, Remark SRP rewrite, row removal).
    df = pd.read_excel(filepath, dtype=str)
    _sno_col = df.columns[0]  # Remember S.No column name for renumbering after clean


    def vectorized_strip_decimal(series):
        s = series.fillna("").astype(str).str.strip()
        has_dot   = s.str.contains(".", regex=False)
        int_part  = s.str.split(".").str[0]
        is_numeric = int_part.str.lstrip("-").str.isdigit()
        return s.where(~(has_dot & is_numeric), int_part).replace("nan", "")

    def vectorized_clean_account(series):
        s        = series.fillna("").astype(str).str.strip()
        ends_dot0 = s.str.endswith(".0")
        base      = s.str[:-2]
        is_safe   = ends_dot0 & base.str.lstrip("-").str.isdigit() & ~base.str.startswith("0")
        return s.where(~is_safe, base).replace("nan", "")

    # Apply account/dialed cleanup — but NEVER touch the DPD column.
    if col_e_name and col_e_name in df.columns and col_e_name != dpd_col_name:
        df[col_e_name] = vectorized_clean_account(df[col_e_name])
    if dialed_col_name and dialed_col_name in df.columns and dialed_col_name != dpd_col_name:
        df[dialed_col_name] = vectorized_strip_decimal(df[dialed_col_name])
    else:
        for col in df.columns:
            if "dialed" in str(col).lower() and col != dpd_col_name:
                df[col] = vectorized_strip_decimal(df[col])
                dialed_col_name = col
                break

    # Restore DPD column: Excel reads integers as floats ("30" -> "30.0").
    # Strip the trailing ".0" for whole-number DPD values so the output
    # matches the raw file exactly (e.g. "30" not "30.0"). Non-integer
    # DPD values (rare) are kept as-is.
    if dpd_col_name and dpd_col_name in df.columns:
        def _fix_dpd(val):
            s = str(val).strip()
            if s.lower() in ("", "nan", "none"):
                return ""
            if s.endswith(".0"):
                base = s[:-2]
                if base.lstrip("-").isdigit():
                    return base
            return s
        df[dpd_col_name] = df[dpd_col_name].apply(_fix_dpd)

    STATUS_COL = STATUS_COL_NAME if STATUS_COL_NAME in df.columns else (
        df.columns[STATUS_COL_INDEX] if STATUS_COL_INDEX < len(df.columns) else None
    )
    if not STATUS_COL:
        raise ValueError("Status column not found.")

    status_normalized = df[STATUS_COL].astype(str).str.strip().str.upper()
    is_blank          = df[STATUS_COL].isna() | (df[STATUS_COL].astype(str).str.strip() == "")
    has_cease         = status_normalized.str.contains("CEASE", na=False)
    is_removable      = status_normalized.isin(REMOVE_STATUSES) | is_blank | has_cease

    if PTP_AMOUNT_COL in df.columns:
        has_ptp       = status_normalized.str.contains(r"\bPTP\b", regex=True, na=False)
        ptp_numeric   = pd.to_numeric(df[PTP_AMOUNT_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        ptp_has_value = has_ptp & (ptp_numeric > 0)
        ptp_no_value  = has_ptp & (ptp_numeric <= 0)
        is_removable  = (is_removable | ptp_no_value) & ~ptp_has_value

    if CLAIM_PAID_COL in df.columns:
        has_kept       = status_normalized.str.contains(r"\bKEPT\b", regex=True, na=False)
        claim_numeric  = pd.to_numeric(df[CLAIM_PAID_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        kept_has_value = has_kept & (claim_numeric > 0)
        kept_no_value  = has_kept & (claim_numeric <= 0)
        is_removable   = (is_removable | kept_no_value) & ~kept_has_value

    # Always remove rows whose Remark contains any of these auto-generated /
    # system phrases, regardless of Status (case-insensitive, partial match).
    REMARK_EXCLUDE_PHRASES = [
        "System Auto Update Remarks For PD", "Updates when case",
        "New Assignment", "New Contacts",
    ]
    if REMARK_COL in df.columns:
        _remark_exclude_pattern = "|".join(re.escape(p) for p in REMARK_EXCLUDE_PHRASES)
        remark_exclude_mask = df[REMARK_COL].astype(str).str.contains(
            _remark_exclude_pattern, case=False, na=False, regex=True)
        is_removable = is_removable | remark_exclude_mask
    else:
        remark_exclude_mask = pd.Series([False] * len(df), index=df.index)

    cleaned_df = df[~is_removable].copy()
    removed_df = df[is_removable].copy()

    # S.No is left untouched — original values are retained as-is from the
    # raw file, even after rows have been removed (no renumbering).

    s_norm  = removed_df[STATUS_COL].fillna("").astype(str).str.strip()
    s_upper = s_norm.str.upper()

    # Priority order (highest first): a row removed because its Status
    # directly matches a removable status (or contains CEASE) is ALWAYS
    # labeled "Status: ..." — even if its Remark also happens to contain an
    # excluded phrase, or other conditions also apply — so it reliably shows
    # up in the "Removed Status" tab instead of being reclassified.
    is_status_match = s_upper.isin(REMOVE_STATUSES) | s_upper.str.contains("CEASE", na=False)
    is_blank_status = s_upper.isin(["", "NAN"])
    has_ptp_reason   = s_upper.str.contains("PTP",  na=False) & ~is_status_match
    has_kept_reason  = s_upper.str.contains("KEPT", na=False) & ~is_status_match & ~has_ptp_reason
    remark_excluded  = remark_exclude_mask.reindex(removed_df.index, fill_value=False) & \
                        ~is_status_match & ~is_blank_status & ~has_ptp_reason & ~has_kept_reason

    reason = pd.Series("Status: " + s_norm, index=removed_df.index)  # default/fallback
    reason = reason.where(~is_blank_status,  "Blank Status")
    reason = reason.where(~has_ptp_reason,   "PTP with no PTP Amount")
    reason = reason.where(~has_kept_reason,  "KEPT with no Claim Paid Amount")
    reason = reason.where(~remark_excluded,  "Remark contains excluded phrase")
    reason = reason.where(~is_status_match,  "Status: " + s_norm)  # re-assert top priority last
    removed_df.insert(0, "Removed Reason", reason)

    srp_mask   = pd.Series([False] * len(cleaned_df), index=cleaned_df.index)
    remarks_df = pd.DataFrame(columns=["Row #", STATUS_COL,
                                        REMARK_COL + " (Before)",
                                        REMARK_COL + " (After)"])  # always empty DataFrame with correct cols

    if REMARK_COL in cleaned_df.columns:
        remark_norm    = cleaned_df[REMARK_COL].astype(str).str.strip().str.upper()
        status_clean   = cleaned_df[STATUS_COL].astype(str).str.strip().str.upper()
        has_action_ptp = remark_norm.str.contains(r"ACTION\s*:\s*PTP", regex=True, na=False)
        status_not_ptp_kept = (
            ~status_clean.str.contains("PTP",  na=False) &
            ~status_clean.str.contains("KEPT", na=False)
        )
        srp_mask = has_action_ptp & status_not_ptp_kept
        if srp_mask.any():
            # Build remarks log strictly from cleaned_df rows matching srp_mask only
            changed_rows = cleaned_df.loc[srp_mask, [STATUS_COL, REMARK_COL]].copy()
            remarks_df = pd.DataFrame({
                "Row #":                  range(1, srp_mask.sum() + 1),
                STATUS_COL:               changed_rows[STATUS_COL].values,
                REMARK_COL + " (Before)": changed_rows[REMARK_COL].astype(str).values,
                REMARK_COL + " (After)":  changed_rows[REMARK_COL].astype(str).str.replace(
                    r"(?i)Action\s*:\s*PTP", "ACTION: SRP", regex=True).values,
            }).reset_index(drop=True)
        cleaned_df.loc[srp_mask, REMARK_COL] = cleaned_df.loc[srp_mask, REMARK_COL].astype(str).str.replace(
            r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True)

    cols_to_drop = [df.columns[i] for i in range(COLS_TO_DROP_START, min(COLS_TO_DROP_END + 1, len(df.columns)))]
    cleaned_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")
    removed_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")




    if DATE_COL_NAME in cleaned_df.columns and TIME_COL_NAME in cleaned_df.columns:
        # Parse the Date column (Column B) for the date portion
        raw_date = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce")
        # Parse the Time column (Column C) — extract time portion only
        raw_time = pd.to_datetime(cleaned_df[TIME_COL_NAME], errors="coerce")
        date_str = raw_date.dt.strftime("%m/%d/%Y")
        # Time only — strip any date portion from Time column
        time_str = raw_time.dt.strftime("%I:%M:%S %p").str.lstrip("0")
        # Combine: "06/16/2026  4:51:16 PM"
        cleaned_df[TIME_COL_NAME] = date_str + "  " + time_str
    if DATE_COL_NAME in cleaned_df.columns:
        cleaned_df[DATE_COL_NAME] = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m/%d/%Y")

    stats = {
        "total":       len(df),
        "retained":    len(cleaned_df),
        "removed":     len(removed_df),
        "srp_changed": int(srp_mask.sum()),
    }

    # ── Unique ID tab: cleaned data + combined key (Account|Date|Time|Debtor) ──
    # Used to verify that cleaned rows still match the raw file exactly.
    unique_df = cleaned_df.copy()
    _acct_col   = col_e_name if (col_e_name and col_e_name in unique_df.columns) else None
    _date_col   = DATE_COL_NAME if DATE_COL_NAME in unique_df.columns else None
    _time_col   = TIME_COL_NAME if TIME_COL_NAME in unique_df.columns else None
    _debtor_col = next((c for c in unique_df.columns if str(c).strip().lower() == "debtor"), None)

    def _safe_series(col_name):
        if col_name and col_name in unique_df.columns:
            return unique_df[col_name].fillna("").astype(str).str.strip()
        return pd.Series([""] * len(unique_df), index=unique_df.index)

    acct_s   = _safe_series(_acct_col)
    date_s   = _safe_series(_date_col)
    time_raw = _safe_series(_time_col)
    # Time column raw value is "mm/dd/yyyy  h:mm:ss AM/PM" — extract time part only
    time_s   = time_raw.str.split("  ").str[-1].str.strip()
    debtor_s = _safe_series(_debtor_col)

    unique_df.insert(1, "Unique ID", acct_s + "|" + date_s + "|" + time_s + "|" + debtor_s)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, frame in [("Cleaned", cleaned_df), ("Removed", removed_df)]:
            frame.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            normal_fill  = PatternFill("solid", fgColor="1E2235")
            orange_fill  = PatternFill("solid", fgColor="C05621")
            white_bold   = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.font = white_bold
                cell.fill = orange_fill if cell.value == "Removed Reason" else normal_fill
            if DATE_COL_NAME in frame.columns:
                date_col_idx = frame.columns.get_loc(DATE_COL_NAME) + 1
                for row in range(2, ws.max_row + 1):
                    date_cell = ws.cell(row=row, column=date_col_idx)
                    if isinstance(date_cell.value, str) and date_cell.value:
                        try:
                            date_cell.value = datetime.strptime(date_cell.value, "%m/%d/%Y")
                        except ValueError:
                            pass
                    date_cell.number_format = "dd-mm-yyyy"
            if TIME_COL_NAME in frame.columns:
                time_col_idx = frame.columns.get_loc(TIME_COL_NAME) + 1
                for row in range(2, ws.max_row + 1):
                    time_cell = ws.cell(row=row, column=time_col_idx)
                    if isinstance(time_cell.value, str) and time_cell.value.strip():
                        parsed_dt = None
                        for fmt in ("%m/%d/%Y  %I:%M:%S %p", "%m/%d/%Y %I:%M:%S %p"):
                            try:
                                parsed_dt = datetime.strptime(time_cell.value.strip(), fmt)
                                break
                            except ValueError:
                                continue
                        if parsed_dt is not None:
                            time_cell.value = parsed_dt
                    time_cell.number_format = "mm/dd/yyyy hh:mm:ss"
            for col in ws.columns:
                header_len = len(str(col[0].value)) if col[0].value else 10
                ws.column_dimensions[col[0].column_letter].width = min(header_len + 6, 40)
        if not remarks_df.empty:
            remarks_df.to_excel(writer, index=False, sheet_name="Remarks Changes")
            ws = writer.sheets["Remarks Changes"]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="7C3AED")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
        # Write Unique ID sheet
        unique_df.to_excel(writer, index=False, sheet_name="Unique ID")
        ws = writer.sheets["Unique ID"]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E7A4E")
        uid_col_idx = list(unique_df.columns).index("Unique ID") + 1
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=uid_col_idx).fill = PatternFill("solid", fgColor="0F3D27")
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf.seek(0)
    return cleaned_df, removed_df, remarks_df, stats, buf.read(), col_e_name, dialed_col_name, unique_df


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 1 — LOGIN
# ══════════════════════════════════════════════════════════════════════════════
class LoginScreen(tk.Frame):
    def __init__(self, master, on_login):
        super().__init__(master, bg=BG)
        self.on_login = on_login
        self._build()

    def _build(self):
        # Animated gradient banner
        banner = tk.Frame(self, bg=HEADER_BG, height=8)
        banner.pack(fill="x")

        # Center card container
        outer = tk.Frame(self, bg=BG)
        outer.pack(expand=True)

        card = tk.Frame(outer, bg=CARD, padx=48, pady=44,
                        highlightthickness=1, highlightbackground=BORDER)
        card.pack(padx=20, pady=20)

        # Logo area
        logo_frame = tk.Frame(card, bg=CARD)
        logo_frame.pack(pady=(0, 24))

        logo_circle = tk.Canvas(logo_frame, width=72, height=72, bg=CARD,
                                highlightthickness=0)
        logo_circle.pack()
        logo_circle.create_oval(4, 4, 68, 68, fill=ACCENT, outline="")
        logo_circle.create_text(36, 36, text="⬡", font=("Segoe UI", 26, "bold"), fill="white")

        tk.Label(card, text="Welcome Back", font=("Segoe UI", 20, "bold"),
                 bg=CARD, fg=TEXT).pack()
        tk.Label(card, text="Sign in to your account to continue",
                 font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(pady=(4, 28))

        # Username field
        self._field(card, "👤  Username", "username")
        tk.Frame(card, bg=BG, height=10).pack()
        # Password field
        self._field(card, "🔒  Password", "password", show="●")

        # Error label
        self.err_var = tk.StringVar()
        self.err_lbl = tk.Label(card, textvariable=self.err_var,
                                font=("Segoe UI", 9), bg=CARD, fg=DANGER)
        self.err_lbl.pack(pady=(8, 0))

        # Login button
        btn_frame = tk.Frame(card, bg=CARD)
        btn_frame.pack(fill="x", pady=(16, 0))

        self.login_btn = tk.Button(
            btn_frame, text="Sign In  →",
            font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg="white", relief="flat",
            pady=11, cursor="hand2",
            activebackground="#3a7be8", activeforeground="white",
            command=self._attempt_login
        )
        self.login_btn.pack(fill="x")

        # Bind Enter key
        self.master.bind("<Return>", lambda e: self._attempt_login())



        # Bottom bar
        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)

    def _field(self, parent, label_text, attr, show=None):
        lbl = tk.Label(parent, text=label_text, font=("Segoe UI", 9, "bold"),
                       bg=CARD, fg=SUBTEXT, anchor="w")
        lbl.pack(fill="x", pady=(0, 4))

        wrapper = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
        wrapper.pack(fill="x")
        inner = tk.Frame(wrapper, bg=CARD)
        inner.pack(fill="x")

        var = tk.StringVar()
        entry = tk.Entry(inner, textvariable=var,
                         font=("Segoe UI", 11), bg=CARD, fg=TEXT,
                         insertbackground=TEXT, relief="flat", bd=8,
                         show=show or "")
        entry.pack(fill="x")
        setattr(self, f"{attr}_var", var)
        setattr(self, f"{attr}_entry", entry)

        # Focus glow effect
        def on_focus_in(e):
            wrapper.config(bg=ACCENT)
        def on_focus_out(e):
            wrapper.config(bg=BORDER)
        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)

    def _attempt_login(self):
        user = self.username_var.get().strip()
        pw   = self.password_var.get().strip()
        if not user or not pw:
            self.err_var.set("⚠  Please enter username and password.")
            return
        role = db_check_login(user, pw)
        if role is not None:
            self.err_var.set("")
            self.on_login(user, role)
        else:
            self.err_var.set("✗  Invalid username or password.")
            self.password_var.set("")
            self.password_entry.focus_set()


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 2 — DASHBOARD  (dropdown selector)
# ══════════════════════════════════════════════════════════════════════════════
class DashboardScreen(tk.Frame):
    def __init__(self, master, username, role, on_select, on_logout, on_manage_accounts):
        super().__init__(master, bg=BG)
        self.username  = username
        self.role      = role
        self.on_select = on_select
        self.on_logout = on_logout
        self.on_manage_accounts = on_manage_accounts
        self._build()

    def _build(self):
        # Header
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        tk.Label(header, text="⬡  Main Menu",
                 font=("Segoe UI", 14, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=20, pady=14)

        # User badge + logout
        right_hdr = tk.Frame(header, bg=HEADER_BG)
        right_hdr.pack(side="right", padx=16)

        role_color = GOLD if self.role == "admin" else ACCENT
        role_icon  = "👑" if self.role == "admin" else "👤"
        tk.Label(right_hdr, text=f"{role_icon}  {self.username}",
                 font=("Segoe UI", 9, "bold"), bg=HEADER_BG, fg=role_color
                 ).pack(side="left", padx=(0, 12))

        if self.role == "admin":
            tk.Button(right_hdr, text="⚙ Manage Accounts",
                      font=("Segoe UI", 8, "bold"),
                      bg=ACCENT2, fg="white", relief="flat",
                      padx=10, pady=4, cursor="hand2",
                      command=self.on_manage_accounts
                      ).pack(side="left", padx=(0, 8))

        tk.Button(right_hdr, text="Log Out",
                  font=("Segoe UI", 8, "bold"),
                  bg=DANGER, fg="white", relief="flat",
                  padx=12, pady=4, cursor="hand2",
                  command=self.on_logout
                  ).pack(side="left")

        # Thin accent stripe
        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")

        # ── Center content ────────────────────────────────────────────────────
        outer = tk.Frame(self, bg=BG)
        outer.pack(expand=True)

        card = tk.Frame(outer, bg=CARD, padx=56, pady=52,
                        highlightthickness=1, highlightbackground=BORDER)
        card.pack(padx=20, pady=20)

        tk.Label(card, text="Select Your Platform",
                 font=("Segoe UI", 18, "bold"), bg=CARD, fg=TEXT).pack()
        tk.Label(card, text="Choose the client account you want to work with",
                 font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(pady=(4, 32))

        # Dropdown label
        tk.Label(card, text="CLIENT ACCOUNT", font=("Segoe UI", 8, "bold"),
                 bg=CARD, fg=SUBTEXT).pack(anchor="w")

        # Custom styled combobox container
        combo_wrapper = tk.Frame(card, bg=BORDER, padx=1, pady=1)
        combo_wrapper.pack(fill="x", pady=(4, 24))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TCombobox",
                        fieldbackground=CARD,
                        background=CARD,
                        foreground=TEXT,
                        arrowcolor=ACCENT,
                        bordercolor=BORDER,
                        relief="flat",
                        padding=10,
                        font=("Segoe UI", 11))
        style.map("Dark.TCombobox",
                  fieldbackground=[("readonly", CARD)],
                  selectbackground=[("readonly", CARD)],
                  selectforeground=[("readonly", TEXT)])

        self.selected = tk.StringVar(value="— Select an account —")
        combo = ttk.Combobox(
            combo_wrapper,
            textvariable=self.selected,
            values=["Orico Auto Loan", "Bank of Makati"],
            state="readonly",
            style="Dark.TCombobox",
            font=("Segoe UI", 11),
        )
        combo.pack(fill="x")

        # Error label
        self.err_var = tk.StringVar()
        tk.Label(card, textvariable=self.err_var,
                 font=("Segoe UI", 9), bg=CARD, fg=DANGER).pack(pady=(0, 8))

        # Proceed button
        tk.Button(card, text="Proceed  →",
                  font=("Segoe UI", 11, "bold"),
                  bg=ACCENT, fg="white", relief="flat",
                  pady=11, cursor="hand2",
                  activebackground="#3a7be8",
                  command=self._proceed
                  ).pack(fill="x")

        # Quick-info chips
        chips_frame = tk.Frame(card, bg=CARD)
        chips_frame.pack(pady=(28, 0))
        for label, color in [("Orico Auto Loan", ACCENT), ("Bank of Makati", ACCENT2)]:
            chip = tk.Frame(chips_frame, bg=color, padx=14, pady=5)
            chip.pack(side="left", padx=6)
            tk.Label(chip, text=label, font=("Segoe UI", 9, "bold"),
                     bg=color, fg="white").pack()

        # Footer
        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)

    def _proceed(self):
        val = self.selected.get()
        if val.startswith("—"):
            self.err_var.set("⚠  Please select an account first.")
            return
        self.err_var.set("")
        self.on_select(val)


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 3 — ORICO TOOLS  (clickable tool cards)
# ══════════════════════════════════════════════════════════════════════════════
class OricoToolsScreen(tk.Frame):
    def __init__(self, master, username, role="user", on_tool=None, on_back=None):
        super().__init__(master, bg=BG)
        self.username = username
        self.role     = role
        self.on_tool  = on_tool
        self.on_back  = on_back
        self._build()

    def _build(self):
        # Header
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  ⬡  Orico Auto Loan",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)

        right_hdr = tk.Frame(header, bg=HEADER_BG)
        right_hdr.pack(side="right", padx=16)
        role_colors = {"admin": DANGER, "supervisor": WARNING, "user": SUCCESS}
        role_color  = role_colors.get(self.role, ACCENT)
        tk.Label(right_hdr, text=f"👤  {self.username}",
                 font=("Segoe UI", 9, "bold"), bg=HEADER_BG, fg=ACCENT).pack()
        tk.Label(right_hdr, text=f"● {self.role.upper()}",
                 font=("Segoe UI", 7, "bold"), bg=HEADER_BG, fg=role_color).pack()

        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")

        # Subtitle
        sub = tk.Frame(self, bg=BG)
        sub.pack(pady=(32, 8))
        tk.Label(sub, text="Tools & Utilities",
                 font=("Segoe UI", 18, "bold"), bg=BG, fg=TEXT).pack()
        tk.Label(sub, text="Select a tool to get started",
                 font=("Segoe UI", 9), bg=BG, fg=SUBTEXT).pack(pady=4)

        # Tool cards grid
        grid = tk.Frame(self, bg=BG)
        grid.pack(expand=True)

        tools = [
            {
                "name":    "DRR Cleaner",
                "icon":    "🧹",
                "desc":    "Clean & process Daily\nRemittance Reports",
                "color":   ACCENT,
                "tag":     "drr_cleaner",
                "badge":   "Excel",
            },
            {
                "name":    "Autostat",
                "icon":    "📥",
                "desc":    "Import Batch",
                "color":   ACCENT2,
                "tag":     "autostat",
                "badge":   "New",
            },
            {
                "name":    "Analytics",
                "icon":    "📊",
                "desc":    "Analytical\nReports",
                "color":   "#0E9F6E",
                "tag":     "analytical_reports",
                "badge":   "New",
            },
        ]

        for i, tool in enumerate(tools):
            col_frame = tk.Frame(grid, bg=BG)
            col_frame.grid(row=0, column=i, padx=14)
            self._tool_card(col_frame, tool)

        # Footer
        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)

    def _tool_card(self, parent, tool):
        is_active = tool["tag"] is not None
        border_color = tool["color"] if is_active else BORDER

        card = tk.Frame(parent, bg=CARD, width=200, height=220,
                        highlightthickness=2,
                        highlightbackground=border_color,
                        cursor="hand2" if is_active else "arrow")
        card.pack()
        card.pack_propagate(False)

        inner = tk.Frame(card, bg=CARD, padx=20, pady=20)
        inner.pack(fill="both", expand=True)

        # Badge
        badge_color = tool["color"] if is_active else "#3D4153"
        badge = tk.Frame(inner, bg=badge_color, padx=8, pady=2)
        badge.pack(anchor="e")
        tk.Label(badge, text=tool["badge"], font=("Segoe UI", 7, "bold"),
                 bg=badge_color, fg="white").pack()

        # Icon
        tk.Label(inner, text=tool["icon"], font=("Segoe UI", 34),
                 bg=CARD, fg=tool["color"] if is_active else SUBTEXT).pack(pady=(8, 8))

        # Name
        tk.Label(inner, text=tool["name"], font=("Segoe UI", 12, "bold"),
                 bg=CARD, fg=TEXT if is_active else SUBTEXT).pack()

        # Description
        tk.Label(inner, text=tool["desc"], font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT, justify="center").pack(pady=(4, 0))

        # Hover effects for active cards
        if is_active:
            tag = tool["tag"]

            def on_enter(e, c=card, col=tool["color"]):
                c.config(highlightbackground=col, bg="#22263A")
                for w in c.winfo_children():
                    _set_bg_recursive(w, "#22263A")

            def on_leave(e, c=card, col=tool["color"]):
                c.config(highlightbackground=col, bg=CARD)
                for w in c.winfo_children():
                    _set_bg_recursive(w, CARD)

            def on_click(e, t=tag):
                self.on_tool(t)

            for widget in [card, inner] + _all_children(inner):
                widget.bind("<Enter>", on_enter)
                widget.bind("<Leave>", on_leave)
                widget.bind("<Button-1>", on_click)


def _all_children(widget):
    result = []
    for child in widget.winfo_children():
        result.append(child)
        result.extend(_all_children(child))
    return result


def _set_bg_recursive(widget, color):
    try:
        widget.config(bg=color)
    except Exception:
        pass
    for child in widget.winfo_children():
        _set_bg_recursive(child, color)


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN 4 — DRR CLEANER APP
# ══════════════════════════════════════════════════════════════════════════════
class DRRCleanerScreen(tk.Frame):
    PAGE_SIZE = 2000

    def __init__(self, master, username, role="user", on_back=None):
        super().__init__(master, bg=BG)
        self.username           = username
        self.role               = role
        self.on_back            = on_back
        self.file_path          = None
        self.output_bytes       = None
        self.cleaned_df         = None
        self.removed_df         = None
        self.removed_reason_df  = None
        self.unique_df         = None
        self.verify_df         = None   # summary DataFrame for Verify tab
        self._raw_filepath     = None   # path of raw file for comparison
        self.remarks_df         = None
        self._current_df        = None
        self._current_page      = 0
        self._iid_to_row        = {}
        self._build_ui()

    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  🧹  DRR Cleaner  —  Orico Auto Loan",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)
        tk.Label(left_hdr, text="Drop · Preview · Verify · Download",
                 font=("Segoe UI", 8), bg=HEADER_BG, fg=SUBTEXT).pack(side="left", padx=4)

        right_hdr = tk.Frame(header, bg=HEADER_BG)
        right_hdr.pack(side="right", padx=16)
        role_colors = {"admin": DANGER, "supervisor": WARNING, "user": SUCCESS}
        role_color  = role_colors.get(self.role, ACCENT)
        tk.Label(right_hdr, text=f"👤  {self.username}",
                 font=("Segoe UI", 9, "bold"), bg=HEADER_BG, fg=ACCENT).pack()
        tk.Label(right_hdr, text=f"● {self.role.upper()}",
                 font=("Segoe UI", 7, "bold"), bg=HEADER_BG, fg=role_color).pack()

        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")

        # ── Body ──────────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # ── LEFT PANEL ────────────────────────────────────────────────────────
        left = tk.Frame(body, bg=BG, width=300)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        # ── Drop Zone ─────────────────────────────────────────────────────────
        self._drop_frame = tk.Frame(left, bg=CARD, highlightthickness=2,
                                    highlightbackground=BORDER)
        self._drop_frame.pack(fill="x", pady=(0, 10))

        # ── State A: empty drop prompt ────────────────────────────────────────
        self._drop_empty = tk.Frame(self._drop_frame, bg=CARD)
        self._drop_empty.pack(fill="both", padx=20, pady=20)

        self._drop_icon  = tk.Label(self._drop_empty, text="📂", font=("Segoe UI", 26),
                                    bg=CARD, fg=ACCENT)
        self._drop_icon.pack()
        self._drop_title = tk.Label(self._drop_empty, text="Drop Excel file here",
                                    font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT)
        self._drop_title.pack(pady=(6, 2))
        tk.Label(self._drop_empty, text=".xlsx  ·  .xlsm  ·  .xls",
                 font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack()
        tk.Label(self._drop_empty, text="or", font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT).pack(pady=(8, 4))
        tk.Button(self._drop_empty, text="Browse File", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat", padx=16, pady=6,
                  cursor="hand2", command=self._browse).pack()

        # ── State B: file loaded display (hidden until a file is picked) ──────
        self._drop_loaded = tk.Frame(self._drop_frame, bg=CARD)
        # NOT packed yet — shown in _load_file

        # Top: green checkmark banner
        loaded_top = tk.Frame(self._drop_loaded, bg="#1A3A2A")
        loaded_top.pack(fill="x")
        tk.Label(loaded_top, text="✅  File Loaded", font=("Segoe UI", 9, "bold"),
                 bg="#1A3A2A", fg=SUCCESS).pack(side="left", padx=12, pady=6)
        tk.Button(loaded_top, text="✕  Clear", font=("Segoe UI", 8),
                  bg="#1A3A2A", fg=SUBTEXT, relief="flat",
                  cursor="hand2", command=self._clear_file
                  ).pack(side="right", padx=8)

        # File icon + details
        loaded_body = tk.Frame(self._drop_loaded, bg=CARD)
        loaded_body.pack(fill="x", padx=14, pady=12)

        tk.Label(loaded_body, text="📄", font=("Segoe UI", 22),
                 bg=CARD, fg=ACCENT).pack(side="left", padx=(0, 10))

        loaded_text = tk.Frame(loaded_body, bg=CARD)
        loaded_text.pack(side="left", fill="x", expand=True)
        self._fi_name = tk.Label(loaded_text, text="", font=("Segoe UI", 9, "bold"),
                                  bg=CARD, fg=TEXT, anchor="w", wraplength=180, justify="left")
        self._fi_name.pack(fill="x")
        self._fi_meta = tk.Label(loaded_text, text="", font=("Segoe UI", 7),
                                  bg=CARD, fg=SUBTEXT, anchor="w")
        self._fi_meta.pack(fill="x", pady=(2, 0))

        # Change file button
        tk.Button(self._drop_loaded, text="🔄  Change File",
                  font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT,
                  relief="flat", cursor="hand2",
                  command=self._browse).pack(pady=(0, 10))

        # Register DnD on the frame AND every child widget so the whole zone is a target
        self._dnd_enabled = False
        self._dnd_widgets = []
        try:
            for widget in [self._drop_frame, self._drop_empty, self._drop_loaded,
                           self._drop_icon, self._drop_title, loaded_top, loaded_body]:
                widget.drop_target_register('DND_Files')
                widget.dnd_bind('<<Drop>>',      self._on_drop)
                widget.dnd_bind('<<DragEnter>>', self._on_drag_enter)
                widget.dnd_bind('<<DragLeave>>', self._on_drag_leave)
                self._dnd_widgets.append(widget)
            self._dnd_enabled = True
        except Exception:
            pass

        # Legacy label kept for compatibility (hidden)
        self.file_label = tk.Label(left, text="", font=("Segoe UI", 8),
                                   bg=BG, fg=SUBTEXT, wraplength=270, justify="left")


        tk.Label(left, text="SUMMARY", font=("Segoe UI", 8, "bold"),
                 bg=BG, fg=SUBTEXT).pack(anchor="w", pady=(0, 6))

        self.stat_vars = {}
        for key, label, color, icon in [
            ("total",    "Total Rows",    TEXT,    "📋"),
            ("retained", "Rows Retained", SUCCESS, "✅"),
            ("removed",  "Rows Removed",  DANGER,  "🗑"),
            ("srp",      "Remarks Fixed", WARNING, "✏️"),
        ]:
            c = tk.Frame(left, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
            c.pack(fill="x", pady=3)
            inn = tk.Frame(c, bg=CARD)
            inn.pack(fill="x", padx=12, pady=8)
            tk.Label(inn, text=icon, font=("Segoe UI", 11), bg=CARD, fg=color).pack(side="left")
            tk.Label(inn, text=label, font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(side="left", padx=8)
            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(inn, textvariable=var, font=("Segoe UI", 11, "bold"),
                     bg=CARD, fg=color).pack(side="right")

        # Supervisor: can view + download but NOT upload/clean
        # User: can upload, clean, AND download
        can_upload        = self.role in ("admin", "user")
        can_download       = self.role in ("admin", "supervisor", "user")
        self._can_upload   = can_upload
        self._can_download = can_download

        self.process_btn = tk.Button(left, text="▶  Run Cleaner",
                                     font=("Segoe UI", 10, "bold"), bg=ACCENT, fg="white",
                                     relief="flat", pady=10, cursor="hand2",
                                     state="disabled" if can_upload else "disabled",
                                     command=self._run_process)
        self.process_btn.pack(fill="x", pady=(14, 6))
        if not can_upload:
            self.process_btn.config(state="disabled", bg=CARD, fg=SUBTEXT,
                                    text="▶  Run Cleaner  (No Access)")

        self.download_btn = tk.Button(left, text="⬇  Download Output",
                                      font=("Segoe UI", 9, "bold"),
                                      bg=SUCCESS if can_download else CARD,
                                      fg="white" if can_download else SUBTEXT,
                                      relief="flat", pady=10, cursor="hand2",
                                      state="disabled", command=self._download)
        self.download_btn.pack(fill="x")
        if not can_download:
            self.download_btn.config(text="⬇  Download  (No Access)")

        style = ttk.Style()
        style.configure("TProgressbar", troughcolor=CARD, background=ACCENT, thickness=4)
        self.progress = ttk.Progressbar(left, mode="indeterminate")
        self.progress.pack(fill="x", pady=(10, 0))

        # ── RIGHT PANEL ───────────────────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        tab_bar = tk.Frame(right, bg=BG)
        tab_bar.pack(fill="x", pady=(0, 8))

        self.active_tab = tk.StringVar(value="cleaned")
        self.tab_btns   = {}
        for tab_id, tab_label, tab_color in [
            ("cleaned",        "✅ Cleaned Rows",    ACCENT),
            ("verify",         "🔍 Verify",          CARD),
            ("removed",        "🗑 Removed Rows",    CARD),
            ("removed_reason", "📋 Removed Status",  CARD),
            ("remarks",        "✏️ Remarks Changes", CARD),
        ]:
            btn = tk.Button(tab_bar, text=tab_label,
                            font=("Segoe UI", 9, "bold"),
                            bg=tab_color, fg="white", relief="flat",
                            padx=14, pady=7, cursor="hand2",
                            command=lambda t=tab_id: self._switch_tab(t))
            btn.pack(side="left", padx=(0, 6))
            self.tab_btns[tab_id] = btn

        # ── Search bar ────────────────────────────────────────────────────────
        search_frame = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        search_frame.pack(fill="x", pady=(0, 4))
        tk.Label(search_frame, text="🔍", bg=CARD, fg=SUBTEXT, font=("Segoe UI", 10)).pack(side="left", padx=8)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                 ).pack(side="left", fill="x", expand=True, pady=8, padx=4)
        tk.Label(search_frame, text="Search all columns…", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        # ── Column Dropdown Filter bar ─────────────────────────────────────────
        filter_outer = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        filter_outer.pack(fill="x", pady=(0, 8))
        tk.Label(filter_outer, text="▼ Column Filter:", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=8, pady=6)
        self._filter_col_var = tk.StringVar(value="— Select Column —")
        self._filter_val_var = tk.StringVar(value="— Select Value —")
        self._filter_col_master = []   # full list of column names, for typeahead filtering
        self._filter_val_master = []   # full list of unique values, for typeahead filtering
        self._filter_col_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_col_var,
                                             width=24, font=("Segoe UI", 9))
        self._filter_col_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_val_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_val_var,
                                             width=28, font=("Segoe UI", 9))
        self._filter_val_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_col_cb.bind("<<ComboboxSelected>>", self._on_filter_col_change)
        self._filter_val_cb.bind("<<ComboboxSelected>>", self._on_filter_apply)
        self._filter_col_cb.bind("<KeyRelease>", self._on_filter_col_type)
        self._filter_val_cb.bind("<KeyRelease>", self._on_filter_val_type)
        tk.Button(filter_outer, text="✕ Clear Filter", font=("Segoe UI", 8, "bold"),
                  bg=DANGER, fg="white", relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self._clear_filter).pack(side="left", padx=(0, 6))
        self._filter_active = False

        table_frame = tk.Frame(right, bg=BG)
        table_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style.configure("Custom.Treeview", background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28, font=("Segoe UI", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading", background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT2)], foreground=[("selected", "white")])

        self.tree = ttk.Treeview(table_frame, style="Custom.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                  show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self._on_cell_click)
        self.tree.bind("<ButtonRelease-1>",  self._on_tree_click)
        self.tree.bind("<Double-1>",         self._on_double_click)
        self.tree.bind("<Return>",           self._on_double_click)

        self.row_count_label = tk.Label(right, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.row_count_label.pack(anchor="e", pady=(4, 0))

        # ── Excel-style formula bar ───────────────────────────────────────────
        fbar = tk.Frame(right, bg="#1C1F2E",
                        highlightthickness=1, highlightbackground=BORDER)
        fbar.pack(fill="x", pady=(4, 2))

        # fx badge
        tk.Label(fbar, text=" fx ", font=("Segoe UI", 9, "bold"),
                 bg="#2E7D32", fg="white", padx=6, pady=4
                 ).pack(side="left")

        # Column name box
        self._fbar_col_var = tk.StringVar(value="")
        col_box = tk.Entry(fbar, textvariable=self._fbar_col_var,
                           font=("Segoe UI", 9, "bold"),
                           bg="#252840", fg=ACCENT,
                           relief="flat", bd=0,
                           width=22, justify="center",
                           insertbackground=ACCENT,
                           readonlybackground="#252840",
                           state="readonly")
        col_box.pack(side="left", padx=(1, 0), ipady=4)

        # Divider
        tk.Frame(fbar, bg=BORDER, width=1).pack(side="left", fill="y", pady=2)

        # Value text — full, selectable, scrollable
        self._fbar_val = tk.Text(fbar, font=("Segoe UI", 9),
                                 bg="#1C1F2E", fg=TEXT,
                                 relief="flat", bd=0,
                                 height=1, wrap="none",
                                 cursor="xterm",
                                 selectbackground=ACCENT,
                                 selectforeground="white",
                                 insertbackground=TEXT,
                                 state="disabled")
        fbar_xsb = ttk.Scrollbar(fbar, orient="horizontal",
                                  command=self._fbar_val.xview)
        self._fbar_val.configure(xscrollcommand=fbar_xsb.set)
        self._fbar_val.pack(side="left", fill="x", expand=True, padx=(6, 0), pady=2)
        # Copy button on right
        tk.Button(fbar, text="⎘", font=("Segoe UI", 10),
                  bg="#1C1F2E", fg=SUBTEXT, relief="flat",
                  cursor="hand2", padx=6,
                  command=self._fbar_copy
                  ).pack(side="right", padx=4)

        nav_frame = tk.Frame(right, bg=BG)
        nav_frame.pack(fill="x", pady=(4, 0))
        self.prev_btn = tk.Button(nav_frame, text="◀  Prev", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._prev_page)
        self.prev_btn.pack(side="left", padx=(0, 6))
        self.page_label = tk.Label(nav_frame, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.page_label.pack(side="left")
        self.next_btn = tk.Button(nav_frame, text="Next  ▶", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._next_page)
        self.next_btn.pack(side="left", padx=6)

        # Status bar
        self.status_var = tk.StringVar(value="Ready — drop or browse an Excel file to begin.")
        tk.Label(self, textvariable=self.status_var, font=("Segoe UI", 8),
                 bg=HEADER_BG, fg=SUBTEXT, anchor="w", padx=16, pady=6).pack(fill="x", side="bottom")

        # Footer
        watermark = tk.Frame(self, bg=DARK)
        watermark.pack(fill="x", side="bottom")
        tk.Label(watermark, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=6)
        tk.Label(watermark, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=6)

    # ── FILE LOAD ─────────────────────────────────────────────────────────────
    def _on_drag_enter(self, event):
        self._drop_frame.config(highlightbackground=ACCENT, highlightthickness=2,
                                bg="#1A2540")
        self._drop_icon.config(bg="#1A2540", fg=ACCENT, text="⬇️")
        self._drop_title.config(bg="#1A2540", fg=ACCENT, text="Release to load file")
        for w in self._drop_frame.winfo_children():
            try: w.config(bg="#1A2540")
            except Exception: pass

    def _on_drag_leave(self, event):
        self._drop_frame.config(highlightbackground=BORDER, bg=CARD)
        self._drop_icon.config(bg=CARD, fg=ACCENT, text="📂")
        self._drop_title.config(bg=CARD, fg=TEXT, text="Drop Excel file here")
        for w in self._drop_frame.winfo_children():
            try: w.config(bg=CARD)
            except Exception: pass

    def _on_drop(self, event):
        self._on_drag_leave(None)
        raw = event.data.strip()
        # tkinterdnd2 may wrap paths with braces when they contain spaces
        paths = []
        if raw.startswith("{"):
            import re
            paths = re.findall(r'\{([^}]+)\}', raw)
        if not paths:
            paths = [raw]
        self._load_file(paths[0])

    def _browse(self):
        path = filedialog.askopenfilename(title="Select Excel file",
                                          filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self._load_file(path)

    def _clear_file(self):
        self.file_path = None
        self._drop_loaded.pack_forget()
        self._drop_empty.pack(fill="both", padx=20, pady=20)
        self._drop_frame.config(highlightbackground=BORDER)
        self.process_btn.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set("Ready — drop or browse an Excel file to begin.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        self.file_path = path
        self._raw_filepath = path  # keep raw path for Verify tab
        fname = os.path.basename(path)
        fsize = os.path.getsize(path)
        fsize_str = (f"{fsize / 1024:.1f} KB" if fsize < 1_048_576
                     else f"{fsize / 1_048_576:.2f} MB")
        import time
        mtime     = os.path.getmtime(path)
        mtime_str = time.strftime("%b %d, %Y  %I:%M %p", time.localtime(mtime))

        # Swap drop zone: hide empty state, show loaded state
        self._drop_empty.pack_forget()
        self._fi_name.config(text=fname)
        self._fi_meta.config(text=f"{fsize_str}  ·  {mtime_str}")
        self._drop_loaded.pack(fill="both")
        self._drop_frame.config(highlightbackground=SUCCESS)

        self.file_label.config(text="")
        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set(f"Loaded: {fname}  —  Click ▶ Run Cleaner to process.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    # ── PROCESSING ────────────────────────────────────────────────────────────
    def _run_process(self):
        if self.role not in ("admin", "user"):
            messagebox.showwarning("Access Denied", "Your role does not have upload/clean access.")
            return
        if not self.file_path:
            return
        self.process_btn.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set("Processing…")
        self.progress.start(10)
        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        try:
            result = process_file(self.file_path)
            self.after(0, lambda: self._on_done(*result))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _on_done(self, cleaned, removed, remarks, stats, out_bytes, col_e, dialed_col, unique_df=None):
        self.progress.stop()
        self.cleaned_df        = cleaned
        self.removed_df        = removed
        self.unique_df         = unique_df if unique_df is not None else cleaned.copy()
        self.verify_df         = None   # reset; built on demand when tab is clicked
        if "Removed Reason" in removed.columns:
            status_col = [c for c in removed.columns if c != "Removed Reason"]
            mask = removed["Removed Reason"].str.startswith("Status: ")
            self.removed_reason_df = removed[mask][["Removed Reason"] + status_col].copy()
            # Rows removed purely because their Status matched a "removable"
            # status belong only in the "Removed Status" tab — keep them out
            # of the general "Removed Rows" tab so they aren't shown twice.
            self.removed_df = removed[~mask].copy()
        else:
            self.removed_reason_df = removed.copy()
            self.removed_df        = removed
        self.remarks_df        = remarks
        self.output_bytes      = out_bytes

        self.stat_vars["total"].set(str(stats["total"]))
        self.stat_vars["retained"].set(str(stats["retained"]))
        self.stat_vars["removed"].set(str(stats["removed"]))
        self.stat_vars["srp"].set(str(stats["srp_changed"]))

        self.download_btn.config(state="normal" if self._can_download else "disabled")
        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.status_var.set(
            f"Done  ·  {stats['retained']} retained  ·  {stats['removed']} removed  ·  "
            f"{stats['srp_changed']} remarks → SRP"
        )
        self._switch_tab("cleaned")
        show_toast(self.master, "Processing complete!", color=SUCCESS)

    def _on_error(self, msg):
        self.progress.stop()
        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.status_var.set(f"Error: {msg}")
        messagebox.showerror("Processing Error", msg)

    # ── TABS & TABLE ──────────────────────────────────────────────────────────
    def _switch_tab(self, tab_id):
        self.active_tab.set(tab_id)
        colors = {k: CARD for k in self.tab_btns}
        if tab_id in ("remarks", "removed_reason"):
            colors[tab_id] = ACCENT2
        elif tab_id == "verify":
            colors[tab_id] = WARNING
        else:
            colors[tab_id] = ACCENT
        for tid, btn in self.tab_btns.items():
            btn.config(bg=colors[tid])
        # Hide verify panel if switching away from it
        if hasattr(self, "_verify_panel") and self._verify_panel.winfo_exists():
            self._verify_panel.destroy()

        # Verify tab: build on demand when clicked
        if tab_id == "verify":
            self._show_verify_panel()
            return

        df_map = {
            "cleaned":        self.cleaned_df,
            "removed":        self.removed_df,
            "removed_reason": self.removed_reason_df,
            "remarks":        self.remarks_df,
        }
        df = df_map.get(tab_id)
        empty_msgs = {
            "remarks":        "No remark changes found.",
            "removed_reason": "No removed rows found.",
            "removed":        "No removed rows.",
            "cleaned":        "No retained rows.",
        }
        # Refresh column filter dropdown for the new tab
        if df is not None and not (hasattr(df, "empty") and df.empty):
            self._filter_col_master = list(df.columns)
            self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
            self._filter_col_var.set("— Select Column —")
            self._filter_val_var.set("— Select Value —")
            self._filter_val_master = []
            self._filter_val_cb["values"] = []
            self._filter_active = False
            self._populate_table(df)
        else:
            self._clear_table()
            self.row_count_label.config(text=empty_msgs.get(tab_id, "No rows."))

    def _populate_table(self, df, filter_text=""):
        self._clear_table()
        if df is None or df.empty:
            self.row_count_label.config(text="No rows to display.")
            return
        cols = list(df.columns)
        self.tree["columns"] = cols
        for col in cols:
            sample    = df[col].astype(str)
            max_len   = max(len(str(col)), int(sample.str.len().quantile(0.90)) if len(df) > 0 else 10)
            col_width = max(100, min(max_len * 8, 280))
            self.tree.heading(col, text=col, anchor="w",
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=col_width, minwidth=60, anchor="w", stretch=False)
        filt = filter_text.lower().strip()
        if filt:
            combined = df.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            filtered = df[combined.str.contains(filt, regex=False)].reset_index(drop=True)
        else:
            filtered = df.reset_index(drop=True)
        self._current_df   = filtered
        self._current_page = 0
        self._render_page()

    def _render_page(self):
        self.tree.delete(*self.tree.get_children())
        self._iid_to_row = {}
        df    = self._current_df
        page  = self._current_page
        start = page * self.PAGE_SIZE
        end   = min(start + self.PAGE_SIZE, len(df))
        str_df = df.iloc[start:end].fillna("").astype(str)
        disp_df = str_df.copy()
        if DATE_COL_NAME in disp_df.columns:
            disp_df[DATE_COL_NAME] = disp_df[DATE_COL_NAME].apply(_display_date_ddmmyyyy)
        if TIME_COL_NAME in disp_df.columns:
            disp_df[TIME_COL_NAME] = disp_df[TIME_COL_NAME].apply(_display_time_only)
        for i, vals in enumerate(disp_df.values.tolist()):
            tag = "even" if i % 2 == 0 else "odd"
            iid = self.tree.insert("", "end", values=vals, tags=(tag,))
            self._iid_to_row[iid] = start + i
        self.tree.tag_configure("even", background=CARD)
        self.tree.tag_configure("odd",  background="#14172A")
        total_pages = max(1, (len(df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        self.row_count_label.config(
            text=(f"Page {page + 1} of {total_pages}  ·  "
                  f"Showing rows {start + 1}–{end} of {len(df)}  ·  "
                  f"{len(df.columns)} columns  ·  ← → to navigate pages")
        )
        if hasattr(self, "prev_btn"):
            self.prev_btn.config(state="normal" if page > 0 else "disabled")
            self.next_btn.config(state="normal" if page < total_pages - 1 else "disabled")
            self.page_label.config(text=f"Page {page + 1} / {total_pages}")

    # ── VERIFY TAB ───────────────────────────────────────────────────────────
    def _show_verify_panel(self):
        """Build and display the Verify summary panel in the table area."""
        # Highlight the verify tab button
        for tid, btn in self.tab_btns.items():
            btn.config(bg=WARNING if tid == "verify" else CARD)
        self.tab_btns["verify"].config(bg=WARNING)
        self.active_tab.set("verify")

        # Clear existing table content
        self._clear_table()

        # Destroy any previous verify panel
        if hasattr(self, "_verify_panel") and self._verify_panel.winfo_exists():
            self._verify_panel.destroy()

        # Build panel inside the tree's parent frame
        panel = tk.Frame(self.tree.master, bg=CARD,
                         highlightthickness=1, highlightbackground=WARNING)
        panel.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._verify_panel = panel

        tk.Label(panel, text="🔍  Verify: Raw vs Cleaned",
                 font=("Segoe UI", 13, "bold"), bg=CARD, fg=TEXT).pack(pady=(24, 4))
        tk.Label(panel, text="Compare every row in the cleaned output against the original raw file.",
                 font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(pady=(0, 20))

        # ── Summary cards area ────────────────────────────────────────────
        cards = tk.Frame(panel, bg=CARD)
        cards.pack(pady=(0, 20))

        def _stat_card(parent, label, var, color):
            c = tk.Frame(parent, bg=DARK, padx=28, pady=18,
                         highlightthickness=1, highlightbackground=color)
            c.pack(side="left", padx=12)
            tk.Label(c, textvariable=var, font=("Segoe UI", 28, "bold"),
                     bg=DARK, fg=color).pack()
            tk.Label(c, text=label, font=("Segoe UI", 9), bg=DARK, fg=SUBTEXT).pack()
            return c

        self._v_total    = tk.StringVar(value="—")
        self._v_matched  = tk.StringVar(value="—")
        self._v_unmatched= tk.StringVar(value="—")
        self._v_status   = tk.StringVar(value="Load the raw file to begin verification.")

        _stat_card(cards, "Cleaned Rows",  self._v_total,     TEXT)
        _stat_card(cards, "✔ Matched",     self._v_matched,   SUCCESS)
        _stat_card(cards, "✗ Unmatched",   self._v_unmatched, DANGER)

        tk.Label(panel, textvariable=self._v_status,
                 font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT,
                 wraplength=600).pack(pady=(0, 16))

        # ── Unmatched detail list (shown after comparison) ────────────────
        self._v_detail_frame = tk.Frame(panel, bg=CARD)
        self._v_detail_frame.pack(fill="x", padx=40, pady=(0, 12))

        # ── Load raw file button ───────────────────────────────────────────
        btn_row = tk.Frame(panel, bg=CARD)
        btn_row.pack()

        tk.Button(btn_row, text="📂  Load Raw File & Compare",
                  font=("Segoe UI", 10, "bold"),
                  bg=WARNING, fg="white", relief="flat",
                  padx=20, pady=10, cursor="hand2",
                  command=self._run_verify).pack(side="left", padx=8)

        tk.Button(btn_row, text="⬇  Download Comparison",
                  font=("Segoe UI", 10, "bold"),
                  bg=ACCENT, fg="white", relief="flat",
                  padx=20, pady=10, cursor="hand2",
                  command=self._download_verify_report).pack(side="left", padx=8)

        tk.Button(btn_row, text="✕  Close",
                  font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat",
                  padx=12, pady=10, cursor="hand2",
                  command=lambda: (panel.destroy(), self._switch_tab("cleaned"))
                  ).pack(side="left", padx=8)

        # Auto-run if raw file path is already known
        if self._raw_filepath and self.unique_df is not None:
            self.after(100, self._run_verify)

    def _run_verify(self):
        """Compare unique_df IDs against the raw file and update summary cards."""
        if self.unique_df is None or self.unique_df.empty:
            self._v_status.set("⚠ No cleaned data. Run the cleaner first.")
            return

        # Ask for file if not already set
        path = self._raw_filepath
        if not path or not os.path.exists(path):
            path = filedialog.askopenfilename(
                title="Select the original raw file",
                filetypes=[("Excel files", "*.xlsx *.xls")]
            )
            if not path:
                return
            self._raw_filepath = path

        self._v_status.set("⏳ Reading raw file…")
        self.update_idletasks()

        try:
            raw_df = pd.read_excel(path, dtype=str)

            # Build the same Unique ID from the raw file
            raw_acct_col   = raw_df.columns[COL_E_INDEX] if COL_E_INDEX < len(raw_df.columns) else None
            raw_date_col   = DATE_COL_NAME if DATE_COL_NAME in raw_df.columns else None
            raw_time_col   = TIME_COL_NAME if TIME_COL_NAME in raw_df.columns else None
            raw_debtor_col = next((c for c in raw_df.columns if str(c).strip().lower() == "debtor"), None)

            def _s(df, col):
                if col and col in df.columns:
                    return df[col].fillna("").astype(str).str.strip()
                return pd.Series([""] * len(df), index=df.index)

            raw_time_s = _s(raw_df, raw_time_col).str.split("  ").str[-1].str.strip()
            # Format raw date to match cleaned (mm/dd/yyyy)
            raw_date_s = pd.to_datetime(_s(raw_df, raw_date_col), errors="coerce").dt.strftime("%m/%d/%Y").fillna("")

            raw_ids = (
                _s(raw_df, raw_acct_col)   + "|" +
                raw_date_s                  + "|" +
                raw_time_s                  + "|" +
                _s(raw_df, raw_debtor_col)
            )
            raw_id_set = set(raw_ids.tolist())

            # Compare against cleaned Unique IDs
            cleaned_ids  = self.unique_df["Unique ID"].astype(str).tolist()
            total        = len(cleaned_ids)
            matched_ids  = [uid for uid in cleaned_ids if uid in raw_id_set]
            unmatched_ids= [uid for uid in cleaned_ids if uid not in raw_id_set]

            matched   = len(matched_ids)
            unmatched = len(unmatched_ids)

            # Stash results so the Download button can export them
            self._v_raw_df        = raw_df
            self._v_matched_ids   = matched_ids
            self._v_unmatched_ids = unmatched_ids

            self._v_total.set(str(total))
            self._v_matched.set(str(matched))
            self._v_unmatched.set(str(unmatched))

            if unmatched == 0:
                self._v_status.set(f"✔ All {total} cleaned rows were found in the raw file. Data integrity confirmed.")
            else:
                self._v_status.set(
                    f"⚠ {unmatched} cleaned row(s) could NOT be matched back to the raw file. "
                    f"They may have been altered or are duplicates. See details below."
                )

            # Show unmatched detail list
            for w in self._v_detail_frame.winfo_children():
                w.destroy()

            if unmatched_ids:
                tk.Label(self._v_detail_frame,
                         text=f"Unmatched Unique IDs ({min(unmatched, 50)} shown):",
                         font=("Segoe UI", 9, "bold"), bg=CARD, fg=DANGER).pack(anchor="w", pady=(8, 4))
                detail_box = tk.Frame(self._v_detail_frame, bg=DARK,
                                      highlightthickness=1, highlightbackground=DANGER)
                detail_box.pack(fill="x")
                for uid in unmatched_ids[:50]:
                    tk.Label(detail_box, text=f"  ✗  {uid}",
                             font=("Segoe UI", 8), bg=DARK, fg=DANGER,
                             anchor="w").pack(fill="x", padx=8, pady=1)

        except Exception as e:
            self._v_status.set(f"✗ Error reading raw file: {e}")

    def _download_verify_report(self):
        """Export the Raw vs Cleaned comparison results to an Excel workbook."""
        if self.unique_df is None or self.unique_df.empty:
            messagebox.showwarning("No Data", "Run the cleaner first, then verify.", parent=self)
            return
        if not hasattr(self, "_v_matched_ids"):
            messagebox.showwarning("Not Verified Yet",
                                    "Click 'Load Raw File & Compare' first to generate the comparison.",
                                    parent=self)
            return
        if self.cleaned_df is None or self._v_raw_df is None:
            messagebox.showwarning("Missing Data",
                                    "Both the cleaned data and raw file are required to build the report.",
                                    parent=self)
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Comparison Report",
            defaultextension=".xlsx",
            initialfile=f"Verify_Comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            total     = len(self.unique_df)
            matched   = len(self._v_matched_ids)
            unmatched = len(self._v_unmatched_ids)

            summary_df = pd.DataFrame({
                "Metric": ["Raw File", "Cleaned Rows", "Matched", "Unmatched"],
                "Value": [
                    os.path.basename(self._raw_filepath) if self._raw_filepath else "",
                    total, matched, unmatched
                ]
            })

            matched_df   = self.unique_df[self.unique_df["Unique ID"].astype(str).isin(self._v_matched_ids)]
            unmatched_df = self.unique_df[self.unique_df["Unique ID"].astype(str).isin(self._v_unmatched_ids)]

            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                matched_df.to_excel(writer, sheet_name="Matched", index=False)
                unmatched_df.to_excel(writer, sheet_name="Unmatched", index=False)
                self._v_raw_df.to_excel(writer, sheet_name="Raw File", index=False)
                self.cleaned_df.to_excel(writer, sheet_name="Cleaned File", index=False)

            # Style header rows
            wb = openpyxl.load_workbook(save_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F8EF7", end_color="4F8EF7", fill_type="solid")
                for col_cells in ws.columns:
                    max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)
            wb.save(save_path)

            show_toast(self.master, f"Comparison report saved: {os.path.basename(save_path)}", color=SUCCESS)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save comparison report:\n{e}", parent=self)


    def _sort_col(self, col):
        tab = self.active_tab.get()
        df  = {
            "cleaned":        self.cleaned_df,
            "removed":        self.removed_df,
            "removed_reason": self.removed_reason_df,
            "remarks":        self.remarks_df,
        }.get(tab)
        if df is None:
            return
        asc = getattr(self, "_sort_asc", {})
        ascending = not asc.get(col, True)
        asc[col] = ascending
        self._sort_asc = asc
        try:
            df_sorted = df.sort_values(by=col, ascending=ascending,
                                        key=lambda x: x.astype(str).str.lower())
        except Exception:
            df_sorted = df
        if tab == "cleaned":
            self.cleaned_df = df_sorted
        elif tab == "removed":
            self.removed_df = df_sorted
        else:
            self.remarks_df = df_sorted
        self._populate_table(df_sorted, self.search_var.get())

    def _filter_typeahead(self, combo, master_list, var):
        """Narrow a combobox's dropdown list live to entries containing the typed text."""
        typed = var.get().strip().lower()
        if not typed or typed.startswith("—") or typed == "(all)":
            shown = master_list
        else:
            shown = [v for v in master_list if typed in str(v).lower()]
        combo["values"] = shown
        if typed and shown:
            try:
                combo.tk.call("ttk::combobox::Post", combo)
                combo.focus_set()
                combo.icursor("end")
            except Exception:
                pass

    def _on_filter_col_type(self, event=None):
        """As the user types, narrow the column dropdown; auto-apply on an exact match."""
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_col_cb, self._filter_col_master, self._filter_col_var)
        if self._filter_col_var.get() in self._filter_col_master:
            self._on_filter_col_change()

    def _on_filter_val_type(self, event=None):
        """As the user types, narrow the value dropdown and live-filter the table."""
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_val_cb, self._filter_val_master, self._filter_val_var)
        self._on_filter_apply()

    def _on_filter_col_change(self, _=None):
        """When a column is chosen, populate the value dropdown with unique values."""
        tab = self.active_tab.get()
        df  = {"cleaned": self.cleaned_df, "removed": self.removed_df,
               "removed_reason": self.removed_reason_df, "remarks": self.remarks_df}.get(tab)
        if df is None:
            return
        col = self._filter_col_var.get()
        if col not in df.columns:
            return
        self._filter_val_master = ["(All)"] + sorted(df[col].fillna("").astype(str).unique().tolist())
        self._filter_val_cb["values"] = self._filter_val_master
        self._filter_val_var.set("— Select Value —")

    def _on_filter_apply(self, _=None):
        """Apply the column+value filter to the current tab. Matches by substring,
        so partial text typed into the value box filters live (faster than exact match)."""
        tab = self.active_tab.get()
        df  = {"cleaned": self.cleaned_df, "removed": self.removed_df,
               "removed_reason": self.removed_reason_df, "remarks": self.remarks_df}.get(tab)
        if df is None:
            return
        col = self._filter_col_var.get()
        val = self._filter_val_var.get().strip()
        if col not in df.columns or val in ("", "— Select Value —", "(All)"):
            self._populate_table(df, self.search_var.get())
            self._filter_active = False
            return
        filtered = df[df[col].fillna("").astype(str).str.contains(val, case=False, na=False, regex=False)].copy()
        self._filter_active = True
        self._populate_table(filtered, self.search_var.get())

    def _clear_filter(self):
        self._filter_col_var.set("— Select Column —")
        self._filter_val_var.set("— Select Value —")
        self._filter_val_master = []
        self._filter_val_cb["values"] = []
        if self._filter_col_master:
            self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
        self._filter_active = False
        tab = self.active_tab.get()
        df  = {"cleaned": self.cleaned_df, "removed": self.removed_df,
               "removed_reason": self.removed_reason_df, "remarks": self.remarks_df}.get(tab)
        if df is not None:
            self._populate_table(df, self.search_var.get())

    def _clear_table(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text="")
        self._current_df   = None
        self._current_page = 0
        self._iid_to_row    = {}
        if hasattr(self, "prev_btn"):
            self.prev_btn.config(state="disabled")
            self.next_btn.config(state="disabled")
            self.page_label.config(text="")

    def _prev_page(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_df is not None:
            total = (len(self._current_df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE
            if self._current_page < total - 1:
                self._current_page += 1
                self._render_page()

    def _on_search(self, *_):
        tab = self.active_tab.get()
        df  = {
            "cleaned":        self.cleaned_df,
            "removed":        self.removed_df,
            "removed_reason": self.removed_reason_df,
            "remarks":        self.remarks_df,
        }.get(tab)
        if df is not None:
            self._populate_table(df, self.search_var.get())

    # ── FORMULA BAR ───────────────────────────────────────────────────────────
    def _get_raw_value(self, item_iid, col_name):
        """Look up the original (unformatted) value for a cell, bypassing the
        grid's display-only formatting (e.g. Date shown as dd-mm-yyyy, Time
        shown without its date) — mirrors how Excel's formula bar shows the
        raw underlying value regardless of the cell's display format."""
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return ""
        try:
            val = self._current_df.iloc[row_idx][col_name]
        except Exception:
            return ""
        if pd.isna(val):
            return ""
        return str(val)

    def _on_tree_click(self, event):
        """Detect which column was clicked and update the formula bar."""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        item   = self.tree.identify_row(event.y)
        if not item or not col_id:
            return
        cols = self.tree["columns"]
        try:
            col_idx = int(col_id.replace("#", "")) - 1
            col_name = cols[col_idx]
        except (ValueError, IndexError):
            return
        val = self._get_raw_value(item, col_name)
        self._update_fbar(col_name, val)

    def _on_cell_click(self, event=None):
        """Selection change — update bar with first column of selected row."""
        sel = self.tree.selection()
        if not sel:
            return
        cols = self.tree["columns"]
        if not cols:
            return
        # Only refresh bar if no specific column was detected by _on_tree_click
        # (this fires for keyboard nav / programmatic selection)
        if not hasattr(self, "_fbar_last_item") or self._fbar_last_item != sel[0]:
            self._fbar_last_item = sel[0]
            val = self._get_raw_value(sel[0], cols[0])
            self._update_fbar(cols[0], val)

    def _update_fbar(self, col_name, value):
        """Push column name + value into the Excel-style formula bar."""
        self._fbar_col_var.set(f"  {col_name}  ")
        self._fbar_val.config(state="normal")
        self._fbar_val.delete("1.0", "end")
        self._fbar_val.insert("1.0", str(value) if value != "" else "")
        self._fbar_val.config(state="disabled")

    def _fbar_copy(self):
        val = self._fbar_val.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(val)
        show_toast(self.master, "Value copied!", color=SUCCESS, duration=1500)

    # ── ROW DETAIL POPUP (double-click) ───────────────────────────────────────
    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        item    = sel[0]
        values  = self.tree.item(item, "values")
        columns = self.tree["columns"]
        if not columns or not values:
            return
        clicked_col = None
        if event and event.x:
            col_id = self.tree.identify_column(event.x)
            try:
                col_idx  = int(col_id.replace("#", "")) - 1
                clicked_col = columns[col_idx]
            except (ValueError, IndexError):
                pass
        self._show_row_detail(columns, values, clicked_col)

    def _show_row_detail(self, columns, values, clicked_col=None):
        popup = tk.Toplevel(self)
        popup.title("Row Details")
        popup.configure(bg=BG)
        popup.resizable(True, True)
        popup.grab_set()

        # ── Header bar ───────────────────────────────────────────────────────
        hdr = tk.Frame(popup, bg=HEADER_BG, height=48)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔍  Row Details",
                 font=("Segoe UI", 12, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text=f"{len(columns)} fields",
                 font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT
                 ).pack(side="left")
        tk.Button(hdr, text="✕  Close",
                  font=("Segoe UI", 8, "bold"),
                  bg=CARD, fg=TEXT, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=popup.destroy
                  ).pack(side="right", padx=12, pady=8)
        tk.Frame(popup, bg=ACCENT, height=2).pack(fill="x")

        # ── Scrollable body ───────────────────────────────────────────────────
        container = tk.Frame(popup, bg=BG)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg=BG, highlightthickness=0, bd=0)
        vsb2   = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner  = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>",  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))

        def _mw(e): canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        popup.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # ── Field rows ────────────────────────────────────────────────────────
        col_data = list(zip(columns, values))
        for i, (col, val) in enumerate(col_data):
            is_hl  = (col == clicked_col)
            row_bg = ACCENT2 if is_hl else (CARD if i % 2 == 0 else "#14172A")
            lbl_fg = "white"  if is_hl else SUBTEXT
            val_fg = "white"  if is_hl else TEXT

            row = tk.Frame(inner, bg=row_bg)
            row.pack(fill="x", padx=12, pady=1)

            tk.Label(row, text=f"{i+1:>3}", font=("Consolas", 8),
                     bg=row_bg, fg=SUBTEXT, width=3, anchor="e"
                     ).pack(side="left", padx=(8, 4), pady=6)

            tk.Label(row, text=col, font=("Segoe UI", 9, "bold"),
                     bg=row_bg, fg=lbl_fg, width=28, anchor="w"
                     ).pack(side="left", padx=(0, 10), pady=6)

            tk.Frame(row, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)

            val_str  = str(val) if val != "" else "—"
            vt = tk.Text(row, font=("Segoe UI", 9),
                         bg=row_bg, fg=val_fg,
                         relief="flat", bd=0, height=1, wrap="none",
                         cursor="xterm",
                         selectbackground=ACCENT, selectforeground="white")
            vt.insert("1.0", val_str)
            if len(val_str) > 80:
                vt.config(height=min((len(val_str) // 80) + 1, 4), wrap="word")
            vt.config(state="disabled")
            vt.pack(side="left", fill="x", expand=True, padx=(10, 8), pady=4)

        # ── Footer ────────────────────────────────────────────────────────────
        footer = tk.Frame(popup, bg=DARK)
        footer.pack(fill="x", side="bottom")

        def _copy_all():
            popup.clipboard_clear()
            popup.clipboard_append("\n".join(f"{c}\t{v}" for c, v in col_data))
            show_toast(popup, "All fields copied!", color=SUCCESS)

        tk.Button(footer, text="📋  Copy All Fields",
                  font=("Segoe UI", 9, "bold"),
                  bg=ACCENT, fg="white", relief="flat",
                  padx=16, pady=7, cursor="hand2",
                  command=_copy_all).pack(side="left", padx=12, pady=8)
        tk.Label(footer, text="Double-click any row to view full details.",
                 font=("Segoe UI", 8), bg=DARK, fg=SUBTEXT
                 ).pack(side="right", padx=16)

        popup.update_idletasks()
        popup.geometry("700x560")
        px = self.winfo_rootx() + (self.winfo_width()  - 700) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 560) // 2
        popup.geometry(f"700x560+{max(px,0)}+{max(py,0)}")

    def _download(self):
        save_path = filedialog.asksaveasfilename(
            title="Save output as",
            defaultextension=".xlsx",
            initialfile="cleaned_output.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            # Full output (Cleaned, Removed, Remarks, Unique ID sheets)
            with open(save_path, "wb") as f:
                f.write(self.output_bytes)

            self.status_var.set(f"Saved → {save_path}")
            show_toast(self.master, f"Saved: {os.path.basename(save_path)}", color=SUCCESS)
        except Exception as e:
            show_toast(self.master, f"Save failed: {e}", color=DANGER, duration=5000)


# ══════════════════════════════════════════════════════════════════════════════
# AUTOSTAT (IMPORT BATCH)
# ══════════════════════════════════════════════════════════════════════════════
class AutostatScreen(tk.Frame):
    PAGE_SIZE = 2000

    # Fixed header names applied positionally to the first columns of every
    # imported batch file (row 1 of the file is skipped/ignored for headers).
    FIXED_COLUMNS = [
        "CH Code", "Account Number", "Status", "Remark", "Remark by",
        "Remark Date", "PTP Amount", "PTP Date",
        "Claim Paid Amount", "Claim Paid Date",
    ]

    def __init__(self, master, username, role="user", on_back=None):
        super().__init__(master, bg=BG)
        self.username        = username
        self.role            = role
        self.on_back         = on_back
        self.file_path       = None
        self.imported_df     = None
        self._current_df     = None
        self._current_page   = 0
        self._iid_to_row     = {}
        self._filter_active  = False
        self._filter_col_master = []
        self._filter_val_master = []
        self._masterlist_path = os.path.join(_app_data_dir(), "remark_masterlist.json")
        self._masterlist = self._load_masterlist()
        self._agent_code_path = os.path.join(_app_data_dir(), "agent_code_masterlist.json")
        self._agent_codes = self._load_agent_codes()
        self._build_ui()

    # ── EDITED MASTERLIST (reference list for the Remark column) ───────────────
    def _load_masterlist(self):
        try:
            with open(self._masterlist_path, "r", encoding="utf-8") as f:
                items = json.load(f)
            return set(str(x).strip().upper() for x in items if str(x).strip())
        except Exception:
            return set()

    def _save_masterlist(self):
        try:
            with open(self._masterlist_path, "w", encoding="utf-8") as f:
                json.dump(sorted(self._masterlist), f, indent=2)
        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save Reference list:\n{e}", parent=self)

    # ── AGENT CODES MASTERLIST (reference list for the "Remark by" column) ─────
    def _load_agent_codes(self):
        try:
            with open(self._agent_code_path, "r", encoding="utf-8") as f:
                items = json.load(f)
            return set(str(x).strip().upper() for x in items if str(x).strip())
        except Exception:
            return set()

    def _save_agent_codes(self):
        try:
            with open(self._agent_code_path, "w", encoding="utf-8") as f:
                json.dump(sorted(self._agent_codes), f, indent=2)
        except Exception as e:
            messagebox.showerror("Save Failed", f"Could not save Agent Codes list:\n{e}", parent=self)

    def _open_masterlist_manager(self):
        is_admin = (self.role == "admin")

        win = tk.Toplevel(self)
        win.title("Reference — Remark Reference List")
        win.configure(bg=BG)
        win.geometry("480x560")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        tk.Label(win, text="📋  Reference", font=("Segoe UI", 12, "bold"),
                 bg=BG, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 2))

        if is_admin:
            subtitle = "Any Remark value not found in this list will be highlighted red in the table."
        else:
            subtitle = "Any Remark value not found in this list will be highlighted red in the table.  🔒 View only — editing is restricted to admins."
        tk.Label(win, text=subtitle, font=("Segoe UI", 8), bg=BG,
                 fg=SUBTEXT if is_admin else WARNING,
                 wraplength=440, justify="left"
                 ).pack(anchor="w", padx=16, pady=(0, 10))

        list_frame = tk.Frame(win, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))
        sb = ttk.Scrollbar(list_frame, orient="vertical")
        sb.pack(side="right", fill="y")
        lb = tk.Listbox(list_frame, bg=CARD, fg=TEXT, selectbackground=ACCENT2,
                         font=("Segoe UI", 9), relief="flat", highlightthickness=0,
                         yscrollcommand=sb.set)
        lb.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        sb.config(command=lb.yview)

        def refresh_list():
            lb.delete(0, "end")
            for item in sorted(self._masterlist):
                lb.insert("end", item)

        refresh_list()

        # ── Add entry row (admin only) ────────────────────────────────────────
        if is_admin:
            entry_row = tk.Frame(win, bg=BG)
            entry_row.pack(fill="x", padx=16, pady=(0, 8))
            entry_var = tk.StringVar()
            tk.Entry(entry_row, textvariable=entry_var, font=("Segoe UI", 9),
                      bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                      ).pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))

            def add_entry():
                val = entry_var.get().strip()
                if val:
                    self._masterlist.add(val.upper())
                    entry_var.set("")
                    refresh_list()
                    self._save_masterlist()

            tk.Button(entry_row, text="+ Add", font=("Segoe UI", 9, "bold"),
                      bg=ACCENT2, fg="white", relief="flat", padx=12, pady=5,
                      cursor="hand2", command=add_entry).pack(side="left")

        def remove_selected():
            sel = list(lb.curselection())
            for idx in reversed(sel):
                self._masterlist.discard(lb.get(idx))
            refresh_list()
            self._save_masterlist()

        def import_from_file():
            path = filedialog.askopenfilename(
                title="Import Remarks From File",
                filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"), ("All files", "*.*")]
            )
            if not path:
                return
            try:
                if path.lower().endswith(".csv"):
                    fdf = pd.read_csv(path, dtype=str, header=None)
                else:
                    fdf = pd.read_excel(path, dtype=str, header=None)
                values = fdf.iloc[:, 0].dropna().astype(str).str.strip()
                added = 0
                for v in values:
                    if v and v.lower() not in {m.lower() for m in self._masterlist}:
                        self._masterlist.add(v.upper())
                        added += 1
                refresh_list()
                self._save_masterlist()
                messagebox.showinfo("Import Complete", f"Added {added} new remark(s) to the Reference list.", parent=win)
            except Exception as e:
                messagebox.showerror("Import Failed", f"Could not read file:\n{e}", parent=win)

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))

        if is_admin:
            # Admin: full edit controls
            tk.Button(btn_row, text="🗑  Remove Selected", font=("Segoe UI", 8, "bold"),
                      bg=DANGER, fg="white", relief="flat", padx=10, pady=6,
                      cursor="hand2", command=remove_selected).pack(side="left")
            tk.Button(btn_row, text="📂  Import From File", font=("Segoe UI", 8, "bold"),
                      bg=CARD, fg=TEXT, relief="flat", padx=10, pady=6,
                      cursor="hand2", command=import_from_file).pack(side="left", padx=6)

        def save_and_close():
            if is_admin:
                self._save_masterlist()
            if self.imported_df is not None:
                self._populate_table(self.imported_df, self.search_var.get())
            win.destroy()

        close_label = "✓  Save & Close" if is_admin else "✕  Close"
        close_color = SUCCESS if is_admin else CARD
        close_fg    = "white" if is_admin else TEXT
        tk.Button(btn_row, text=close_label, font=("Segoe UI", 9, "bold"),
                  bg=close_color, fg=close_fg, relief="flat", padx=14, pady=6,
                  cursor="hand2", command=save_and_close).pack(side="right")

        # Ensure the Reference list is always written to disk on close (admin only).
        win.protocol("WM_DELETE_WINDOW", save_and_close)

    # ── AGENT CODES (reference list for the "Remark by" column) ────────────────
    def _open_agent_code_manager(self):
        is_admin = (self.role == "admin")

        win = tk.Toplevel(self)
        win.title("Agent Codes — Reference List")
        win.configure(bg=BG)
        win.geometry("480x560")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        tk.Label(win, text="🧑‍💼  Agent Codes", font=("Segoe UI", 12, "bold"),
                 bg=BG, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 2))

        if is_admin:
            subtitle = "Any \"Remark by\" value not found in this list will be highlighted red in the table."
        else:
            subtitle = "Any \"Remark by\" value not found in this list will be highlighted red in the table.  🔒 View only — editing is restricted to admins."
        tk.Label(win, text=subtitle, font=("Segoe UI", 8), bg=BG,
                 fg=SUBTEXT if is_admin else WARNING,
                 wraplength=440, justify="left"
                 ).pack(anchor="w", padx=16, pady=(0, 10))

        list_frame = tk.Frame(win, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))
        sb = ttk.Scrollbar(list_frame, orient="vertical")
        sb.pack(side="right", fill="y")
        lb = tk.Listbox(list_frame, bg=CARD, fg=TEXT, selectbackground=ACCENT2,
                         font=("Segoe UI", 9), relief="flat", highlightthickness=0,
                         yscrollcommand=sb.set)
        lb.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        sb.config(command=lb.yview)

        def refresh_list():
            lb.delete(0, "end")
            for item in sorted(self._agent_codes):
                lb.insert("end", item)

        refresh_list()

        # ── Add entry row (admin only) ────────────────────────────────────────
        if is_admin:
            entry_row = tk.Frame(win, bg=BG)
            entry_row.pack(fill="x", padx=16, pady=(0, 8))
            entry_var = tk.StringVar()
            tk.Entry(entry_row, textvariable=entry_var, font=("Segoe UI", 9),
                      bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                      ).pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))

            def add_entry():
                val = entry_var.get().strip()
                if val:
                    self._agent_codes.add(val.upper())
                    entry_var.set("")
                    refresh_list()
                    self._save_agent_codes()

            tk.Button(entry_row, text="+ Add", font=("Segoe UI", 9, "bold"),
                      bg=ACCENT2, fg="white", relief="flat", padx=12, pady=5,
                      cursor="hand2", command=add_entry).pack(side="left")

        def remove_selected():
            sel = list(lb.curselection())
            for idx in reversed(sel):
                self._agent_codes.discard(lb.get(idx))
            refresh_list()
            self._save_agent_codes()

        def import_from_file():
            path = filedialog.askopenfilename(
                title="Import Agent Codes From File",
                filetypes=[("Excel/CSV files", "*.xlsx *.xls *.csv"), ("All files", "*.*")]
            )
            if not path:
                return
            try:
                if path.lower().endswith(".csv"):
                    fdf = pd.read_csv(path, dtype=str, header=None)
                else:
                    fdf = pd.read_excel(path, dtype=str, header=None)
                values = fdf.iloc[:, 0].dropna().astype(str).str.strip()
                added = 0
                for v in values:
                    if v and v.lower() not in {m.lower() for m in self._agent_codes}:
                        self._agent_codes.add(v.upper())
                        added += 1
                refresh_list()
                self._save_agent_codes()
                messagebox.showinfo("Import Complete", f"Added {added} new agent code(s) to the list.", parent=win)
            except Exception as e:
                messagebox.showerror("Import Failed", f"Could not read file:\n{e}", parent=win)

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))

        if is_admin:
            # Admin: full edit controls
            tk.Button(btn_row, text="🗑  Remove Selected", font=("Segoe UI", 8, "bold"),
                      bg=DANGER, fg="white", relief="flat", padx=10, pady=6,
                      cursor="hand2", command=remove_selected).pack(side="left")
            tk.Button(btn_row, text="📂  Import From File", font=("Segoe UI", 8, "bold"),
                      bg=CARD, fg=TEXT, relief="flat", padx=10, pady=6,
                      cursor="hand2", command=import_from_file).pack(side="left", padx=6)

        def save_and_close():
            if is_admin:
                self._save_agent_codes()
            if self.imported_df is not None:
                self._populate_table(self.imported_df, self.search_var.get())
            win.destroy()

        close_label = "✓  Save & Close" if is_admin else "✕  Close"
        close_color = SUCCESS if is_admin else CARD
        close_fg    = "white" if is_admin else TEXT
        tk.Button(btn_row, text=close_label, font=("Segoe UI", 9, "bold"),
                  bg=close_color, fg=close_fg, relief="flat", padx=14, pady=6,
                  cursor="hand2", command=save_and_close).pack(side="right")

        # Ensure the Agent Codes list is always written to disk on close (admin only).
        win.protocol("WM_DELETE_WINDOW", save_and_close)

    # ── AUTO-CORRECT REMARKS against the Edited Masterlist ──────────────────────
    @staticmethod
    def _normalize_remark(text):
        """Loosely normalize a remark for comparison: lowercase, treat
        underscores/hyphens/multiple spaces as equivalent separators, so
        'VIBER - PTP_PARTIAL' and 'VIBER - PTP PARTIAL' are recognized
        as the same remark."""
        s = str(text).strip().lower()
        s = s.replace("_", " ").replace("-", " ")
        s = re.sub(r"\s+", " ", s)
        return s.strip()

    @staticmethod
    def _partial_ratio(a, b):
        """Like fuzzywuzzy's partial_ratio: finds the best-aligned substring
        match so that extra trailing/leading text (e.g. agent notes appended
        after the remark category) doesn't tank the similarity score.

        Guards against a false-positive failure mode: a very short string
        (e.g. "PU") can find some small window inside a much longer string
        that coincidentally matches near-perfectly, which would otherwise
        outrank a genuinely close, full-length match. The raw score is
        scaled down by how short the shorter string is — both on its own
        and relative to the longer string — so short candidates only score
        highly when they're a real, substantial chunk of the comparison,
        not just luck."""
        if not a or not b:
            return 0.0
        shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
        matcher = difflib.SequenceMatcher(None, shorter, longer)
        best = 0.0
        for block in matcher.get_matching_blocks():
            start = max(0, block.b - block.a)
            end   = start + len(shorter)
            window = longer[start:end]
            ratio = difflib.SequenceMatcher(None, shorter, window).ratio()
            if ratio > best:
                best = ratio

        shorter_len, longer_len = len(shorter), len(longer)
        if longer_len > 0:
            length_factor = (min(1.0, shorter_len / 6) *
                              min(1.0, shorter_len / (longer_len * 0.4)))
            best *= length_factor
        return best

    def _auto_correct_remarks(self):
        if self.imported_df is None or self.imported_df.empty:
            messagebox.showwarning("No Data", "Import a file first.", parent=self)
            return
        if "Remark" not in self.imported_df.columns:
            messagebox.showwarning("No Remark Column", "This file has no 'Remark' column.", parent=self)
            return
        if not self._masterlist:
            messagebox.showwarning("Masterlist Empty",
                                    "Add entries to the Reference list first.", parent=self)
            return

        masterlist_lower = {m.lower(): m for m in self._masterlist}
        norm_map = {}
        for m in self._masterlist:
            norm_map.setdefault(self._normalize_remark(m), m)
        norm_keys = list(norm_map.keys())

        df = self.imported_df
        corrections = []  # (row_index, original, corrected)

        def find_correction(original):
            text = str(original).strip()
            if not text:
                return None, None
            if text.lower() in masterlist_lower:
                return None, None  # already an exact match — nothing to do
            norm_text = self._normalize_remark(text)
            if norm_text in norm_map:
                return norm_map[norm_text], 1.0
            # Always pick the single nearest entry in the masterlist by
            # similarity ratio (using the best of full-string and
            # partial/substring similarity, so extra trailing text in the
            # remark doesn't prevent a good match), not just ones above a
            # strict cutoff.
            best_key, best_ratio = None, 0.0
            for key in norm_keys:
                full_ratio    = difflib.SequenceMatcher(None, norm_text, key).ratio()
                partial_ratio = self._partial_ratio(norm_text, key)
                ratio = max(full_ratio, partial_ratio)
                if ratio > best_ratio:
                    best_ratio, best_key = ratio, key
            if best_key is not None and best_ratio >= 0.55:
                return norm_map[best_key], best_ratio
            return None, None

        new_remarks = df["Remark"].copy()
        for idx, val in df["Remark"].items():
            corrected, ratio = find_correction(val)
            if corrected is not None and corrected != val:
                corrections.append((idx, str(val), corrected, ratio))
                new_remarks.at[idx] = corrected

        if not corrections:
            messagebox.showinfo("Auto-Correct", "No remarks needed correction — everything already matches the Reference list (or no close match was found).", parent=self)
            return

        # Preview dialog before applying
        win = tk.Toplevel(self)
        win.title("Auto-Correct Preview")
        win.configure(bg=BG)
        win.geometry("620x480")
        win.transient(self.winfo_toplevel())
        win.grab_set()

        tk.Label(win, text=f"🩹  {len(corrections)} remark(s) will be auto-corrected",
                 font=("Segoe UI", 11, "bold"), bg=BG, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 6))

        list_frame = tk.Frame(win, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        list_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))
        sb = ttk.Scrollbar(list_frame, orient="vertical")
        sb.pack(side="right", fill="y")
        preview_tree = ttk.Treeview(list_frame, columns=("orig", "corrected", "match"), show="headings",
                                     yscrollcommand=sb.set, height=14)
        preview_tree.heading("orig", text="Original")
        preview_tree.heading("corrected", text="Corrected To")
        preview_tree.heading("match", text="Match %")
        preview_tree.column("orig", width=250, anchor="w")
        preview_tree.column("corrected", width=250, anchor="w")
        preview_tree.column("match", width=70, anchor="center")
        preview_tree.pack(side="left", fill="both", expand=True)
        sb.config(command=preview_tree.yview)
        for _, orig, corr, ratio in corrections[:500]:
            pct = f"{ratio * 100:.0f}%" if ratio is not None else ""
            preview_tree.insert("", "end", values=(orig, corr, pct))
        if len(corrections) > 500:
            tk.Label(win, text=f"...and {len(corrections) - 500} more not shown.",
                     font=("Segoe UI", 8), bg=BG, fg=SUBTEXT).pack(anchor="w", padx=16)

        def apply_corrections():
            for idx, _orig, corr, _ratio in corrections:
                self.imported_df.at[idx, "Remark"] = str(corr).upper()
            self._populate_table(self.imported_df, self.search_var.get())
            win.destroy()
            show_toast(self.master, f"Auto-corrected {len(corrections)} remark(s).", color=SUCCESS)

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Button(btn_row, text="✕  Cancel", font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text=f"✓  Apply {len(corrections)} Correction(s)", font=("Segoe UI", 9, "bold"),
                  bg=SUCCESS, fg="white", relief="flat", padx=14, pady=7,
                  cursor="hand2", command=apply_corrections).pack(side="right")

    # ── DOWNLOAD ─────────────────────────────────────────────────────────────
    def _write_autostat_workbook(self, df, path):
        """Write a single styled Autostat workbook (header style + Remark
        highlighting against the Edited Masterlist) to `path`."""
        df.to_excel(path, sheet_name="Imported Data", index=False)

        wb = openpyxl.load_workbook(path)
        ws = wb["Imported Data"]

        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F8EF7", end_color="4F8EF7", fill_type="solid")

        if self._masterlist and "Remark" in df.columns:
            masterlist_norm = {self._normalize_remark(m) for m in self._masterlist}
            remark_col_idx = list(df.columns).index("Remark") + 1
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            red_font = Font(color="9C0006")
            for row in range(2, ws.max_row + 1):
                remark_val = self._normalize_remark(ws.cell(row=row, column=remark_col_idx).value or "")
                if remark_val and remark_val not in masterlist_norm:
                    ws.cell(row=row, column=remark_col_idx).fill = red_fill
                    ws.cell(row=row, column=remark_col_idx).font = red_font

        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        wb.save(path)

    def _download_autostat(self):
        if self.imported_df is None or self.imported_df.empty:
            messagebox.showwarning("No Data", "Import a file first.", parent=self)
            return

        save_path = filedialog.asksaveasfilename(
            title="Save Imported Data",
            defaultextension=".xlsx",
            initialfile=f"Autostat_Imported_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            base, ext = os.path.splitext(save_path)
            test_path = f"{base}_test{ext}"
            full_path = f"{base}_full{ext}"

            df_full = self.imported_df.copy()
            df_test = df_full.head(1).copy()

            self._write_autostat_workbook(df_test, test_path)
            self._write_autostat_workbook(df_full, full_path)

            show_toast(self.master,
                       f"Saved 2 files: {os.path.basename(test_path)} & {os.path.basename(full_path)}",
                       color=SUCCESS)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not save file:\n{e}", parent=self)

    # ── UI BUILD ─────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  📥  Autostat  —  Orico Auto Loan",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)
        tk.Label(left_hdr, text="Drop · Import · Browse Batch",
                 font=("Segoe UI", 8), bg=HEADER_BG, fg=SUBTEXT).pack(side="left", padx=4)

        right_hdr = tk.Frame(header, bg=HEADER_BG)
        right_hdr.pack(side="right", padx=16)
        role_colors = {"admin": DANGER, "supervisor": WARNING, "user": SUCCESS}
        role_color  = role_colors.get(self.role, ACCENT)
        tk.Label(right_hdr, text=f"👤  {self.username}",
                 font=("Segoe UI", 9, "bold"), bg=HEADER_BG, fg=ACCENT).pack()
        tk.Label(right_hdr, text=f"● {self.role.upper()}",
                 font=("Segoe UI", 7, "bold"), bg=HEADER_BG, fg=role_color).pack()

        tk.Frame(self, bg=ACCENT2, height=3).pack(fill="x")

        # ── Body ──────────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # ── LEFT PANEL ────────────────────────────────────────────────────────
        left = tk.Frame(body, bg=BG, width=300)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        # ── Drop Zone ─────────────────────────────────────────────────────────
        self._drop_frame = tk.Frame(left, bg=CARD, highlightthickness=2,
                                    highlightbackground=BORDER)
        self._drop_frame.pack(fill="x", pady=(0, 10))

        self._drop_empty = tk.Frame(self._drop_frame, bg=CARD)
        self._drop_empty.pack(fill="both", padx=20, pady=20)

        self._drop_icon  = tk.Label(self._drop_empty, text="📂", font=("Segoe UI", 26),
                                    bg=CARD, fg=ACCENT2)
        self._drop_icon.pack()
        self._drop_title = tk.Label(self._drop_empty, text="Drop batch file here",
                                    font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT)
        self._drop_title.pack(pady=(6, 2))
        tk.Label(self._drop_empty, text=".xlsx  ·  .xlsm  ·  .xls  ·  .csv",
                 font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack()
        tk.Label(self._drop_empty, text="or", font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT).pack(pady=(8, 4))
        tk.Button(self._drop_empty, text="Browse File", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat", padx=16, pady=6,
                  cursor="hand2", command=self._browse).pack()

        # ── State B: file loaded display ───────────────────────────────────────
        self._drop_loaded = tk.Frame(self._drop_frame, bg=CARD)

        loaded_top = tk.Frame(self._drop_loaded, bg="#1A3A2A")
        loaded_top.pack(fill="x")
        tk.Label(loaded_top, text="✅  File Loaded", font=("Segoe UI", 9, "bold"),
                 bg="#1A3A2A", fg=SUCCESS).pack(side="left", padx=12, pady=6)
        tk.Button(loaded_top, text="✕  Clear", font=("Segoe UI", 8),
                  bg="#1A3A2A", fg=SUBTEXT, relief="flat",
                  cursor="hand2", command=self._clear_file
                  ).pack(side="right", padx=8)

        loaded_body = tk.Frame(self._drop_loaded, bg=CARD)
        loaded_body.pack(fill="x", padx=14, pady=12)

        tk.Label(loaded_body, text="📄", font=("Segoe UI", 22),
                 bg=CARD, fg=ACCENT2).pack(side="left", padx=(0, 10))

        loaded_text = tk.Frame(loaded_body, bg=CARD)
        loaded_text.pack(side="left", fill="x", expand=True)
        self._fi_name = tk.Label(loaded_text, text="", font=("Segoe UI", 9, "bold"),
                                  bg=CARD, fg=TEXT, anchor="w", wraplength=180, justify="left")
        self._fi_name.pack(fill="x")
        self._fi_meta = tk.Label(loaded_text, text="", font=("Segoe UI", 7),
                                  bg=CARD, fg=SUBTEXT, anchor="w")
        self._fi_meta.pack(fill="x", pady=(2, 0))

        tk.Button(self._drop_loaded, text="🔄  Change File",
                  font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT,
                  relief="flat", cursor="hand2",
                  command=self._browse).pack(pady=(0, 10))

        # Register DnD on the drop zone (and its children) if available
        self._dnd_enabled = False
        try:
            for widget in [self._drop_frame, self._drop_empty, self._drop_loaded,
                           self._drop_icon, self._drop_title, loaded_top, loaded_body]:
                widget.drop_target_register('DND_Files')
                widget.dnd_bind('<<Drop>>',      self._on_drop)
                widget.dnd_bind('<<DragEnter>>', self._on_drag_enter)
                widget.dnd_bind('<<DragLeave>>', self._on_drag_leave)
            self._dnd_enabled = True
        except Exception:
            pass

        tk.Label(left, text="SUMMARY", font=("Segoe UI", 8, "bold"),
                 bg=BG, fg=SUBTEXT).pack(anchor="w", pady=(0, 6))

        self.stat_vars = {}
        for key, label, color, icon in [
            ("total",   "Total Rows",    TEXT,    "📋"),
            ("columns", "Columns Found", ACCENT2, "🗂"),
        ]:
            c = tk.Frame(left, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
            c.pack(fill="x", pady=3)
            inn = tk.Frame(c, bg=CARD)
            inn.pack(fill="x", padx=12, pady=8)
            tk.Label(inn, text=icon, font=("Segoe UI", 11), bg=CARD, fg=color).pack(side="left")
            tk.Label(inn, text=label, font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(side="left", padx=8)
            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(inn, textvariable=var, font=("Segoe UI", 11, "bold"),
                     bg=CARD, fg=color).pack(side="right")

        can_upload         = self.role in ("admin", "user")
        self._can_upload   = can_upload

        self.process_btn = tk.Button(left, text="▶  Import File",
                                     font=("Segoe UI", 10, "bold"), bg=ACCENT2, fg="white",
                                     relief="flat", pady=10, cursor="hand2",
                                     state="disabled",
                                     command=self._run_import)
        self.process_btn.pack(fill="x", pady=(14, 6))
        if not can_upload:
            self.process_btn.config(state="disabled", bg=CARD, fg=SUBTEXT,
                                    text="▶  Import File  (No Access)")

        self.add_more_btn = tk.Button(left, text="➕  Add More Data",
                                      font=("Segoe UI", 9, "bold"), bg=CARD, fg=TEXT,
                                      relief="flat", pady=8, cursor="hand2",
                                      state="disabled",
                                      command=self._open_add_more_data)
        self.add_more_btn.pack(fill="x", pady=(0, 6))
        if not can_upload:
            self.add_more_btn.config(state="disabled", fg=SUBTEXT,
                                     text="➕  Add More Data  (No Access)")

        style = ttk.Style()
        style.configure("TProgressbar", troughcolor=CARD, background=ACCENT2, thickness=4)
        self.progress = ttk.Progressbar(left, mode="indeterminate")
        self.progress.pack(fill="x", pady=(10, 0))

        # ── RIGHT PANEL ───────────────────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        tab_bar = tk.Frame(right, bg=BG)
        tab_bar.pack(fill="x", pady=(0, 8))
        tk.Label(tab_bar, text="✅ Imported Rows", font=("Segoe UI", 9, "bold"),
                 bg=ACCENT2, fg="white", padx=14, pady=7).pack(side="left")
        if self.role == "admin":
            tk.Button(tab_bar, text="📋 Reference", font=("Segoe UI", 9, "bold"),
                      bg=CARD, fg=TEXT, relief="flat", padx=14, pady=7, cursor="hand2",
                      command=self._open_masterlist_manager).pack(side="left", padx=(8, 0))
            tk.Button(tab_bar, text="🧑‍💼 Agent Codes", font=("Segoe UI", 9, "bold"),
                      bg=CARD, fg=TEXT, relief="flat", padx=14, pady=7, cursor="hand2",
                      command=self._open_agent_code_manager).pack(side="left", padx=(8, 0))
        else:
            tk.Button(tab_bar, text="📋 Reference  🔒", font=("Segoe UI", 9, "bold"),
                      bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7, cursor="hand2",
                      command=self._open_masterlist_manager).pack(side="left", padx=(8, 0))
            tk.Button(tab_bar, text="🧑‍💼 Agent Codes  🔒", font=("Segoe UI", 9, "bold"),
                      bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7, cursor="hand2",
                      command=self._open_agent_code_manager).pack(side="left", padx=(8, 0))
        tk.Button(tab_bar, text="⬇  Download (2 files)", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat", padx=14, pady=7, cursor="hand2",
                  command=self._download_autostat).pack(side="right")

        # ── Search bar ────────────────────────────────────────────────────────
        search_frame = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        search_frame.pack(fill="x", pady=(0, 4))
        tk.Label(search_frame, text="🔍", bg=CARD, fg=SUBTEXT, font=("Segoe UI", 10)).pack(side="left", padx=8)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                 ).pack(side="left", fill="x", expand=True, pady=8, padx=4)
        tk.Label(search_frame, text="Search all columns…", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        # ── Column Dropdown Filter bar ─────────────────────────────────────────
        filter_outer = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        filter_outer.pack(fill="x", pady=(0, 8))
        tk.Label(filter_outer, text="▼ Column Filter:", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=8, pady=6)
        self._filter_col_var = tk.StringVar(value="— Select Column —")
        self._filter_val_var = tk.StringVar(value="— Select Value —")
        self._filter_col_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_col_var,
                                             width=24, font=("Segoe UI", 9))
        self._filter_col_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_val_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_val_var,
                                             width=28, font=("Segoe UI", 9))
        self._filter_val_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_col_cb.bind("<<ComboboxSelected>>", self._on_filter_col_change)
        self._filter_val_cb.bind("<<ComboboxSelected>>", self._on_filter_apply)
        self._filter_col_cb.bind("<KeyRelease>", self._on_filter_col_type)
        self._filter_val_cb.bind("<KeyRelease>", self._on_filter_val_type)
        tk.Button(filter_outer, text="✕ Clear Filter", font=("Segoe UI", 8, "bold"),
                  bg=DANGER, fg="white", relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self._clear_filter).pack(side="left", padx=(0, 6))

        # ── Error textboxes (Status / Agent Codes) ──────────────────────────────
        # Each box only appears while there is at least one flagged row for
        # that category, and disappears automatically once fixed.
        self._status_err_frame = tk.Frame(right, bg="#2A1414",
                                          highlightthickness=1, highlightbackground=DANGER)
        status_err_hdr = tk.Frame(self._status_err_frame, bg="#2A1414")
        status_err_hdr.pack(fill="x", padx=10, pady=(6, 0))
        tk.Label(status_err_hdr, text="🔴  Status Issues", font=("Segoe UI", 9, "bold"),
                 bg="#2A1414", fg="#FF6B6B").pack(side="left")
        self._status_err_text = tk.Text(self._status_err_frame, font=("Segoe UI", 9),
                                        bg="#2A1414", fg="#FF6B6B", relief="flat", bd=0,
                                        height=2, wrap="word", cursor="xterm",
                                        selectbackground=DANGER, selectforeground="white",
                                        state="disabled")
        self._status_err_text.pack(fill="x", padx=10, pady=(2, 8))

        self._agent_err_frame = tk.Frame(right, bg="#2A2014",
                                         highlightthickness=1, highlightbackground=WARNING)
        agent_err_hdr = tk.Frame(self._agent_err_frame, bg="#2A2014")
        agent_err_hdr.pack(fill="x", padx=10, pady=(6, 0))
        tk.Label(agent_err_hdr, text="🟠  Remark By / Agent Code Issues", font=("Segoe UI", 9, "bold"),
                 bg="#2A2014", fg="#F5A623").pack(side="left")
        self._agent_err_text = tk.Text(self._agent_err_frame, font=("Segoe UI", 9),
                                       bg="#2A2014", fg="#F5A623", relief="flat", bd=0,
                                       height=2, wrap="word", cursor="xterm",
                                       selectbackground=WARNING, selectforeground="white",
                                       state="disabled")
        self._agent_err_text.pack(fill="x", padx=10, pady=(2, 8))
        # Both start hidden — _render_page packs them back in when there's something to show.

        table_frame = tk.Frame(right, bg=BG)
        table_frame.pack(fill="both", expand=True)
        self._table_frame = table_frame
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style.configure("Custom.Treeview", background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28, font=("Segoe UI", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading", background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT2)], foreground=[("selected", "white")])

        self.tree = ttk.Treeview(table_frame, style="Custom.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                  show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self._on_cell_click)
        self.tree.bind("<ButtonRelease-1>",  self._on_tree_click)
        self.tree.bind("<Double-1>",         self._on_double_click)
        self.tree.bind("<Return>",           self._on_double_click)

        self.row_count_label = tk.Label(right, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.row_count_label.pack(anchor="e", pady=(4, 0))

        # ── Excel-style formula bar ───────────────────────────────────────────
        fbar = tk.Frame(right, bg="#1C1F2E",
                        highlightthickness=1, highlightbackground=BORDER)
        fbar.pack(fill="x", pady=(4, 2))

        tk.Label(fbar, text=" fx ", font=("Segoe UI", 9, "bold"),
                 bg="#2E7D32", fg="white", padx=6, pady=4
                 ).pack(side="left")

        self._fbar_col_var = tk.StringVar(value="")
        col_box = tk.Entry(fbar, textvariable=self._fbar_col_var,
                           font=("Segoe UI", 9, "bold"),
                           bg="#252840", fg=ACCENT2,
                           relief="flat", bd=0,
                           width=22, justify="center",
                           insertbackground=ACCENT2,
                           readonlybackground="#252840",
                           state="readonly")
        col_box.pack(side="left", padx=(1, 0), ipady=4)

        tk.Frame(fbar, bg=BORDER, width=1).pack(side="left", fill="y", pady=2)

        self._fbar_val = tk.Text(fbar, font=("Segoe UI", 9),
                                 bg="#1C1F2E", fg=TEXT,
                                 relief="flat", bd=0,
                                 height=1, wrap="none",
                                 cursor="xterm",
                                 selectbackground=ACCENT2,
                                 selectforeground="white",
                                 insertbackground=TEXT,
                                 state="disabled")
        fbar_xsb = ttk.Scrollbar(fbar, orient="horizontal",
                                  command=self._fbar_val.xview)
        self._fbar_val.configure(xscrollcommand=fbar_xsb.set)
        self._fbar_val.pack(side="left", fill="x", expand=True, padx=(6, 0), pady=2)
        tk.Button(fbar, text="⎘", font=("Segoe UI", 10),
                  bg="#1C1F2E", fg=SUBTEXT, relief="flat",
                  cursor="hand2", padx=6,
                  command=self._fbar_copy
                  ).pack(side="right", padx=4)

        nav_frame = tk.Frame(right, bg=BG)
        nav_frame.pack(fill="x", pady=(4, 0))
        self.prev_btn = tk.Button(nav_frame, text="◀  Prev", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._prev_page)
        self.prev_btn.pack(side="left", padx=(0, 6))
        self.page_label = tk.Label(nav_frame, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.page_label.pack(side="left")
        self.next_btn = tk.Button(nav_frame, text="Next  ▶", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._next_page)
        self.next_btn.pack(side="left", padx=6)

        # Status bar
        self.status_var = tk.StringVar(value="Ready — drop or browse a batch file to begin.")
        tk.Label(self, textvariable=self.status_var, font=("Segoe UI", 8),
                 bg=HEADER_BG, fg=SUBTEXT, anchor="w", padx=16, pady=6).pack(fill="x", side="bottom")

        # Footer
        watermark = tk.Frame(self, bg=DARK)
        watermark.pack(fill="x", side="bottom")
        tk.Label(watermark, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=6)
        tk.Label(watermark, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=6)

    # ── FILE LOAD ─────────────────────────────────────────────────────────────
    def _on_drag_enter(self, event):
        self._drop_frame.config(highlightbackground=ACCENT2, highlightthickness=2,
                                bg="#241A40")
        self._drop_icon.config(bg="#241A40", fg=ACCENT2, text="⬇️")
        self._drop_title.config(bg="#241A40", fg=ACCENT2, text="Release to load file")
        for w in self._drop_frame.winfo_children():
            try: w.config(bg="#241A40")
            except Exception: pass

    def _on_drag_leave(self, event):
        self._drop_frame.config(highlightbackground=BORDER, bg=CARD)
        self._drop_icon.config(bg=CARD, fg=ACCENT2, text="📂")
        self._drop_title.config(bg=CARD, fg=TEXT, text="Drop batch file here")
        for w in self._drop_frame.winfo_children():
            try: w.config(bg=CARD)
            except Exception: pass

    def _on_drop(self, event):
        self._on_drag_leave(None)
        raw = event.data.strip()
        paths = []
        if raw.startswith("{"):
            import re
            paths = re.findall(r'\{([^}]+)\}', raw)
        if not paths:
            paths = [raw]
        self._load_file(paths[0])

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select batch file",
            filetypes=[("Excel/CSV files", "*.xlsx *.xlsm *.xls *.csv")]
        )
        if path:
            self._load_file(path)

    def _clear_file(self):
        self.file_path = None
        self._drop_loaded.pack_forget()
        self._drop_empty.pack(fill="both", padx=20, pady=20)
        self._drop_frame.config(highlightbackground=BORDER)
        self.process_btn.config(state="disabled")
        if hasattr(self, "add_more_btn"):
            self.add_more_btn.config(state="disabled")
        self.status_var.set("Ready — drop or browse a batch file to begin.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        self.file_path = path
        fname = os.path.basename(path)
        fsize = os.path.getsize(path)
        fsize_str = (f"{fsize / 1024:.1f} KB" if fsize < 1_048_576
                     else f"{fsize / 1_048_576:.2f} MB")
        import time
        mtime     = os.path.getmtime(path)
        mtime_str = time.strftime("%b %d, %Y  %I:%M %p", time.localtime(mtime))

        self._drop_empty.pack_forget()
        self._fi_name.config(text=fname)
        self._fi_meta.config(text=f"{fsize_str}  ·  {mtime_str}")
        self._drop_loaded.pack(fill="both")
        self._drop_frame.config(highlightbackground=SUCCESS)

        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.status_var.set(f"Loaded: {fname}  —  Click ▶ Import File to load it.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    # ── IMPORT ───────────────────────────────────────────────────────────────
    def _run_import(self):
        try:
            if self.role not in ("admin", "user"):
                messagebox.showwarning("Access Denied", "Your role does not have upload/import access.")
                return
            if not self.file_path:
                messagebox.showwarning("No File", "Please select a file first.")
                return
            self.process_btn.config(state="disabled")
            self.status_var.set("Importing…")
            self.progress.start(10)
            threading.Thread(target=self._import_thread, daemon=True).start()
        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to start import:\n{e}")

    def _clean_batch_file(self, path):
        """Read and clean a single batch file (xlsx/xlsm/xls/csv) into a
        DataFrame using the same rules as the main import: fixed column
        names, blank-row removal, date reformatting, Remark normalisation,
        and the ACTION: rewrite. Shared by both the main "Import File"
        button and the "Add More Data" merge window so every file goes
        through identical cleaning."""
        is_csv = path.lower().endswith(".csv")

        # Row 1 is always a title/junk row — skip it.
        # Row 2 becomes the header; data starts from row 3.
        if is_csv:
            df = pd.read_csv(path, dtype=str, skiprows=1)
        else:
            df = pd.read_excel(path, dtype=str, skiprows=1)

        # Apply fixed column names positionally for the first N columns.
        new_cols = list(df.columns)
        for i, name in enumerate(self.FIXED_COLUMNS):
            if i < len(new_cols):
                new_cols[i] = name
        df.columns = new_cols

        # Drop fully blank rows (NaN or whitespace-only across every cell).
        df = df.dropna(how="all")
        df = df.fillna("")
        blank_mask = (df.astype(str).apply(lambda c: c.str.strip()) == "").all(axis=1)
        df = df[~blank_mask].reset_index(drop=True)

        # Reformat Remark Date to "MM/DD/YYYY HH:MM:SS" (24-hour, no AM/PM).
        if "Remark Date" in df.columns:
            df["Remark Date"] = df["Remark Date"].apply(self._format_remark_date)

        # Claim Paid Date — date only, no time portion.
        if "Claim Paid Date" in df.columns:
            def _format_claim_paid_date(value):
                if value is None:
                    return value
                text = str(value).strip()
                if text == "" or text.lower() == "nan":
                    return value
                parsed = pd.to_datetime(text, errors="coerce")
                if pd.notna(parsed):
                    return parsed.strftime("%m/%d/%Y")
                return value
            df["Claim Paid Date"] = df["Claim Paid Date"].apply(_format_claim_paid_date)

        # PTP Date — date only, no time portion.
        if "PTP Date" in df.columns:
            def _format_date_only(value):
                if value is None:
                    return value
                text = str(value).strip()
                if text == "" or text.lower() == "nan":
                    return value
                parsed = pd.to_datetime(text, errors="coerce")
                if pd.notna(parsed):
                    return parsed.strftime("%m/%d/%Y")
                return value
            df["PTP Date"] = df["PTP Date"].apply(_format_date_only)

        # Normalise Remark to ALL CAPS.
        if "Remark" in df.columns:
            df["Remark"] = df["Remark"].astype(str).str.strip().str.upper()
            df["Remark"] = df["Remark"].replace({"NAN": "", "REMARK": ""})

        # Rewrite ACTION: <word> in Remark based on Status:
        #   Status contains KEPT → ACTION: KEPT
        #   Status contains PTP  → ACTION: PTP
        #   Neither              → ACTION: SRP
        if "Remark" in df.columns and "Status" in df.columns:
            status_up = df["Status"].astype(str).str.strip().str.upper()
            has_action = df["Remark"].str.contains(r"ACTION\s*:\s*\w+", regex=True, na=False)
            for idx in df.index[has_action]:
                st = status_up.at[idx]
                if "KEPT" in st:
                    target = "KEPT"
                elif "PTP" in st:
                    target = "PTP"
                else:
                    target = "SRP"
                df.at[idx, "Remark"] = re.sub(
                    r"ACTION\s*:\s*\w+", f"ACTION: {target}", df.at[idx, "Remark"]
                )

        return df

    def _import_thread(self):
        try:
            df = self._clean_batch_file(self.file_path)
            self.after(0, lambda: self._on_import_done(df))
        except Exception as e:
            self.after(0, lambda: self._on_import_error(str(e)))

    @staticmethod
    def _format_remark_date(value):
        """Force any date/time value into 'MM/DD/YYYY HH:MM:SS' (24-hour,
        zero-padded, no AM/PM). Falls back to the original text if it
        truly cannot be parsed as a date."""
        if value is None:
            return value
        text = str(value).strip()
        if text == "" or text.lower() == "nan":
            return value

        # First try pandas' general parser (handles AM/PM, ISO, etc.)
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.notna(parsed):
            return parsed.strftime("%m/%d/%Y %H:%M:%S")

        # Fallback: manually pull out an "M/D/YYYY H:MM:SS [AM/PM]" style
        # pattern and zero-pad / convert to 24-hour by hand.
        m = re.match(
            r"^\s*(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})"
            r"(?:\s+(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?\s*([AaPp][Mm])?)?\s*$",
            text
        )
        if m:
            mo, da, yr, hh, mm, ss, ampm = m.groups()
            mo, da = int(mo), int(da)
            yr = int(yr)
            if yr < 100:
                yr += 2000
            hh = int(hh) if hh else 0
            mm = int(mm) if mm else 0
            ss = int(ss) if ss else 0
            # Only treat the hour as a 12-hour value (and convert) if it's
            # actually in the valid 12-hour range. If the source text is
            # malformed — e.g. an already-24-hour value like "13:00:00 PM"
            # that pandas' parser rejected and we fell back to here — adding
            # 12 again would push it past 23 (13 -> 25). Guard against that.
            if ampm and 1 <= hh <= 12:
                ampm = ampm.upper()
                if ampm == "PM" and hh != 12:
                    hh += 12
                elif ampm == "AM" and hh == 12:
                    hh = 0
            try:
                return f"{mo:02d}/{da:02d}/{yr:04d} {hh:02d}:{mm:02d}:{ss:02d}"
            except Exception:
                return value

        return value

    def _on_import_done(self, df):
        self.progress.stop()
        self.imported_df = df
        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        if hasattr(self, "add_more_btn"):
            self.add_more_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.stat_vars["total"].set(str(len(df)))
        self.stat_vars["columns"].set(str(len(df.columns)))
        self.status_var.set(f"Done  ·  {len(df)} rows imported  ·  {len(df.columns)} columns")
        self._filter_col_master = list(df.columns)
        self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
        self._filter_col_var.set("— Select Column —")
        self._filter_val_var.set("— Select Value —")
        self._filter_val_master = []
        self._filter_val_cb["values"] = []
        self._filter_active = False
        self._populate_table(df)
        show_toast(self.master, "Batch imported!", color=SUCCESS)

    def _on_import_error(self, msg):
        self.progress.stop()
        self.process_btn.config(state="normal" if self.role in ("admin", "user") else "disabled")
        self.status_var.set(f"Error: {msg}")
        messagebox.showerror("Import Error", msg)

    # ── ADD MORE DATA (merge an additional file into the current import) ───────
    def _open_add_more_data(self):
        if self.role not in ("admin", "user"):
            messagebox.showwarning("Access Denied", "Your role does not have upload/import access.", parent=self)
            return
        if self.imported_df is None or self.imported_df.empty:
            messagebox.showwarning("No Data", "Import a main file first, then add more data on top of it.", parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Add More Data")
        win.configure(bg=BG)
        win.geometry("460x460")
        win.resizable(False, False)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(win, bg=HEADER_BG)
        hdr.pack(fill="x", side="top")
        tk.Label(hdr, text="➕  Add More Data", font=("Segoe UI", 12, "bold"),
                 bg=HEADER_BG, fg=TEXT).pack(side="left", padx=16, pady=10)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT,
                  relief="flat", cursor="hand2", command=win.destroy
                  ).pack(side="right", padx=12)
        tk.Frame(win, bg=ACCENT2, height=3).pack(fill="x", side="top")

        # ── Button row: pack to the bottom FIRST so it always reserves its
        # own space and can never be squeezed off-screen by the content
        # above, no matter how tall that content ends up being. ─────────────
        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(8, 14), side="bottom")

        tk.Label(win, text=f"Currently loaded: {len(self.imported_df)} row(s).\n"
                            "Pick another file below to append its (cleaned) rows on top "
                            "of what's already imported.",
                 font=("Segoe UI", 8), bg=BG, fg=SUBTEXT, wraplength=420, justify="left"
                 ).pack(anchor="w", padx=16, pady=(10, 8), side="top")

        # ── Drop / browse zone ───────────────────────────────────────────────
        drop_frame = tk.Frame(win, bg=CARD, highlightthickness=2, highlightbackground=BORDER)
        drop_frame.pack(fill="x", padx=16, pady=(0, 10), side="top")
        drop_inner = tk.Frame(drop_frame, bg=CARD)
        drop_inner.pack(fill="both", padx=16, pady=16)
        drop_icon = tk.Label(drop_inner, text="📂", font=("Segoe UI", 22), bg=CARD, fg=ACCENT2)
        drop_icon.pack()
        drop_title = tk.Label(drop_inner, text="Drop additional file here",
                              font=("Segoe UI", 10, "bold"), bg=CARD, fg=TEXT)
        drop_title.pack(pady=(6, 2))
        tk.Label(drop_inner, text=".xlsx  ·  .xlsm  ·  .xls  ·  .csv",
                 font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack()

        file_var   = tk.StringVar(value="")
        file_label = tk.Label(win, textvariable=file_var, font=("Segoe UI", 9, "bold"),
                              bg=BG, fg=SUCCESS, wraplength=420, justify="left")
        file_label.pack(anchor="w", padx=16, pady=(0, 4), side="top")

        state = {"path": None}

        def pick_file(path):
            if not path or not os.path.exists(path):
                return
            state["path"] = path
            fname = os.path.basename(path)
            file_var.set(f"✅  {fname}")
            merge_btn.config(state="normal")

        def browse_file():
            path = filedialog.askopenfilename(
                title="Select additional batch file",
                filetypes=[("Excel/CSV files", "*.xlsx *.xlsm *.xls *.csv")]
            )
            if path:
                pick_file(path)

        tk.Button(drop_inner, text="Browse File", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat", padx=16, pady=6,
                  cursor="hand2", command=browse_file).pack(pady=(8, 0))

        # Drag-and-drop support, same as the main drop zone.
        try:
            def on_drop(event):
                raw = event.data.strip()
                paths = []
                if raw.startswith("{"):
                    paths = re.findall(r'\{([^}]+)\}', raw)
                if not paths:
                    paths = [raw]
                pick_file(paths[0])
            for widget in [drop_frame, drop_inner, drop_icon, drop_title]:
                widget.drop_target_register('DND_Files')
                widget.dnd_bind('<<Drop>>', on_drop)
        except Exception:
            pass

        progress_bar = ttk.Progressbar(win, mode="indeterminate")
        progress_bar.pack(fill="x", padx=16, pady=(0, 6), side="top")

        status_var = tk.StringVar(value="")
        tk.Label(win, textvariable=status_var, font=("Segoe UI", 8),
                 bg=BG, fg=SUBTEXT).pack(anchor="w", padx=16, side="top")

        def do_merge():
            path = state["path"]
            if not path:
                return
            merge_btn.config(state="disabled")
            cancel_btn.config(state="disabled")
            status_var.set("Reading and cleaning the additional file…")
            progress_bar.start(10)

            def worker():
                try:
                    new_df = self._clean_batch_file(path)
                    self.after(0, lambda: on_merge_done(new_df))
                except Exception as e:
                    self.after(0, lambda: on_merge_error(str(e)))

            threading.Thread(target=worker, daemon=True).start()

        def on_merge_done(new_df):
            progress_bar.stop()
            before = len(self.imported_df)
            # Align columns: any new columns the additional file doesn't have
            # come through blank; extra columns it has but the main file
            # doesn't are kept too, so nothing is silently dropped.
            combined = pd.concat([self.imported_df, new_df], ignore_index=True, sort=False)
            combined = combined.fillna("")
            self.imported_df = combined
            added = len(combined) - before
            self.stat_vars["total"].set(str(len(combined)))
            self.stat_vars["columns"].set(str(len(combined.columns)))
            self.status_var.set(f"Done  ·  {len(combined)} rows total  ·  +{added} added")
            self._filter_col_master = list(combined.columns)
            self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
            self._filter_col_var.set("— Select Column —")
            self._filter_val_var.set("— Select Value —")
            self._filter_val_master = []
            self._filter_val_cb["values"] = []
            self._filter_active = False
            self._populate_table(combined, self.search_var.get())
            win.destroy()
            show_toast(self.master, f"Added {added} row(s)  ·  {len(combined)} total.", color=SUCCESS)

        def on_merge_error(msg):
            progress_bar.stop()
            merge_btn.config(state="normal")
            cancel_btn.config(state="normal")
            status_var.set("")
            messagebox.showerror("Add More Data — Error", msg, parent=win)

        cancel_btn = tk.Button(btn_row, text="✕  Cancel", font=("Segoe UI", 9),
                               bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7,
                               cursor="hand2", command=win.destroy)
        cancel_btn.pack(side="right", padx=(6, 0))
        merge_btn = tk.Button(btn_row, text="➕  Add To Current Data", font=("Segoe UI", 9, "bold"),
                              bg=SUCCESS, fg="white", relief="flat", padx=14, pady=7,
                              cursor="hand2", state="disabled", command=do_merge)
        merge_btn.pack(side="right")

        win.update_idletasks()
        px = self.winfo_rootx() + (self.winfo_width()  - 460) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 460) // 2
        win.geometry(f"460x460+{max(px,0)}+{max(py,0)}")

    # ── TABLE ────────────────────────────────────────────────────────────────
    def _populate_table(self, df, filter_text=""):
        self._clear_table()
        if df is None or df.empty:
            self.row_count_label.config(text="No rows to display.")
            return
        cols = list(df.columns)
        self.tree["columns"] = cols
        for col in cols:
            sample      = df[col].astype(str)
            # Auto-fit: width based on the longest actual value in the column
            # (header included), not a sampled estimate, so every column
            # is wide enough to show its content without truncation.
            max_len     = max([len(str(col))] + [len(v) for v in sample] or [10])
            col_width   = max(70, min(max_len * 8 + 16, 400))
            self.tree.heading(col, text=col, anchor="w",
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=col_width, minwidth=60, anchor="w", stretch=False)
        filt = filter_text.lower().strip()
        if filt:
            combined = df.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            filtered = df[combined.str.contains(filt, regex=False)]
        else:
            filtered = df
        self._current_df   = filtered
        self._current_page = 0
        self._render_page()

    def _set_error_box(self, frame, text_widget, lines, total_count, label):
        """Show/update an error textbox if there's something to report,
        otherwise hide it completely (auto-vanishes once fixed)."""
        if total_count <= 0:
            frame.pack_forget()
            return
        shown = lines[:20]
        body = "\n".join(shown)
        if total_count > len(shown):
            body += f"\n…and {total_count - len(shown)} more."
        text_widget.config(state="normal")
        text_widget.delete("1.0", "end")
        text_widget.insert("1.0", body)
        text_widget.config(state="disabled", height=min(max(2, len(shown) + (1 if total_count > len(shown) else 0)), 6))
        frame.pack(fill="x", pady=(0, 8), before=self._table_frame)

    def _render_page(self):
        self.tree.delete(*self.tree.get_children())
        self._iid_to_row = {}
        df    = self._current_df
        page  = self._current_page
        start = page * self.PAGE_SIZE
        end   = min(start + self.PAGE_SIZE, len(df))
        disp_df = df.iloc[start:end].fillna("").astype(str)

        cols = list(df.columns)
        status_idx     = cols.index("Status") if "Status" in cols else None
        agent_code_idx = cols.index("Remark by") if "Remark by" in cols else None

        # Compare Status against the Reference list (exact, case-insensitive match)
        masterlist_lower = {m.lower() for m in self._masterlist} if self._masterlist else set()
        # Compare "Remark by" against the Agent Codes reference list (exact, case-insensitive match)
        agent_codes_lower = {a.lower() for a in self._agent_codes} if self._agent_codes else set()

        status_flagged       = 0
        status_blank         = 0
        agent_code_flagged   = 0
        agent_code_blank     = 0
        status_lines  = []
        agent_lines   = []

        for i, vals in enumerate(disp_df.values.tolist()):
            abs_row = start + i
            tag     = "even" if i % 2 == 0 else "odd"
            tags    = [tag]

            # ── Status check: blank, or not in Reference list ──────────────────
            if status_idx is not None:
                status_val = str(vals[status_idx]).strip()
                if not status_val:
                    tags.append("status_invalid")
                    status_blank += 1
                    status_lines.append(f"Row {abs_row + 1}: Status is missing")
                elif masterlist_lower and status_val.lower() not in masterlist_lower:
                    tags.append("status_invalid")
                    status_flagged += 1
                    status_lines.append(f"Row {abs_row + 1}: \"{status_val}\" not in Reference list")

            # ── "Remark by" check: blank, or not in the Agent Codes list ───────
            if agent_code_idx is not None:
                agent_val = str(vals[agent_code_idx]).strip()
                if not agent_val:
                    tags.append("agent_code_invalid")
                    agent_code_blank += 1
                    agent_lines.append(f"Row {abs_row + 1}: Remark by is missing")
                elif agent_codes_lower and agent_val.lower() not in agent_codes_lower:
                    tags.append("agent_code_invalid")
                    agent_code_flagged += 1
                    agent_lines.append(f"Row {abs_row + 1}: \"{agent_val}\" not in Agent Codes list")

            iid = self.tree.insert("", "end", values=vals, tags=tuple(tags))
            self._iid_to_row[iid] = abs_row

        self.tree.tag_configure("even",               background=CARD)
        self.tree.tag_configure("odd",                background="#14172A")
        self.tree.tag_configure("status_invalid",      background="#3B1111", foreground="#FF6B6B")
        self.tree.tag_configure("agent_code_invalid",  background="#3B2A11", foreground="#F5A623")

        total_pages = max(1, (len(df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)

        flag_note = ""
        if status_flagged:
            flag_note += f"  ·  🔴 {status_flagged} Status not in Reference"
        if status_blank:
            flag_note += f"  ·  🔴 {status_blank} Status missing"
        if agent_code_flagged:
            flag_note += f"  ·  🟠 {agent_code_flagged} Remark by not in Agent Codes"
        if agent_code_blank:
            flag_note += f"  ·  🟠 {agent_code_blank} Remark by missing"

        self.row_count_label.config(
            text=(f"Page {page + 1} of {total_pages}  ·  "
                  f"Showing rows {start + 1}–{end} of {len(df)}  ·  "
                  f"{len(df.columns)} columns  ·  ← → to navigate pages{flag_note}")
        )
        if hasattr(self, "prev_btn"):
            self.prev_btn.config(state="normal" if page > 0 else "disabled")
            self.next_btn.config(state="normal" if page < total_pages - 1 else "disabled")
            self.page_label.config(text=f"Page {page + 1} / {total_pages}")

        # ── Error textboxes: appear only while there's something to fix,
        # vanish automatically once the underlying issue is resolved. ────────
        self._set_error_box(self._status_err_frame, self._status_err_text,
                             status_lines, status_flagged + status_blank, "Status")
        self._set_error_box(self._agent_err_frame, self._agent_err_text,
                             agent_lines, agent_code_flagged + agent_code_blank, "Remark by")

    def _clear_table(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text="")
        self._current_df   = None
        self._current_page = 0
        self._iid_to_row    = {}
        if hasattr(self, "_status_err_frame"):
            self._status_err_frame.pack_forget()
            self._agent_err_frame.pack_forget()
        if hasattr(self, "prev_btn"):
            self.prev_btn.config(state="disabled")
            self.next_btn.config(state="disabled")
            self.page_label.config(text="")

    def _prev_page(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_df is not None:
            total = (len(self._current_df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE
            if self._current_page < total - 1:
                self._current_page += 1
                self._render_page()

    def _sort_col(self, col):
        df = self.imported_df
        if df is None:
            return
        asc = getattr(self, "_sort_asc", {})
        ascending = not asc.get(col, True)
        asc[col] = ascending
        self._sort_asc = asc
        try:
            df_sorted = df.sort_values(by=col, ascending=ascending,
                                        key=lambda x: x.astype(str).str.lower())
        except Exception:
            df_sorted = df
        self.imported_df = df_sorted
        self._populate_table(df_sorted, self.search_var.get())

    def _on_search(self, *_):
        if self.imported_df is not None:
            self._populate_table(self.imported_df, self.search_var.get())

    # ── COLUMN FILTER (typeable, with live typeahead) ───────────────────────
    def _filter_typeahead(self, combo, master_list, var):
        typed = var.get().strip().lower()
        if not typed or typed.startswith("—") or typed == "(all)":
            shown = master_list
        else:
            shown = [v for v in master_list if typed in str(v).lower()]
        combo["values"] = shown
        if typed and shown:
            try:
                combo.tk.call("ttk::combobox::Post", combo)
                combo.focus_set()
                combo.icursor("end")
            except Exception:
                pass

    def _on_filter_col_type(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_col_cb, self._filter_col_master, self._filter_col_var)
        if self._filter_col_var.get() in self._filter_col_master:
            self._on_filter_col_change()

    def _on_filter_val_type(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_val_cb, self._filter_val_master, self._filter_val_var)
        self._on_filter_apply()

    def _on_filter_col_change(self, _=None):
        df = self.imported_df
        if df is None:
            return
        col = self._filter_col_var.get()
        if col not in df.columns:
            return
        self._filter_val_master = ["(All)"] + sorted(df[col].fillna("").astype(str).unique().tolist())
        self._filter_val_cb["values"] = self._filter_val_master
        self._filter_val_var.set("— Select Value —")

    def _on_filter_apply(self, _=None):
        df = self.imported_df
        if df is None:
            return
        col = self._filter_col_var.get()
        val = self._filter_val_var.get().strip()
        if col not in df.columns or val in ("", "— Select Value —", "(All)"):
            self._populate_table(df, self.search_var.get())
            self._filter_active = False
            return
        filtered = df[df[col].fillna("").astype(str).str.contains(val, case=False, na=False, regex=False)].copy()
        self._filter_active = True
        self._populate_table(filtered, self.search_var.get())

    def _clear_filter(self):
        self._filter_col_var.set("— Select Column —")
        self._filter_val_var.set("— Select Value —")
        self._filter_val_master = []
        self._filter_val_cb["values"] = []
        if self._filter_col_master:
            self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
        self._filter_active = False
        if self.imported_df is not None:
            self._populate_table(self.imported_df, self.search_var.get())

    # ── FORMULA BAR ───────────────────────────────────────────────────────────
    def _get_raw_value(self, item_iid, col_name):
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return ""
        try:
            val = self._current_df.iloc[row_idx][col_name]
        except Exception:
            return ""
        if pd.isna(val):
            return ""
        return str(val)

    def _on_tree_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        item   = self.tree.identify_row(event.y)
        if not item or not col_id:
            return
        cols = self.tree["columns"]
        try:
            col_idx  = int(col_id.replace("#", "")) - 1
            col_name = cols[col_idx]
        except (ValueError, IndexError):
            return
        val = self._get_raw_value(item, col_name)
        self._update_fbar(col_name, val)

        # Single-click on a Status cell that doesn't match the Reference list
        # → open the ranked suggestion dropdown so the user can fix it inline.
        if col_name == "Status" and self._masterlist:
            masterlist_lower = {m.lower() for m in self._masterlist}
            if val.strip() and val.strip().lower() not in masterlist_lower:
                self._edit_status_cell(item)

        # Single-click on a "Remark by" cell that doesn't match the Agent
        # Codes reference list → open the ranked suggestion dropdown so the
        # user can fix it inline, scanning for the nearest available code.
        if col_name == "Remark by" and self._agent_codes:
            agent_codes_lower = {a.lower() for a in self._agent_codes}
            if val.strip() and val.strip().lower() not in agent_codes_lower:
                self._edit_agent_code_cell(item)

    def _on_cell_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        cols = self.tree["columns"]
        if not cols:
            return
        if not hasattr(self, "_fbar_last_item") or self._fbar_last_item != sel[0]:
            self._fbar_last_item = sel[0]
            val = self._get_raw_value(sel[0], cols[0])
            self._update_fbar(cols[0], val)

    def _update_fbar(self, col_name, value):
        self._fbar_col_var.set(f"  {col_name}  ")
        self._fbar_val.config(state="normal")
        self._fbar_val.delete("1.0", "end")
        self._fbar_val.insert("1.0", str(value) if value != "" else "")
        self._fbar_val.config(state="disabled")

    def _fbar_copy(self):
        val = self._fbar_val.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(val)
        show_toast(self.master, "Value copied!", color=SUCCESS, duration=1500)

    # ── ROW DETAIL POPUP (double-click) ───────────────────────────────────────
    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        item    = sel[0]
        values  = self.tree.item(item, "values")
        columns = self.tree["columns"]
        if not columns or not values:
            return
        clicked_col = None
        if event and event.x:
            col_id = self.tree.identify_column(event.x)
            try:
                col_idx  = int(col_id.replace("#", "")) - 1
                clicked_col = columns[col_idx]
            except (ValueError, IndexError):
                pass
        self._show_row_detail(columns, values, clicked_col)

    def _rank_masterlist_suggestions(self, text, limit=8):
        """Return masterlist entries ranked by similarity to `text`
        (best match first), using the same scoring as auto-correct."""
        if not self._masterlist:
            return []
        norm_text = self._normalize_remark(text) if text else ""
        scored = []
        for m in self._masterlist:
            norm_m = self._normalize_remark(m)
            if not norm_text:
                score = 0.0
            else:
                score = max(difflib.SequenceMatcher(None, norm_text, norm_m).ratio(),
                            self._partial_ratio(norm_text, norm_m))
            scored.append((score, m))
        scored.sort(key=lambda x: (-x[0], x[1].lower()))
        return [m for _score, m in scored[:limit]]

    def _edit_remark_cell(self, item_iid):
        """Inline-edit the Remark value for a row (double-click on the Remark cell).
        Shows a dropdown of the closest Reference list suggestions so you
        can pick one instead of retyping, while still allowing free text."""
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return
        orig_label  = self._current_df.iloc[row_idx].name
        current_val = self.imported_df.at[orig_label, "Remark"]
        current_val = "" if pd.isna(current_val) else str(current_val)

        win = tk.Toplevel(self)
        win.title("Edit Remark")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        tk.Label(win, text="✏️  Edit Remark", font=("Segoe UI", 11, "bold"),
                 bg=BG, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Label(win, text="Start typing to see the closest matches from the Reference list.",
                 font=("Segoe UI", 8), bg=BG, fg=SUBTEXT).pack(anchor="w", padx=16, pady=(0, 8))

        entry_var = tk.StringVar(value=current_val)
        style = ttk.Style()
        style.configure("RemarkEdit.TCombobox", font=("Segoe UI", 10))
        combo = ttk.Combobox(win, textvariable=entry_var, font=("Segoe UI", 10),
                              width=44, style="RemarkEdit.TCombobox")
        combo["values"] = self._rank_masterlist_suggestions(current_val)
        combo.pack(padx=16, pady=(0, 4), ipady=4, fill="x")
        combo.focus_set()
        combo.icursor("end")

        hint_var = tk.StringVar(value="")
        hint = tk.Label(win, textvariable=hint_var, font=("Segoe UI", 8),
                         bg=BG, fg=SUBTEXT, wraplength=360, justify="left")
        hint.pack(anchor="w", padx=16, pady=(0, 10))

        def update_hint_and_suggestions(*_):
            val = entry_var.get().strip()
            combo["values"] = self._rank_masterlist_suggestions(val)
            if not val:
                hint_var.set("")
                return
            if self._masterlist and val.lower() not in {m.lower() for m in self._masterlist}:
                hint_var.set("⚠ Not in Reference list — pick a suggestion below or it will be highlighted red.")
            else:
                hint_var.set("✓ Matches the Reference list.")

        def on_key(event=None):
            if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
                return
            update_hint_and_suggestions()
            try:
                if entry_var.get().strip() and combo["values"]:
                    combo.tk.call("ttk::combobox::Post", combo)
                    combo.focus_set()
                    combo.icursor("end")
            except Exception:
                pass

        combo.bind("<KeyRelease>", on_key)
        combo.bind("<<ComboboxSelected>>", lambda e: update_hint_and_suggestions())
        update_hint_and_suggestions()

        def save_edit():
            new_val = entry_var.get().strip().upper()
            self.imported_df.at[orig_label, "Remark"] = new_val
            win.destroy()
            self._populate_table(self.imported_df, self.search_var.get())

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Button(btn_row, text="✕  Cancel", font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text="✓  Save", font=("Segoe UI", 9, "bold"),
                  bg=SUCCESS, fg="white", relief="flat", padx=14, pady=7,
                  cursor="hand2", command=save_edit).pack(side="right")
        win.bind("<Return>", lambda e: save_edit())
        win.bind("<Escape>", lambda e: win.destroy())

    def _edit_status_cell(self, item_iid):
        """Popup with ranked Reference-list suggestions for a red (invalid) Status cell."""
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return
        orig_label  = self._current_df.iloc[row_idx].name
        current_val = self.imported_df.at[orig_label, "Status"]
        current_val = "" if pd.isna(current_val) else str(current_val)

        if not self._masterlist:
            messagebox.showinfo("Reference Empty",
                                "Add entries to the Reference list first.",
                                parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Fix Status")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(win, bg=HEADER_BG)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🛡  Fix Invalid Status",
                 font=("Segoe UI", 11, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=16, pady=10)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9),
                  bg=HEADER_BG, fg=SUBTEXT, relief="flat",
                  cursor="hand2", command=win.destroy
                  ).pack(side="right", padx=12)
        tk.Frame(win, bg=DANGER, height=2).pack(fill="x")

        # ── Current value display ──────────────────────────────────────────────
        cur_frame = tk.Frame(win, bg=CARD,
                             highlightthickness=1, highlightbackground=DANGER)
        cur_frame.pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(cur_frame, text="Current (invalid):", font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT).pack(anchor="w", padx=10, pady=(6, 0))
        tk.Label(cur_frame, text=current_val or "—", font=("Segoe UI", 10, "bold"),
                 bg=CARD, fg="#FF6B6B").pack(anchor="w", padx=10, pady=(0, 6))

        tk.Label(win, text="Select or type a replacement from the Reference list:",
                 font=("Segoe UI", 8), bg=BG, fg=SUBTEXT
                 ).pack(anchor="w", padx=16, pady=(6, 2))

        entry_var = tk.StringVar(value=current_val)
        combo = ttk.Combobox(win, textvariable=entry_var,
                              font=("Segoe UI", 10), width=44)
        combo["values"] = self._rank_masterlist_suggestions(current_val)
        combo.pack(padx=16, pady=(0, 4), ipady=4, fill="x")
        combo.focus_set()
        combo.icursor("end")

        hint_var = tk.StringVar(value="")
        hint_lbl = tk.Label(win, textvariable=hint_var, font=("Segoe UI", 8),
                             bg=BG, wraplength=360, justify="left")
        hint_lbl.pack(anchor="w", padx=16, pady=(0, 8))

        def _update(*_):
            val = entry_var.get().strip()
            combo["values"] = self._rank_masterlist_suggestions(val)
            if not val:
                hint_var.set("")
                hint_lbl.config(fg=SUBTEXT)
                return
            if val.lower() in {m.lower() for m in self._masterlist}:
                hint_var.set("✓ Matches the Reference list.")
                hint_lbl.config(fg=SUCCESS)
            else:
                hint_var.set("⚠ Still not in the Reference list.")
                hint_lbl.config(fg=WARNING)

        def _on_key(event=None):
            if event and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
                return
            _update()
            try:
                if entry_var.get().strip() and combo["values"]:
                    combo.tk.call("ttk::combobox::Post", combo)
                    combo.focus_set()
                    combo.icursor("end")
            except Exception:
                pass

        combo.bind("<KeyRelease>", _on_key)
        combo.bind("<<ComboboxSelected>>", lambda e: _update())
        _update()

        def _save():
            new_val = entry_var.get().strip()
            if not new_val:
                return
            self.imported_df.at[orig_label, "Status"] = new_val
            win.destroy()
            self._populate_table(self.imported_df, self.search_var.get())

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Button(btn_row, text="✕  Cancel", font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text="✓  Apply", font=("Segoe UI", 9, "bold"),
                  bg=SUCCESS, fg="white", relief="flat", padx=14, pady=7,
                  cursor="hand2", command=_save).pack(side="right")
        win.bind("<Return>", lambda e: _save())
        win.bind("<Escape>", lambda e: win.destroy())

        win.update_idletasks()
        win.geometry("420x300")
        px = self.winfo_rootx() + (self.winfo_width()  - 420) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 300) // 2
        win.geometry(f"420x300+{max(px,0)}+{max(py,0)}")

    def _rank_agent_code_suggestions(self, text, limit=8):
        """Return Agent Codes entries ranked by similarity to `text`
        (nearest match first), using the same scoring as the Status/Remark
        suggestion ranking."""
        if not self._agent_codes:
            return []
        norm_text = self._normalize_remark(text) if text else ""
        scored = []
        for a in self._agent_codes:
            norm_a = self._normalize_remark(a)
            if not norm_text:
                score = 0.0
            else:
                score = max(difflib.SequenceMatcher(None, norm_text, norm_a).ratio(),
                            self._partial_ratio(norm_text, norm_a))
            scored.append((score, a))
        scored.sort(key=lambda x: (-x[0], x[1].lower()))
        return [a for _score, a in scored[:limit]]

    def _edit_agent_code_cell(self, item_iid):
        """Popup with ranked Agent Codes suggestions for a "Remark by" value
        not found in the Agent Codes reference list — scans the list for
        the nearest usable code instead of just flagging it."""
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return
        orig_label  = self._current_df.iloc[row_idx].name
        current_val = self.imported_df.at[orig_label, "Remark by"]
        current_val = "" if pd.isna(current_val) else str(current_val)

        if not self._agent_codes:
            messagebox.showinfo("Agent Codes Empty",
                                "Add entries to the Agent Codes list first.",
                                parent=self)
            return

        win = tk.Toplevel(self)
        win.title("Fix Agent Code")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(win, bg=HEADER_BG)
        hdr.pack(fill="x")
        tk.Label(hdr, text="🧑‍💼  Fix Invalid Agent Code",
                 font=("Segoe UI", 11, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=16, pady=10)
        tk.Button(hdr, text="✕", font=("Segoe UI", 9),
                  bg=HEADER_BG, fg=SUBTEXT, relief="flat",
                  cursor="hand2", command=win.destroy
                  ).pack(side="right", padx=12)
        tk.Frame(win, bg=WARNING, height=2).pack(fill="x")

        # ── Current value display ──────────────────────────────────────────────
        cur_frame = tk.Frame(win, bg=CARD,
                             highlightthickness=1, highlightbackground=WARNING)
        cur_frame.pack(fill="x", padx=16, pady=(12, 4))
        tk.Label(cur_frame, text="Current (not in Agent Codes):", font=("Segoe UI", 8),
                 bg=CARD, fg=SUBTEXT).pack(anchor="w", padx=10, pady=(6, 0))
        tk.Label(cur_frame, text=current_val or "—", font=("Segoe UI", 10, "bold"),
                 bg=CARD, fg="#F5A623").pack(anchor="w", padx=10, pady=(0, 6))

        tk.Label(win, text="Nearest matches found in the Agent Codes list:",
                 font=("Segoe UI", 8), bg=BG, fg=SUBTEXT
                 ).pack(anchor="w", padx=16, pady=(6, 2))

        entry_var = tk.StringVar(value=current_val)
        combo = ttk.Combobox(win, textvariable=entry_var,
                              font=("Segoe UI", 10), width=44)
        combo["values"] = self._rank_agent_code_suggestions(current_val)
        combo.pack(padx=16, pady=(0, 4), ipady=4, fill="x")
        combo.focus_set()
        combo.icursor("end")

        hint_var = tk.StringVar(value="")
        hint_lbl = tk.Label(win, textvariable=hint_var, font=("Segoe UI", 8),
                             bg=BG, wraplength=360, justify="left")
        hint_lbl.pack(anchor="w", padx=16, pady=(0, 8))

        def _update(*_):
            val = entry_var.get().strip()
            combo["values"] = self._rank_agent_code_suggestions(val)
            if not val:
                hint_var.set("")
                hint_lbl.config(fg=SUBTEXT)
                return
            if val.lower() in {a.lower() for a in self._agent_codes}:
                hint_var.set("✓ Matches the Agent Codes list.")
                hint_lbl.config(fg=SUCCESS)
            else:
                hint_var.set("⚠ Still not in the Agent Codes list.")
                hint_lbl.config(fg=WARNING)

        def _on_key(event=None):
            if event and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
                return
            _update()
            try:
                if entry_var.get().strip() and combo["values"]:
                    combo.tk.call("ttk::combobox::Post", combo)
                    combo.focus_set()
                    combo.icursor("end")
            except Exception:
                pass

        combo.bind("<KeyRelease>", _on_key)
        combo.bind("<<ComboboxSelected>>", lambda e: _update())
        _update()

        def _save():
            new_val = entry_var.get().strip()
            if not new_val:
                return
            self.imported_df.at[orig_label, "Remark by"] = new_val
            win.destroy()
            self._populate_table(self.imported_df, self.search_var.get())

        btn_row = tk.Frame(win, bg=BG)
        btn_row.pack(fill="x", padx=16, pady=(0, 14))
        tk.Button(btn_row, text="✕  Cancel", font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat", padx=14, pady=7,
                  cursor="hand2", command=win.destroy).pack(side="right", padx=(6, 0))
        tk.Button(btn_row, text="✓  Apply", font=("Segoe UI", 9, "bold"),
                  bg=SUCCESS, fg="white", relief="flat", padx=14, pady=7,
                  cursor="hand2", command=_save).pack(side="right")
        win.bind("<Return>", lambda e: _save())
        win.bind("<Escape>", lambda e: win.destroy())

        win.update_idletasks()
        win.geometry("420x300")
        px = self.winfo_rootx() + (self.winfo_width()  - 420) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 300) // 2
        win.geometry(f"420x300+{max(px,0)}+{max(py,0)}")

    def _show_row_detail(self, columns, values, clicked_col=None):
        popup = tk.Toplevel(self)
        popup.title("Row Details")
        popup.configure(bg=BG)
        popup.resizable(True, True)
        popup.grab_set()

        hdr = tk.Frame(popup, bg=HEADER_BG, height=48)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔍  Row Details",
                 font=("Segoe UI", 12, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text=f"{len(columns)} fields",
                 font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT
                 ).pack(side="left")
        tk.Button(hdr, text="✕  Close",
                  font=("Segoe UI", 8, "bold"),
                  bg=CARD, fg=TEXT, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=popup.destroy
                  ).pack(side="right", padx=12, pady=8)
        tk.Frame(popup, bg=ACCENT2, height=2).pack(fill="x")

        container = tk.Frame(popup, bg=BG)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg=BG, highlightthickness=0, bd=0)
        vsb2   = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner  = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>",  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))

        def _mw(e): canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        popup.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        col_data = list(zip(columns, values))
        for i, (col, val) in enumerate(col_data):
            is_hl  = (col == clicked_col)
            row_bg = ACCENT2 if is_hl else (CARD if i % 2 == 0 else "#14172A")
            lbl_fg = "white"  if is_hl else SUBTEXT
            val_fg = "white"  if is_hl else TEXT

            row = tk.Frame(inner, bg=row_bg)
            row.pack(fill="x", padx=12, pady=1)

            tk.Label(row, text=f"{i+1:>3}", font=("Consolas", 8),
                     bg=row_bg, fg=SUBTEXT, width=3, anchor="e"
                     ).pack(side="left", padx=(8, 4), pady=6)

            tk.Label(row, text=col, font=("Segoe UI", 9, "bold"),
                     bg=row_bg, fg=lbl_fg, width=28, anchor="w"
                     ).pack(side="left", padx=(0, 10), pady=6)

            tk.Frame(row, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)

            val_str  = str(val) if val != "" else "—"
            vt = tk.Text(row, font=("Segoe UI", 9),
                         bg=row_bg, fg=val_fg,
                         relief="flat", bd=0, height=1, wrap="none",
                         cursor="xterm",
                         selectbackground=ACCENT2, selectforeground="white")
            vt.insert("1.0", val_str)
            if len(val_str) > 80:
                vt.config(height=min((len(val_str) // 80) + 1, 4), wrap="word")
            vt.config(state="disabled")
            vt.pack(side="left", fill="x", expand=True, padx=(10, 8), pady=4)

        footer = tk.Frame(popup, bg=DARK)
        footer.pack(fill="x", side="bottom")

        def _copy_all():
            popup.clipboard_clear()
            popup.clipboard_append("\n".join(f"{c}\t{v}" for c, v in col_data))
            show_toast(popup, "All fields copied!", color=SUCCESS)

        tk.Button(footer, text="📋  Copy All Fields",
                  font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat",
                  padx=16, pady=7, cursor="hand2",
                  command=_copy_all).pack(side="left", padx=12, pady=8)
        tk.Label(footer, text="Double-click any row to view full details.",
                 font=("Segoe UI", 8), bg=DARK, fg=SUBTEXT
                 ).pack(side="right", padx=16)

        popup.update_idletasks()
        popup.geometry("700x560")
        px = self.winfo_rootx() + (self.winfo_width()  - 700) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 560) // 2
        popup.geometry(f"700x560+{max(px,0)}+{max(py,0)}")


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN — ANALYTICAL REPORTS (standalone tool)
# ══════════════════════════════════════════════════════════════════════════════
class _AnalyticsGridPanel(tk.Frame):
    """Reusable grid panel: browse/load an Excel file, then search, filter,
    sort, paginate, inspect rows, and export. Used for both tabs in the
    Analytics screen — Field Result (RESULT sheet, ORICO AUTO LOAN filter)
    and Masterfile (Masterlist headers, no bank filter)."""

    PAGE_SIZE = 2000

    def __init__(self, master, color, drop_text, sheet_picker,
                 bank_filter_col=None, bank_filter_value=None,
                 default_headers=None, export_filename="analytical_view.xlsx",
                 empty_msg="No data loaded. Browse a file on the left to get started.",
                 usecols=None, force_date_columns=None,
                 visit_count_col=None, visit_count_key=None,
                 visit_status_col=None,
                 field_result_panel=None, on_data_changed=None,
                 coverage_status_col=None, coverage_status_source_col=None):
        super().__init__(master, bg=BG)
        self.color              = color
        self.drop_text          = drop_text
        self.sheet_picker       = sheet_picker          # fn(sheet_names) -> sheet name (raises ValueError if not found)
        self.bank_filter_col    = bank_filter_col
        self.bank_filter_value  = bank_filter_value
        self.default_headers    = default_headers or []
        self.export_filename    = export_filename
        self.empty_msg          = empty_msg
        self.usecols            = usecols   # e.g. "D:R" — restrict which Excel columns are read
        self.force_date_columns = force_date_columns or []  # column names always normalized to MM/dd/yyyy text
        self.visit_count_col    = visit_count_col   # name of the computed "visits" column, e.g. "No. of Visits"
        self.visit_status_col   = visit_status_col  # name of the "Visited"/"Not Visited" column, derived from visit_count_col
        self.visit_count_key    = visit_count_key   # column to match on, e.g. "CHCODE"
        self.field_result_panel = field_result_panel  # the other tab's panel, read for cross-tab counting
        self.on_data_changed    = on_data_changed   # callable(self) fired after a successful load
        self.coverage_status_col        = coverage_status_col         # name of computed column, e.g. "Coverage Status"
        self.coverage_status_source_col = coverage_status_source_col   # column to look up, e.g. "STATE"
        self.file_path      = None
        self.df             = None   # full raw dataframe, exactly as loaded
        self._current_df    = None   # after search/filter, before pagination
        self._current_page  = 0
        self._iid_to_row     = {}
        self._filter_active  = False
        self._filter_col_master = []
        self._filter_val_master = []
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        # ── Body ──────────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # ── LEFT PANEL ────────────────────────────────────────────────────────
        left = tk.Frame(body, bg=BG, width=280)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        self._drop_frame = tk.Frame(left, bg=CARD, highlightthickness=2,
                                    highlightbackground=BORDER)
        self._drop_frame.pack(fill="x", pady=(0, 10))

        drop_inner = tk.Frame(self._drop_frame, bg=CARD)
        drop_inner.pack(fill="both", padx=20, pady=20)

        tk.Label(drop_inner, text="📂", font=("Segoe UI", 26),
                 bg=CARD, fg=self.color).pack()
        self._drop_title = tk.Label(drop_inner, text=self.drop_text,
                                    font=("Segoe UI", 10, "bold"), bg=CARD, fg=TEXT,
                                    wraplength=200, justify="center")
        self._drop_title.pack(pady=(6, 2))
        tk.Label(drop_inner, text=".xlsx  ·  .xlsm  ·  .xls",
                 font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack()
        tk.Button(drop_inner, text="📁  Browse File", font=("Segoe UI", 9, "bold"),
                  bg=self.color, fg="white", relief="flat", padx=16, pady=6,
                  cursor="hand2", command=self._browse).pack(pady=(10, 0))

        self._fi_name = tk.Label(left, text="", font=("Segoe UI", 9, "bold"),
                                 bg=BG, fg=TEXT, wraplength=260, justify="left", anchor="w")
        self._fi_name.pack(fill="x", pady=(4, 0))

        self.load_btn = tk.Button(left, text="📊  Load Data",
                                  font=("Segoe UI", 10, "bold"),
                                  bg=self.color, fg="white",
                                  relief="flat", pady=10, cursor="hand2",
                                  state="disabled", command=self._run_load)
        self.load_btn.pack(fill="x", pady=(14, 2))
        tk.Label(left, text="Shows every column and row exactly as in the file",
                 font=("Segoe UI", 7), bg=BG, fg=SUBTEXT).pack(anchor="w")

        # ── Quick stats (rows / columns) ─────────────────────────────────────
        stats_row = tk.Frame(left, bg=BG)
        stats_row.pack(fill="x", pady=(12, 0))

        def _mini_stat(parent, label, color):
            c = tk.Frame(parent, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
            c.pack(side="left", fill="both", expand=True, padx=(0, 6))
            inn = tk.Frame(c, bg=CARD)
            inn.pack(fill="both", padx=10, pady=8)
            var = tk.StringVar(value="—")
            tk.Label(inn, textvariable=var, font=("Segoe UI", 15, "bold"),
                     bg=CARD, fg=color).pack(anchor="w")
            tk.Label(inn, text=label, font=("Segoe UI", 7, "bold"),
                     bg=CARD, fg=SUBTEXT).pack(anchor="w")
            return var

        self._rows_var = _mini_stat(stats_row, "ROWS",    TEXT)
        self._cols_var = _mini_stat(stats_row, "COLUMNS", self.color)

        style = ttk.Style()
        style.configure("Report.Horizontal.TProgressbar", troughcolor=CARD, background=self.color, thickness=4)
        self.progress = ttk.Progressbar(left, mode="indeterminate", style="Report.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(10, 0))

        self.status_var = tk.StringVar(value="Ready — load a file to begin.")
        tk.Label(left, textvariable=self.status_var, font=("Segoe UI", 8),
                 bg=BG, fg=SUBTEXT, wraplength=260, justify="left", anchor="w"
                 ).pack(fill="x", pady=(10, 0))

        self.download_btn = tk.Button(left, text="⬇  Export View",
                                      font=("Segoe UI", 9, "bold"),
                                      bg=CARD, fg=SUBTEXT,
                                      relief="flat", pady=8, cursor="hand2",
                                      state="disabled", command=self._download)
        self.download_btn.pack(fill="x", pady=(10, 0))

        # ── RIGHT PANEL — search, filter, grid ───────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Search bar
        search_frame = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        search_frame.pack(fill="x", pady=(0, 4))
        tk.Label(search_frame, text="🔍", bg=CARD, fg=SUBTEXT, font=("Segoe UI", 10)).pack(side="left", padx=8)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 9),
                 bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=0
                 ).pack(side="left", fill="x", expand=True, pady=8, padx=4)
        tk.Label(search_frame, text="Search all columns…", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        # Column dropdown filter
        filter_outer = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        filter_outer.pack(fill="x", pady=(0, 8))
        tk.Label(filter_outer, text="▼ Column Filter:", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=8, pady=6)
        self._filter_col_var = tk.StringVar(value="— Select Column —")
        self._filter_val_var = tk.StringVar(value="— Select Value —")
        self._filter_col_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_col_var,
                                             width=24, font=("Segoe UI", 9))
        self._filter_col_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_val_cb  = ttk.Combobox(filter_outer, textvariable=self._filter_val_var,
                                             width=28, font=("Segoe UI", 9))
        self._filter_val_cb.pack(side="left", padx=(0, 6), pady=6)
        self._filter_col_cb.bind("<<ComboboxSelected>>", self._on_filter_col_change)
        self._filter_val_cb.bind("<<ComboboxSelected>>", self._on_filter_apply)
        self._filter_col_cb.bind("<KeyRelease>", self._on_filter_col_type)
        self._filter_val_cb.bind("<KeyRelease>", self._on_filter_val_type)
        tk.Button(filter_outer, text="✕ Clear Filter", font=("Segoe UI", 8, "bold"),
                  bg=DANGER, fg="white", relief="flat", padx=10, pady=4,
                  cursor="hand2", command=self._clear_filter).pack(side="left", padx=(0, 6))

        # Grid
        table_frame = tk.Frame(right, bg=BG)
        table_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style.configure("Custom.Treeview", background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28, font=("Segoe UI", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading", background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT2)], foreground=[("selected", "white")])

        self.tree = ttk.Treeview(table_frame, style="Custom.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                  show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self._on_cell_click)
        self.tree.bind("<ButtonRelease-1>",  self._on_tree_click)
        self.tree.bind("<Double-1>",         self._on_double_click)
        self.tree.bind("<Return>",           self._on_double_click)

        self.row_count_label = tk.Label(right, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.row_count_label.pack(anchor="e", pady=(4, 0))

        # ── Excel-style formula bar ───────────────────────────────────────────
        fbar = tk.Frame(right, bg="#1C1F2E",
                        highlightthickness=1, highlightbackground=BORDER)
        fbar.pack(fill="x", pady=(4, 2))

        tk.Label(fbar, text=" fx ", font=("Segoe UI", 9, "bold"),
                 bg="#2E7D32", fg="white", padx=6, pady=4
                 ).pack(side="left")

        self._fbar_col_var = tk.StringVar(value="")
        col_box = tk.Entry(fbar, textvariable=self._fbar_col_var,
                           font=("Segoe UI", 9, "bold"),
                           bg="#252840", fg=ACCENT,
                           relief="flat", bd=0,
                           width=22, justify="center",
                           insertbackground=ACCENT,
                           readonlybackground="#252840",
                           state="readonly")
        col_box.pack(side="left", padx=(1, 0), ipady=4)

        tk.Frame(fbar, bg=BORDER, width=1).pack(side="left", fill="y", pady=2)

        self._fbar_val = tk.Text(fbar, font=("Segoe UI", 9),
                                 bg="#1C1F2E", fg=TEXT,
                                 relief="flat", bd=0,
                                 height=1, wrap="none",
                                 cursor="xterm",
                                 selectbackground=ACCENT,
                                 selectforeground="white",
                                 insertbackground=TEXT,
                                 state="disabled")
        fbar_xsb = ttk.Scrollbar(fbar, orient="horizontal",
                                  command=self._fbar_val.xview)
        self._fbar_val.configure(xscrollcommand=fbar_xsb.set)
        self._fbar_val.pack(side="left", fill="x", expand=True, padx=(6, 0), pady=2)
        tk.Button(fbar, text="⎘", font=("Segoe UI", 10),
                  bg="#1C1F2E", fg=SUBTEXT, relief="flat",
                  cursor="hand2", padx=6,
                  command=self._fbar_copy
                  ).pack(side="right", padx=4)

        nav_frame = tk.Frame(right, bg=BG)
        nav_frame.pack(fill="x", pady=(4, 0))
        self.prev_btn = tk.Button(nav_frame, text="◀  Prev", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._prev_page)
        self.prev_btn.pack(side="left", padx=(0, 6))
        self.page_label = tk.Label(nav_frame, text="", font=("Segoe UI", 8), bg=BG, fg=SUBTEXT)
        self.page_label.pack(side="left")
        self.next_btn = tk.Button(nav_frame, text="Next  ▶", font=("Segoe UI", 8, "bold"),
                                   bg=CARD, fg=TEXT, relief="flat", padx=12, pady=4,
                                   cursor="hand2", state="disabled", command=self._next_page)
        self.next_btn.pack(side="left", padx=6)

        self._show_initial_state()

    def _show_empty_message(self, msg):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text=msg)

    def _show_initial_state(self):
        """On first build (and after Clear), show either the plain empty
        message, or — for tabs configured with default_headers (e.g. the
        Masterfile tab) — a ready-to-import template with those header
        columns already laid out, so the expected layout is visible before
        anything is loaded."""
        if self.default_headers:
            self.tree.delete(*self.tree.get_children())
            self.tree["columns"] = self.default_headers
            for col in self.default_headers:
                self.tree.heading(col, text=col, anchor="w")
                self.tree.column(col, width=max(100, min(len(col) * 9, 220)),
                                  minwidth=60, anchor="w", stretch=False)
            self.row_count_label.config(
                text=f"Template ready — {len(self.default_headers)} Masterlist columns. "
                     f"Browse a file on the left to import data."
            )
        else:
            self._show_empty_message(self.empty_msg)

    # ── FILE LOAD ─────────────────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(title="Select Excel file",
                                          filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")])
        if path:
            self._load_file(path)

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        self.file_path = path
        fname = os.path.basename(path)
        self._drop_title.config(text="File ready to load", fg=self.color)
        self._drop_frame.config(highlightbackground=self.color)
        self._fi_name.config(text=f"📄  {fname}")
        self.load_btn.config(state="normal")
        self.status_var.set(f"Loaded: {fname}  —  Click Load Data to view it.")
        self._show_empty_message("Click “Load Data” to display every row and column.")

    # ── LOAD & DISPLAY (every column, every row, no transformation) ─────────
    def _run_load(self):
        if not self.file_path:
            return
        self.load_btn.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set("Reading file…")
        self.progress.start(10)
        threading.Thread(target=self._load_thread, daemon=True).start()

    @staticmethod
    def _reformat_date_columns(df, raw):
        """DRR-style date cleanup: any column that Excel actually stored as
        a date/time value gets rewritten as plain text — dates become
        MM/DD/YYYY with no HH:MM:SS, columns that are genuinely time-of-day
        (name contains TIME but not DATE) become hh:mm:ss AM/PM. Columns
        that were already plain text in the file are left untouched."""
        for col in raw.columns:
            if col not in df.columns or not pd.api.types.is_datetime64_any_dtype(raw[col]):
                continue
            name_upper = str(col).strip().upper()
            is_time_only = "TIME" in name_upper and "DATE" not in name_upper
            if is_time_only:
                formatted = raw[col].dt.strftime("%I:%M:%S %p").str.lstrip("0")
            else:
                formatted = raw[col].dt.strftime("%m/%d/%Y")
            formatted = formatted.where(raw[col].notna(), "").reset_index(drop=True)
            df[col] = formatted.values
        return df

    def _force_date_text(self, df, raw):
        """Always normalize specific columns (e.g. "ENDO DATE") to plain
        MM/dd/yyyy text — matching the raw file's date value exactly, with
        no time component — regardless of whether Excel stored the cell as
        a real date or as text (e.g. "5/2/2024 0:00:00", "2024-05-02").
        Values that can't be parsed as a date are left exactly as found."""
        for target in self.force_date_columns:
            col = next(
                (c for c in df.columns if str(c).strip().upper() == target.strip().upper()),
                None
            )
            if col is None:
                continue
            source = raw[col] if col in raw.columns else df[col]
            if pd.api.types.is_datetime64_any_dtype(source):
                parsed = pd.to_datetime(source, errors="coerce")
            else:
                parsed = pd.to_datetime(source.astype(str), errors="coerce", format="mixed")
            formatted = parsed.dt.strftime("%m/%d/%Y")
            original = df[col].reset_index(drop=True)
            formatted = formatted.reset_index(drop=True)
            df[col] = formatted.where(parsed.reset_index(drop=True).notna(), original).values
        return df

    def _find_col(self, df, name):
        if df is None:
            return None
        return next(
            (c for c in df.columns if str(c).strip().upper() == name.strip().upper()), None
        )

    def _compute_visit_counts_column(self):
        """Add/refresh the visit-count column: for each Masterfile row, how
        many times that row's CHCODE actually appears in the Field Result
        tab's data. A code that never shows up in Field Result gets 0 —
        the Masterfile row itself is never counted as a "visit", so a code
        with no Field Result match never shows a false 1."""
        if not self.visit_count_col or self.df is None or self.df.empty:
            return
        key_col = self._find_col(self.df, self.visit_count_key)
        if key_col is None:
            return
        own_codes = self.df[key_col].astype(str).str.strip().str.upper()

        fr_panel = self.field_result_panel
        fr_counts = None
        if fr_panel is not None and fr_panel.df is not None and not fr_panel.df.empty:
            fr_key_col = self._find_col(fr_panel.df, self.visit_count_key)
            if fr_key_col is not None:
                fr_codes = fr_panel.df[fr_key_col].astype(str).str.strip().str.upper()
                fr_counts = fr_codes.value_counts()

        if fr_counts is not None:
            self.df[self.visit_count_col] = own_codes.map(fr_counts).fillna(0).astype(int)
        else:
            self.df[self.visit_count_col] = 0

        if self.visit_status_col:
            self.df[self.visit_status_col] = self.df[self.visit_count_col].apply(
                lambda n: "Visited" if n > 0 else "Not Visited"
            )

    def recompute_visit_counts(self):
        """Public hook: recompute and redraw the visit-count column without
        re-importing a file. Called when the OTHER tab (Field Result)
        finishes loading new data, so visit counts here stay current."""
        if not self.visit_count_col or self.df is None or self.df.empty:
            return
        self._compute_visit_counts_column()
        self._populate_table(self.df, self.search_var.get())

    def _compute_coverage_status_column(self):
        """Add/refresh the computed coverage-status column: for each row,
        look up this tab's STATE value against the saved Geo Reference
        table (Geo Reference tab / geo_reference DB table) — first by
        PROVINCE, falling back to MUNICIPALITY if there's no province
        match — and pull that Geo Reference row's AREA STATUS (Covered /
        Not Covered / Special Field).

        Mirrors the Excel formula:
        =XLOOKUP(state, AreaRef[PROVINCE], AreaRef[AREA STATUS],
                 XLOOKUP(state, AreaRef[MUNICIPALITY], AreaRef[AREA STATUS], , 0), 0)
        A STATE value with no match in either column is left blank, same
        as the formula's fallback.
        """
        if not self.coverage_status_col or self.df is None or self.df.empty:
            return
        source_col = self._find_col(self.df, self.coverage_status_source_col)
        if source_col is None:
            return

        # Build PROVINCE -> AREA STATUS and MUNICIPALITY -> AREA STATUS
        # lookup maps from the saved Geo Reference rows. First occurrence
        # wins for a given key, matching XLOOKUP's default "first match".
        province_map, municipality_map = {}, {}
        for row in db_list_geo_reference():
            # row = (id, unique_code, province, municipality, final_area,
            #        geocode, cluster, area_status, created_at)
            province     = str(row[2]).strip().upper()
            municipality = str(row[3]).strip().upper()
            area_status  = row[7]
            if province and province not in province_map:
                province_map[province] = area_status
            if municipality and municipality not in municipality_map:
                municipality_map[municipality] = area_status

        def _lookup(val):
            key = str(val).strip().upper()
            if not key:
                return ""
            if key in province_map:
                return province_map[key]
            if key in municipality_map:
                return municipality_map[key]
            return ""

        values = self.df[source_col].apply(_lookup)
        if self.coverage_status_col in self.df.columns:
            # Already present (e.g. a recompute) — update in place, keep its position.
            self.df[self.coverage_status_col] = values
        else:
            # First time adding it — insert right next to STATE instead of
            # tacking it onto the far-right end of the table, so it's easy
            # to spot without scrolling all the way across.
            insert_pos = self.df.columns.get_loc(source_col) + 1
            self.df.insert(insert_pos, self.coverage_status_col, values)

    def recompute_coverage_status(self):
        """Public hook: recompute and redraw the coverage-status column
        without re-importing a file. Called when the Geo Reference tab's
        saved data changes (add/edit/delete/import/clear), so this stays
        current."""
        if not self.coverage_status_col or self.df is None or self.df.empty:
            return
        self._compute_coverage_status_column()
        self._populate_table(self.df, self.search_var.get())

    def _load_thread(self):
        try:
            xl = pd.ExcelFile(self.file_path)
            # Which sheet to read is decided by this tab's sheet_picker
            # (e.g. Field Result requires a "RESULT" sheet; Masterfile
            # accepts a "MASTERLIST"/"MASTERFILE" sheet or just the first
            # sheet in the workbook).
            sheet_name = self.sheet_picker(xl.sheet_names)

            # First pass with auto dtype detection — this is the only way
            # to know which columns are real Excel date/time cells (so we
            # can reformat just those), before everything gets cast to str.
            raw = xl.parse(sheet_name, usecols=self.usecols)
            # Second pass with dtype=str preserves every other column
            # exactly as typed: no lost leading zeros, no auto-appended
            # ".0" on ID-style numbers.
            df = xl.parse(sheet_name, dtype=str, usecols=self.usecols)
            df = df.fillna("")
            df = self._reformat_date_columns(df, raw)
            df = self._force_date_text(df, raw)

            # Reorder/align columns to the expected header template, if
            # this tab has one (e.g. the Masterfile tab's 15 columns).
            # Any expected header missing from the file is added as an
            # empty column; any extra columns in the file are kept and
            # appended after the template columns.
            if self.default_headers:
                file_cols_norm = {str(c).strip().upper(): c for c in df.columns}
                ordered = []
                for h in self.default_headers:
                    actual = file_cols_norm.get(h.strip().upper())
                    if actual is not None:
                        ordered.append(actual)
                    else:
                        df[h] = ""
                        ordered.append(h)
                extra = [c for c in df.columns if c not in ordered]
                df = df[ordered + extra]

            total_before = total_after = None
            # Optional bank filter (Field Result only): keep only rows
            # belonging to a given bank/account — matched case-insensitively
            # and trimmed, so "Orico Autoloan" / " ORICO AUTOLOAN " etc.
            # still count.
            if self.bank_filter_col:
                bank_col = next(
                    (c for c in df.columns if str(c).strip().upper() == self.bank_filter_col),
                    None
                )
                if bank_col is None:
                    raise ValueError(
                        f"No \"{self.bank_filter_col}\" column found in the {sheet_name} sheet. "
                        f"Cannot filter for {self.bank_filter_value}."
                    )
                total_before = len(df)
                bank_norm = df[bank_col].astype(str).str.strip().str.upper()
                df = df[bank_norm == self.bank_filter_value].reset_index(drop=True)
                total_after = len(df)

            self.after(0, lambda: self._on_loaded(df, total_before, total_after))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _on_loaded(self, df, total_before=None, total_after=None):
        self.progress.stop()
        self.df = df
        try:
            self._compute_visit_counts_column()
        except Exception as e:
            print(f"[visit-count] skipped: {e}")
        try:
            self._compute_coverage_status_column()
        except Exception as e:
            print(f"[coverage-status] skipped: {e}")
        self.load_btn.config(state="normal")
        self.download_btn.config(state="normal" if not df.empty else "disabled")
        self._rows_var.set(f"{len(df):,}")
        self._cols_var.set(str(len(df.columns)))
        if total_before is not None and total_after is not None:
            dropped = total_before - total_after
            self.status_var.set(
                f"Done  ·  {total_after:,} of {total_before:,} rows are "
                f"{self.bank_filter_value}  ·  {dropped:,} other rows excluded  ·  "
                f"{len(df.columns)} columns."
            )
        else:
            self.status_var.set(
                f"Done  ·  {len(df):,} rows  ·  {len(df.columns)} columns loaded."
            )
        self._filter_col_master = list(df.columns)
        self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
        self._filter_col_var.set("— Select Column —")
        self._filter_val_var.set("— Select Value —")
        self._filter_val_master = []
        self._filter_val_cb["values"] = []
        self._filter_active = False
        self.search_var.set("")
        self._populate_table(df)
        show_toast(self.master, "Data loaded!", color=self.color)
        if self.on_data_changed:
            self.on_data_changed(self)

    def _on_error(self, msg):
        self.progress.stop()
        self.load_btn.config(state="normal")
        self.status_var.set(f"Error: {msg}")
        messagebox.showerror("Load Error", msg)

    # ── TABLE ─────────────────────────────────────────────────────────────────
    def _populate_table(self, df, filter_text=""):
        self._clear_table()
        if df is None or df.empty:
            self.row_count_label.config(text="No rows to display.")
            return
        cols = list(df.columns)
        self.tree["columns"] = cols
        for col in cols:
            sample    = df[col].astype(str)
            max_len   = max(len(str(col)), int(sample.str.len().quantile(0.90)) if len(df) > 0 else 10)
            col_width = max(100, min(max_len * 8, 280))
            self.tree.heading(col, text=col, anchor="w",
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=col_width, minwidth=60, anchor="w", stretch=False)
        filt = filter_text.lower().strip()
        if filt:
            combined = df.fillna("").astype(str).agg(" ".join, axis=1).str.lower()
            filtered = df[combined.str.contains(filt, regex=False)].reset_index(drop=True)
        else:
            filtered = df.reset_index(drop=True)
        self._current_df   = filtered
        self._current_page = 0
        self._render_page()

    def _render_page(self):
        self.tree.delete(*self.tree.get_children())
        self._iid_to_row = {}
        df    = self._current_df
        page  = self._current_page
        start = page * self.PAGE_SIZE
        end   = min(start + self.PAGE_SIZE, len(df))
        str_df = df.iloc[start:end].fillna("").astype(str)
        for i, vals in enumerate(str_df.values.tolist()):
            tag = "even" if i % 2 == 0 else "odd"
            iid = self.tree.insert("", "end", values=vals, tags=(tag,))
            self._iid_to_row[iid] = start + i
        self.tree.tag_configure("even", background=CARD)
        self.tree.tag_configure("odd",  background="#14172A")
        total_pages = max(1, (len(df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE)
        self.row_count_label.config(
            text=(f"Page {page + 1} of {total_pages}  ·  "
                  f"Showing rows {start + 1}–{end} of {len(df)}  ·  "
                  f"{len(df.columns)} columns  ·  ← → to navigate pages")
        )
        self.prev_btn.config(state="normal" if page > 0 else "disabled")
        self.next_btn.config(state="normal" if page < total_pages - 1 else "disabled")
        self.page_label.config(text=f"Page {page + 1} / {total_pages}")

    def _prev_page(self):
        if self._current_page > 0:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_df is not None:
            total = (len(self._current_df) + self.PAGE_SIZE - 1) // self.PAGE_SIZE
            if self._current_page < total - 1:
                self._current_page += 1
                self._render_page()

    def _on_search(self, *_):
        if self.df is not None:
            self._populate_table(self.df, self.search_var.get())

    def _sort_col(self, col):
        if self.df is None or col not in self.df.columns:
            return
        asc = getattr(self, "_sort_asc", {})
        ascending = not asc.get(col, True)
        asc[col] = ascending
        self._sort_asc = asc
        try:
            self.df = self.df.sort_values(by=col, ascending=ascending,
                                           key=lambda x: x.astype(str).str.lower())
        except Exception:
            pass
        self._populate_table(self.df, self.search_var.get())

    # ── COLUMN FILTER ─────────────────────────────────────────────────────────
    def _filter_typeahead(self, combo, master_list, var):
        typed = var.get().strip().lower()
        if not typed or typed.startswith("—") or typed == "(all)":
            shown = master_list
        else:
            shown = [v for v in master_list if typed in str(v).lower()]
        combo["values"] = shown
        if typed and shown:
            try:
                combo.tk.call("ttk::combobox::Post", combo)
                combo.focus_set()
                combo.icursor("end")
            except Exception:
                pass

    def _on_filter_col_type(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_col_cb, self._filter_col_master, self._filter_col_var)
        if self._filter_col_var.get() in self._filter_col_master:
            self._on_filter_col_change()

    def _on_filter_val_type(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Escape", "Tab"):
            return
        self._filter_typeahead(self._filter_val_cb, self._filter_val_master, self._filter_val_var)
        self._on_filter_apply()

    def _on_filter_col_change(self, _=None):
        if self.df is None:
            return
        col = self._filter_col_var.get()
        if col not in self.df.columns:
            return
        self._filter_val_master = ["(All)"] + sorted(self.df[col].fillna("").astype(str).unique().tolist())
        self._filter_val_cb["values"] = self._filter_val_master
        self._filter_val_var.set("— Select Value —")

    def _on_filter_apply(self, _=None):
        if self.df is None:
            return
        col = self._filter_col_var.get()
        val = self._filter_val_var.get().strip()
        if col not in self.df.columns or val in ("", "— Select Value —", "(All)"):
            self._populate_table(self.df, self.search_var.get())
            self._filter_active = False
            return
        filtered = self.df[self.df[col].fillna("").astype(str).str.contains(val, case=False, na=False, regex=False)].copy()
        self._filter_active = True
        self._populate_table(filtered, self.search_var.get())

    def _clear_filter(self):
        self._filter_col_var.set("— Select Column —")
        self._filter_val_var.set("— Select Value —")
        self._filter_val_master = []
        self._filter_val_cb["values"] = []
        if self._filter_col_master:
            self._filter_col_cb["values"] = ["— Select Column —"] + self._filter_col_master
        self._filter_active = False
        if self.df is not None:
            self._populate_table(self.df, self.search_var.get())

    def _clear_table(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text="")
        self._current_df   = None
        self._current_page = 0
        self._iid_to_row    = {}
        self.prev_btn.config(state="disabled")
        self.next_btn.config(state="disabled")
        self.page_label.config(text="")

    # ── FORMULA BAR ───────────────────────────────────────────────────────────
    def _get_raw_value(self, item_iid, col_name):
        row_idx = self._iid_to_row.get(item_iid)
        if row_idx is None or self._current_df is None:
            return ""
        try:
            val = self._current_df.iloc[row_idx][col_name]
        except Exception:
            return ""
        if pd.isna(val):
            return ""
        return str(val)

    def _on_tree_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        item   = self.tree.identify_row(event.y)
        if not item or not col_id:
            return
        cols = self.tree["columns"]
        try:
            col_idx = int(col_id.replace("#", "")) - 1
            col_name = cols[col_idx]
        except (ValueError, IndexError):
            return
        val = self._get_raw_value(item, col_name)
        self._update_fbar(col_name, val)

    def _on_cell_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        cols = self.tree["columns"]
        if not cols:
            return
        if not hasattr(self, "_fbar_last_item") or self._fbar_last_item != sel[0]:
            self._fbar_last_item = sel[0]
            val = self._get_raw_value(sel[0], cols[0])
            self._update_fbar(cols[0], val)

    def _update_fbar(self, col_name, value):
        self._fbar_col_var.set(f"  {col_name}  ")
        self._fbar_val.config(state="normal")
        self._fbar_val.delete("1.0", "end")
        self._fbar_val.insert("1.0", str(value) if value != "" else "")
        self._fbar_val.config(state="disabled")

    def _fbar_copy(self):
        val = self._fbar_val.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(val)
        show_toast(self.master, "Value copied!", color=SUCCESS, duration=1500)

    # ── ROW DETAIL POPUP (double-click) ───────────────────────────────────────
    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        item    = sel[0]
        values  = self.tree.item(item, "values")
        columns = self.tree["columns"]
        if not columns or not values:
            return
        clicked_col = None
        if event and event.x:
            col_id = self.tree.identify_column(event.x)
            try:
                col_idx  = int(col_id.replace("#", "")) - 1
                clicked_col = columns[col_idx]
            except (ValueError, IndexError):
                pass
        self._show_row_detail(columns, values, clicked_col)

    def _show_row_detail(self, columns, values, clicked_col=None):
        popup = tk.Toplevel(self)
        popup.title("Row Details")
        popup.configure(bg=BG)
        popup.resizable(True, True)
        popup.grab_set()

        hdr = tk.Frame(popup, bg=HEADER_BG, height=48)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="🔍  Row Details",
                 font=("Segoe UI", 12, "bold"), bg=HEADER_BG, fg=TEXT
                 ).pack(side="left", padx=16, pady=10)
        tk.Label(hdr, text=f"{len(columns)} fields",
                 font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT
                 ).pack(side="left")
        tk.Button(hdr, text="✕  Close",
                  font=("Segoe UI", 8, "bold"),
                  bg=CARD, fg=TEXT, relief="flat", padx=10, pady=4,
                  cursor="hand2", command=popup.destroy
                  ).pack(side="right", padx=12, pady=8)
        tk.Frame(popup, bg=self.color, height=2).pack(fill="x")

        container = tk.Frame(popup, bg=BG)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg=BG, highlightthickness=0, bd=0)
        vsb2   = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner  = tk.Frame(canvas, bg=BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        inner.bind("<Configure>",  lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))

        def _mw(e): canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        popup.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        col_data = list(zip(columns, values))
        for i, (col, val) in enumerate(col_data):
            is_hl  = (col == clicked_col)
            row_bg = ACCENT2 if is_hl else (CARD if i % 2 == 0 else "#14172A")
            lbl_fg = "white"  if is_hl else SUBTEXT
            val_fg = "white"  if is_hl else TEXT

            row = tk.Frame(inner, bg=row_bg)
            row.pack(fill="x", padx=12, pady=1)

            tk.Label(row, text=f"{i+1:>3}", font=("Consolas", 8),
                     bg=row_bg, fg=SUBTEXT, width=3, anchor="e"
                     ).pack(side="left", padx=(8, 4), pady=6)

            tk.Label(row, text=col, font=("Segoe UI", 9, "bold"),
                     bg=row_bg, fg=lbl_fg, width=28, anchor="w"
                     ).pack(side="left", padx=(0, 10), pady=6)

            tk.Frame(row, bg=BORDER, width=1).pack(side="left", fill="y", pady=4)

            val_str  = str(val) if val != "" else "—"
            vt = tk.Text(row, font=("Segoe UI", 9),
                         bg=row_bg, fg=val_fg,
                         relief="flat", bd=0, height=1, wrap="none",
                         cursor="xterm",
                         selectbackground=ACCENT, selectforeground="white")
            vt.insert("1.0", val_str)
            if len(val_str) > 80:
                vt.config(height=min((len(val_str) // 80) + 1, 4), wrap="word")
            vt.config(state="disabled")
            vt.pack(side="left", fill="x", expand=True, padx=(10, 8), pady=4)

        footer = tk.Frame(popup, bg=DARK)
        footer.pack(fill="x", side="bottom")

        def _copy_all():
            popup.clipboard_clear()
            popup.clipboard_append("\n".join(f"{c}\t{v}" for c, v in col_data))
            show_toast(popup, "All fields copied!", color=SUCCESS)

        tk.Button(footer, text="📋  Copy All Fields",
                  font=("Segoe UI", 9, "bold"),
                  bg=self.color, fg="white", relief="flat",
                  padx=16, pady=7, cursor="hand2",
                  command=_copy_all).pack(side="left", padx=12, pady=8)
        tk.Label(footer, text="Double-click any row to view full details.",
                 font=("Segoe UI", 8), bg=DARK, fg=SUBTEXT
                 ).pack(side="right", padx=16)

        popup.update_idletasks()
        popup.geometry("700x560")
        px = self.winfo_rootx() + (self.winfo_width()  - 700) // 2
        py = self.winfo_rooty() + (self.winfo_height() - 560) // 2
        popup.geometry(f"700x560+{max(px,0)}+{max(py,0)}")

    # ── EXPORT ────────────────────────────────────────────────────────────────
    def _download(self):
        if self._current_df is None or self._current_df.empty:
            return
        save_path = filedialog.asksaveasfilename(
            title="Save view as",
            defaultextension=".xlsx",
            initialfile=self.export_filename,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                self._current_df.to_excel(writer, index=False, sheet_name="Data")
                ws = writer.sheets["Data"]
                white_bold  = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="1E2235")
                for cell in ws[1]:
                    cell.font = white_bold
                    cell.fill = header_fill
                for col in ws.columns:
                    header_len = len(str(col[0].value)) if col[0].value else 10
                    ws.column_dimensions[col[0].column_letter].width = min(header_len + 6, 40)
            self.status_var.set(f"Saved → {save_path}")
            show_toast(self.master, f"Saved: {os.path.basename(save_path)}", color=self.color)
        except Exception as e:
            show_toast(self.master, f"Save failed: {e}", color=DANGER, duration=5000)



# ══════════════════════════════════════════════════════════════════════════════
# TAB — GEO REFERENCE  (Unique / Province / Municipality / Final Area /
#        Geocode / Cluster / Area Status)
#
# A saved reference table, independent from the Field Result / Masterfile
# tabs. Every row is written straight to the shared SQLite DB (see
# db_add_geo_reference etc., DB_PATH under _app_data_dir()), so the data is
# still there the next time the app is opened — closing the app never
# deletes it.
# ══════════════════════════════════════════════════════════════════════════════
class _GeoReferencePanel(tk.Frame):
    COLUMNS = GEO_REFERENCE_COLUMNS  # [(db_field, display_label), ...]

    def __init__(self, master, on_data_changed=None):
        super().__init__(master, bg=BG)
        self._editing_id = None
        self.on_data_changed = on_data_changed  # callable(self) fired after saved data changes
        self._build_ui()
        self._refresh()

    # ── layout ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # LEFT — saved table
        left = tk.Frame(body, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))

        top_row = tk.Frame(left, bg=CARD)
        top_row.pack(fill="x", padx=16, pady=(14, 8))
        tk.Label(top_row, text="🌍  Geo Reference  ·  saved locally",
                 font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT).pack(side="left")
        self._count_var = tk.StringVar(value="0 rows")
        tk.Label(top_row, textvariable=self._count_var,
                 font=("Segoe UI", 8, "bold"), bg=CARD, fg=SUBTEXT).pack(side="right")

        tree_frame = tk.Frame(left, bg=CARD)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        style = ttk.Style()
        style.configure("GeoRef.Treeview",
                        background=DARK, fieldbackground=DARK,
                        foreground=TEXT, rowheight=28,
                        font=("Segoe UI", 9), bordercolor=BORDER, relief="flat")
        style.configure("GeoRef.Treeview.Heading",
                        background=HEADER_BG, foreground=ACCENT,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("GeoRef.Treeview",
                  background=[("selected", ACCENT)], foreground=[("selected", "white")])

        cols = [f for f, _ in self.COLUMNS]
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                 style="GeoRef.Treeview", selectmode="browse",
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        for field, label in self.COLUMNS:
            self.tree.heading(field, text=label)
            self.tree.column(field, width=110, anchor="w")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # Action buttons
        btns = tk.Frame(left, bg=CARD)
        btns.pack(fill="x", padx=12, pady=(0, 14))
        tk.Button(btns, text="🗑  Delete Selected", font=("Segoe UI", 9, "bold"),
                  bg=DANGER, fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._delete_selected).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="↺  Refresh", font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=SUBTEXT, relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._refresh).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="📥  Import Excel", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._import_excel).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="⬇  Export Excel", font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._export_excel).pack(side="left")

        # RIGHT — add / edit form
        right = tk.Frame(body, bg=CARD, highlightthickness=1,
                         highlightbackground=BORDER, width=300)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        tk.Label(right, text="➕  Add / Edit Row", font=("Segoe UI", 11, "bold"),
                 bg=CARD, fg=TEXT).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x", padx=12, pady=(0, 8))

        self._form_vars = {}
        for field, label in self.COLUMNS:
            self._form_vars[field] = self._rfield(right, label)

        self._form_err = tk.StringVar()
        tk.Label(right, textvariable=self._form_err, font=("Segoe UI", 8),
                 bg=CARD, fg=DANGER, wraplength=260, justify="left"
                 ).pack(anchor="w", padx=16, pady=(6, 0))

        tk.Button(right, text="✚  Add New Row", font=("Segoe UI", 10, "bold"),
                  bg=SUCCESS, fg="white", relief="flat", pady=8, cursor="hand2",
                  command=self._add_row).pack(fill="x", padx=16, pady=(10, 6))
        tk.Button(right, text="✎  Update Selected", font=("Segoe UI", 10, "bold"),
                  bg=ACCENT, fg="white", relief="flat", pady=8, cursor="hand2",
                  command=self._update_row).pack(fill="x", padx=16, pady=(0, 6))
        tk.Button(right, text="✕  Clear Form", font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat", pady=6, cursor="hand2",
                  command=self._clear_form).pack(fill="x", padx=16)

    def _rfield(self, parent, label_text):
        tk.Label(parent, text=label_text.upper(), font=("Segoe UI", 8, "bold"),
                 bg=CARD, fg=SUBTEXT).pack(anchor="w", padx=16, pady=(6, 2))
        var = tk.StringVar()
        wrapper = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
        wrapper.pack(fill="x", padx=12)
        inner = tk.Frame(wrapper, bg=CARD)
        inner.pack(fill="x")
        entry = tk.Entry(inner, textvariable=var, font=("Segoe UI", 10),
                         bg=CARD, fg=TEXT, insertbackground=TEXT, relief="flat", bd=6)
        entry.pack(fill="x")
        entry.bind("<FocusIn>",  lambda e: wrapper.config(bg=ACCENT))
        entry.bind("<FocusOut>", lambda e: wrapper.config(bg=BORDER))
        return var

    # ── data ──────────────────────────────────────────────────────────────────
    def _refresh(self):
        self.tree.delete(*self.tree.get_children())
        rows = db_list_geo_reference()
        for row in rows:
            row_id = row[0]
            values = row[1:8]   # unique_code..area_status, skip id and created_at
            self.tree.insert("", "end", iid=str(row_id), values=values)
        self._count_var.set(f"{len(rows):,} row{'s' if len(rows) != 1 else ''}")
        if self.on_data_changed:
            self.on_data_changed(self)

    def _on_select(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        self._editing_id = int(sel[0])
        vals = self.tree.item(sel[0], "values")
        for (field, _label), val in zip(self.COLUMNS, vals):
            self._form_vars[field].set(val)
        self._form_err.set("")

    def _clear_form(self):
        self._editing_id = None
        for var in self._form_vars.values():
            var.set("")
        self._form_err.set("")
        self.tree.selection_remove(self.tree.selection())

    def _form_values(self):
        return [self._form_vars[field].get().strip() for field, _ in self.COLUMNS]

    def _add_row(self):
        vals = self._form_values()
        if not any(vals):
            self._form_err.set("⚠ Fill in at least one field.")
            return
        db_add_geo_reference(*vals)
        self._form_err.set("")
        self._clear_form()
        self._refresh()
        show_toast(self.winfo_toplevel(), "Row added.", color=SUCCESS)

    def _update_row(self):
        if self._editing_id is None:
            self._form_err.set("⚠ Select a row from the table first.")
            return
        vals = self._form_values()
        db_update_geo_reference(self._editing_id, *vals)
        self._form_err.set("")
        self._clear_form()
        self._refresh()
        show_toast(self.winfo_toplevel(), "Row updated.", color=SUCCESS)

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete {len(sel)} selected row(s)? This cannot be undone.",
                                   parent=self):
            return
        for iid in sel:
            db_delete_geo_reference(int(iid))
        self._clear_form()
        self._refresh()
        show_toast(self.winfo_toplevel(), "Row(s) deleted.", color=WARNING)

    # ── import / export ──────────────────────────────────────────────────────
    def _import_excel(self):
        path = filedialog.askopenfilename(
            title="Import Geo Reference",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if not path:
            return
        try:
            df = pd.read_excel(path, dtype=str).fillna("")
            file_cols_norm = {str(c).strip().upper(): c for c in df.columns}
            missing = [label for _, label in self.COLUMNS if label.upper() not in file_cols_norm]
            if missing:
                messagebox.showerror(
                    "Import Error",
                    f"Missing expected column(s): {', '.join(missing)}",
                    parent=self
                )
                return
            rows = []
            for _, r in df.iterrows():
                rows.append(tuple(
                    str(r[file_cols_norm[label.upper()]]).strip() for _, label in self.COLUMNS
                ))
            if not rows:
                messagebox.showinfo("Import", "No rows found in that file.", parent=self)
                return
            n = db_import_geo_reference_bulk(rows)
            self._refresh()
            show_toast(self.winfo_toplevel(), f"Imported {n:,} row(s).", color=SUCCESS)
        except Exception as e:
            messagebox.showerror("Import Error", str(e), parent=self)

    def _export_excel(self):
        rows = db_list_geo_reference()
        if not rows:
            messagebox.showinfo("Export", "No rows to export yet.", parent=self)
            return
        save_path = filedialog.asksaveasfilename(
            title="Export Geo Reference",
            defaultextension=".xlsx",
            initialfile="geo_reference.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            labels = [label for _, label in self.COLUMNS]
            data = [row[1:8] for row in rows]
            df = pd.DataFrame(data, columns=labels)
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Geo Reference")
                ws = writer.sheets["Geo Reference"]
                white_bold  = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="1E2235")
                for cell in ws[1]:
                    cell.font = white_bold
                    cell.fill = header_fill
                for col in ws.columns:
                    header_len = len(str(col[0].value)) if col[0].value else 10
                    ws.column_dimensions[col[0].column_letter].width = min(header_len + 6, 40)
            show_toast(self.winfo_toplevel(), f"Saved: {os.path.basename(save_path)}", color=SUCCESS)
        except Exception as e:
            show_toast(self.winfo_toplevel(), f"Save failed: {e}", color=DANGER, duration=5000)


# ══════════════════════════════════════════════════════════════════════════════
# TAB — SUMMARY  (live pivot: Field Status × [Coverage Status, State] → Bucket)
#
# Reads straight from the Masterfile tab's currently loaded data — no
# separate file to load here. Shaped like an Excel PivotTable:
#   Rows    = Coverage Status, State
#   Columns = Field Status
#   Values  = Count of Placement
# Refreshes automatically whenever the Masterfile's data changes (a new
# load, or a recompute triggered by the Field Result / Geo Reference tabs).
# ══════════════════════════════════════════════════════════════════════════════
class _SummaryPanel(tk.Frame):
    ROW_FIELDS  = ["Coverage Status", "State"]
    COL_FIELD   = "Visit Status"
    VALUE_FIELD = "Placement"

    def __init__(self, master, masterfile_tab):
        super().__init__(master, bg=BG)
        self.masterfile_tab = masterfile_tab
        self._pivot_df = None   # last built pivot table (for export)
        self._build_ui()
        self.refresh()

    # ── layout ────────────────────────────────────────────────────────────────
    def _build_ui(self):
        header = tk.Frame(self, bg=BG)
        header.pack(fill="x", padx=20, pady=(16, 4))
        tk.Label(header,
                 text="📊  Summary  ·  Rows: Coverage Status, State  ·  Columns: Visit Status  ·  Values: Count of Placement",
                 font=("Segoe UI", 10, "bold"), bg=BG, fg=TEXT).pack(side="left")
        tk.Button(header, text="⬇  Export Excel", font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self._export_excel).pack(side="right")
        tk.Button(header, text="↺  Refresh", font=("Segoe UI", 9, "bold"),
                  bg=ACCENT2, fg="white", relief="flat", padx=10, pady=6,
                  cursor="hand2", command=self.refresh).pack(side="right", padx=(0, 8))

        self._info_var = tk.StringVar(value="No Masterfile data loaded yet.")
        tk.Label(self, textvariable=self._info_var, font=("Segoe UI", 8),
                 bg=BG, fg=SUBTEXT).pack(anchor="w", padx=22, pady=(0, 8))

        paned = ttk.PanedWindow(self, orient="vertical")
        paned.pack(fill="both", expand=True, padx=20, pady=(0, 16))

        body = tk.Frame(paned, bg=BG)
        paned.add(body, weight=3)

        style = ttk.Style()
        style.configure("Summary.Treeview", background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28,
                         font=("Segoe UI", 9), borderwidth=0)
        style.configure("Summary.Treeview.Heading", background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Summary.Treeview",
                  background=[("selected", ACCENT2)], foreground=[("selected", "white")])

        vsb = ttk.Scrollbar(body, orient="vertical")
        hsb = ttk.Scrollbar(body, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree = ttk.Treeview(body, style="Summary.Treeview",
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                 show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        # ── Bar graph pane ────────────────────────────────────────────────────
        chart_container = tk.Frame(paned, bg=CARD)
        paned.add(chart_container, weight=2)

        chart_header = tk.Frame(chart_container, bg=CARD)
        chart_header.pack(fill="x", padx=14, pady=(10, 0))
        tk.Label(chart_header, text="📈  Placement Count by Visit Status",
                 font=("Segoe UI", 10, "bold"), bg=CARD, fg=TEXT).pack(side="left")

        self._chart_holder = tk.Frame(chart_container, bg=CARD)
        self._chart_holder.pack(fill="both", expand=True, padx=14, pady=(4, 12))

        if _MPL_AVAILABLE:
            self._fig = Figure(figsize=(5, 3), dpi=100, facecolor=CARD)
            self._ax = self._fig.add_subplot(111)
            self._canvas = FigureCanvasTkAgg(self._fig, master=self._chart_holder)
            self._canvas.get_tk_widget().pack(fill="both", expand=True)
        else:
            tk.Label(self._chart_holder,
                     text="Bar graph unavailable — install matplotlib (pip install matplotlib) to enable it.",
                     font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT, wraplength=500,
                     justify="center").pack(expand=True)

    # ── helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _find_col(df, name):
        if df is None:
            return None
        return next(
            (c for c in df.columns if str(c).strip().upper() == name.strip().upper()), None
        )

    def _build_pivot(self):
        mdf = self.masterfile_tab.df
        if mdf is None or mdf.empty:
            return None, "No Masterfile data loaded yet. Load it in the Masterfile tab first."

        col_field = self._find_col(mdf, self.COL_FIELD)
        row_cols  = [self._find_col(mdf, f) for f in self.ROW_FIELDS]
        val_col   = self._find_col(mdf, self.VALUE_FIELD)

        missing = []
        if col_field is None:
            missing.append(self.COL_FIELD)
        missing += [f for f, c in zip(self.ROW_FIELDS, row_cols) if c is None]
        if val_col is None:
            missing.append(self.VALUE_FIELD)
        if missing:
            return None, (
                f"Masterfile is missing column(s): {', '.join(missing)}. "
                f"These need to be present in the loaded Masterfile file."
            )

        work = mdf[row_cols + [col_field, val_col]].copy()
        for c in row_cols + [col_field]:
            work[c] = work[c].astype(str).str.strip()
            work[c] = work[c].replace("", "(blank)")

        try:
            pivot = pd.pivot_table(
                work, index=row_cols, columns=col_field, values=val_col,
                aggfunc="count", fill_value=0,
                margins=True, margins_name="Grand Total",
            )
        except Exception as e:
            return None, f"Could not build pivot table: {e}"

        pivot = pivot.reset_index()
        # Friendly labels for the row-grouping columns (e.g. actual "STATE" -> "State")
        pivot = pivot.rename(columns={
            row_cols[i]: self.ROW_FIELDS[i] for i in range(len(row_cols))
        })
        return pivot, None

    # ── data / render ─────────────────────────────────────────────────────────
    def refresh(self):
        pivot, msg = self._build_pivot()
        self._pivot_df = pivot
        self.tree.delete(*self.tree.get_children())
        if pivot is None:
            self.tree["columns"] = []
            self._info_var.set(msg or "No data.")
            self._update_chart(None)
            return

        cols = [str(c) for c in pivot.columns]
        self.tree["columns"] = cols
        for label in cols:
            self.tree.heading(label, text=label, anchor="w")
            width = max(110, min(len(label) * 9, 220))
            self.tree.column(label, width=width, anchor="w", stretch=False)

        for _, row in pivot.iterrows():
            self.tree.insert("", "end", values=[row[c] for c in pivot.columns])

        self._info_var.set(
            f"{len(pivot):,} row(s)  ·  {len(cols):,} column(s)  ·  "
            f"live from the current Masterfile data."
        )
        self._update_chart(pivot)

    def _update_chart(self, pivot):
        """Draw/refresh the bar graph once pivot data has been generated.

        Mirrors a PivotChart layout: bars grouped by the outer row field
        (Coverage Status) with the inner row field (State) as individual
        category ticks, one colored bar series per column field value
        (Visit Status), legend titled "Field Status"."""
        if not _MPL_AVAILABLE:
            return

        self._ax.clear()
        self._fig.patch.set_facecolor(CARD)
        self._ax.set_facecolor(CARD)

        if pivot is None or pivot.empty:
            self._ax.axis("off")
            self._ax.text(0.5, 0.5, "No data yet — load the Masterfile to generate the chart.",
                           ha="center", va="center", color=SUBTEXT, fontsize=9,
                           transform=self._ax.transAxes, wrap=True)
            self._canvas.draw()
            return

        outer_field = self.ROW_FIELDS[0]
        inner_field = self.ROW_FIELDS[1] if len(self.ROW_FIELDS) > 1 else None
        skip_cols  = set(self.ROW_FIELDS) | {"Grand Total"}
        value_cols = [c for c in pivot.columns if c not in skip_cols]

        # Drop the overall "Grand Total" row — only chart the real data rows.
        data_rows = pivot[pivot[outer_field].astype(str) != "Grand Total"]
        if data_rows.empty or not value_cols:
            self._ax.axis("off")
            self._ax.text(0.5, 0.5, "Not enough data to chart yet.",
                           ha="center", va="center", color=SUBTEXT, fontsize=9,
                           transform=self._ax.transAxes)
            self._canvas.draw()
            return

        # Preserve first-seen order of each outer group, then walk rows in
        # that order, leaving one empty slot as a gap between groups.
        group_order = list(dict.fromkeys(str(v) for v in data_rows[outer_field]))
        grouped = {g: [] for g in group_order}
        for _, r in data_rows.iterrows():
            grouped[str(r[outer_field])].append(r)

        x_labels, x_positions, group_spans = [], [], []
        series_x  = {c: [] for c in value_cols}
        series_y  = {c: [] for c in value_cols}
        pos = 0
        for g in group_order:
            start = pos
            for r in grouped[g]:
                label = str(r[inner_field]) if inner_field else g
                x_labels.append(label)
                x_positions.append(pos)
                for c in value_cols:
                    series_x[c].append(pos)
                    series_y[c].append(r[c])
                pos += 1
            group_spans.append((start, pos - 1, g))
            pos += 1  # gap before next group

        palette = [ACCENT, WARNING, ACCENT2, SUCCESS, DANGER]
        n_series = len(value_cols)
        bar_width = min(0.8 / max(n_series, 1), 0.4)

        for i, c in enumerate(value_cols):
            offset = (i - (n_series - 1) / 2) * bar_width
            bars = self._ax.bar(
                [x + offset for x in series_x[c]], series_y[c],
                width=bar_width * 0.95, color=palette[i % len(palette)], label=str(c),
            )

        self._ax.set_xticks(x_positions)
        self._ax.set_xticklabels(x_labels, rotation=45, ha="right", color=TEXT, fontsize=7)
        self._ax.tick_params(axis="y", colors=TEXT, labelsize=8)
        for spine in self._ax.spines.values():
            spine.set_color(BORDER)
        self._ax.grid(axis="y", color=BORDER, alpha=0.35, linewidth=0.6)
        self._ax.set_axisbelow(True)
        self._ax.margins(x=0.01)

        # Group labels (e.g. COVERED / NOT COVERED / SPECIAL FIELD) beneath
        # the State ticks, plus thin separators between groups.
        for start, end, g in group_spans:
            mid = (start + end) / 2
            self._ax.text(mid, -0.42, g, transform=self._ax.get_xaxis_transform(),
                          ha="center", va="top", color=TEXT, fontsize=8,
                          fontweight="bold", clip_on=False)
        for idx in range(len(group_spans) - 1):
            sep_x = (group_spans[idx][1] + group_spans[idx + 1][0]) / 2
            self._ax.axvline(sep_x, color=BORDER, linewidth=1, alpha=0.8)

        legend = self._ax.legend(
            title="Field Status", loc="center left", bbox_to_anchor=(1.01, 0.5),
            facecolor=CARD, edgecolor=BORDER, labelcolor=TEXT, fontsize=8,
        )
        legend.get_title().set_color(TEXT)
        legend.get_title().set_fontsize(8)

        self._ax.set_title("Visited Accounts", loc="left", color=TEXT, fontsize=9,
                            fontweight="bold",
                            bbox=dict(facecolor=HEADER_BG, edgecolor=BORDER,
                                      boxstyle="round,pad=0.35"))

        self._fig.subplots_adjust(bottom=0.40, right=0.78, top=0.88, left=0.08)
        self._canvas.draw()

    def _export_excel(self):
        if self._pivot_df is None or self._pivot_df.empty:
            messagebox.showinfo("Export", "Nothing to export yet — load the Masterfile first.", parent=self)
            return
        save_path = filedialog.asksaveasfilename(
            title="Export Summary Pivot",
            defaultextension=".xlsx",
            initialfile="summary_pivot.xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                self._pivot_df.to_excel(writer, index=False, sheet_name="Summary")
                ws = writer.sheets["Summary"]
                white_bold  = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill("solid", fgColor="1E2235")
                for cell in ws[1]:
                    cell.font = white_bold
                    cell.fill = header_fill
                for col in ws.columns:
                    header_len = len(str(col[0].value)) if col[0].value else 10
                    ws.column_dimensions[col[0].column_letter].width = min(header_len + 6, 40)
                self._add_excel_chart(ws, self._pivot_df)
            show_toast(self.winfo_toplevel(), f"Saved: {os.path.basename(save_path)}", color=SUCCESS)
        except Exception as e:
            show_toast(self.winfo_toplevel(), f"Save failed: {e}", color=DANGER, duration=5000)

    def _add_excel_chart(self, ws, pivot_df):
        """Insert a native, editable Excel bar chart next to the pivot table:
        one clustered bar series per Visit Status column, categories grouped
        by Coverage Status / State (matching the in-app chart)."""
        n_row_fields = len(self.ROW_FIELDS)          # 2 -> Coverage Status, State
        total_cols   = len(pivot_df.columns)
        header_row   = 1
        first_data_row      = 2
        grand_total_row     = header_row + len(pivot_df)   # last row = "Grand Total"
        last_value_data_row = grand_total_row - 1          # exclude it from the chart

        value_col_start = n_row_fields + 1     # first Visit Status column
        value_col_end   = total_cols - 1       # exclude the "Grand Total" column

        if last_value_data_row < first_data_row or value_col_end < value_col_start:
            return  # nothing meaningful to chart

        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.overlap  = -20
        chart.gapWidth = 60
        chart.title    = "Visited Accounts"
        chart.y_axis.title = "Count of Placement"
        chart.x_axis.title = None
        chart.style  = 10
        chart.height = 14      # taller so rotated State labels don't crowd
        chart.width  = 34      # the Coverage Status tier row below them

        data = Reference(ws, min_col=value_col_start, max_col=value_col_end,
                          min_row=header_row, max_row=last_value_data_row)
        chart.add_data(data, titles_from_data=True)

        # Multi-column category reference -> Excel groups the axis by
        # Coverage Status (outer) then State (inner), same as a PivotChart.
        cats = Reference(ws, min_col=1, max_col=n_row_fields,
                          min_row=first_data_row, max_row=last_value_data_row)
        chart.set_categories(cats)
        chart.legend.position = "r"

        # Angle the (inner) State labels at 45° instead of a full 90° so they
        # take up far less vertical space and stop overlapping the
        # Coverage Status tier row printed underneath them.
        label_props = CharacterProperties(sz=800)          # 8pt
        chart.x_axis.txPr = RichText(
            bodyPr=RichTextProperties(rot=-2700000, vert="horz"),  # -45°
            p=[Paragraph(pPr=ParagraphProperties(defRPr=label_props), endParaRPr=label_props)],
        )
        chart.x_axis.majorTickMark = "none"
        chart.x_axis.minorTickMark = "none"

        # Data table: shows the per-category values (with color-coded keys
        # matching the series) directly under the plot, per user request.
        chart.plot_area.dTable = DataTable(
            showHorzBorder=True, showVertBorder=True,
            showOutline=True, showKeys=True,
        )

        palette_hex = ["4F8EF7", "F59E0B", "7C3AED", "22C55E", "EF4444"]
        for i, series in enumerate(chart.series):
            series.graphicalProperties.solidFill = palette_hex[i % len(palette_hex)]

        anchor_row = grand_total_row + 3
        ws.add_chart(chart, f"A{anchor_row}")


# ══════════════════════════════════════════════════════════════════════════════
# SCREEN — ANALYTICAL REPORTS (Field Result + Masterfile tabs)
# ══════════════════════════════════════════════════════════════════════════════
class AnalyticalReportsScreen(tk.Frame):
    """Container screen: header + a two-tab Notebook (Field Result,
    Masterfile) + footer. Each tab is an independent _AnalyticsGridPanel
    with its own loaded file, grid, search/filter, and export — importing
    data into one tab never touches the other."""

    def __init__(self, master, username, role="user", on_back=None):
        super().__init__(master, bg=BG)
        self.username = username
        self.role     = role
        self.on_back  = on_back
        self._build_ui()

    @staticmethod
    def _field_result_sheet_picker(sheet_names):
        sheet_name = next(
            (s for s in sheet_names if s.strip().upper() == "RESULT"), None
        )
        if sheet_name is None:
            raise ValueError(
                f"No \"RESULT\" sheet found in this file. "
                f"Sheets present: {', '.join(sheet_names)}"
            )
        return sheet_name

    @staticmethod
    def _masterfile_sheet_picker(sheet_names):
        # The Masterfile tab only ever reads the "Complete Accounts" sheet,
        # matched case-insensitively and trimmed (so "COMPLETE ACCOUNTS",
        # "Complete accounts " etc. all count).
        match = next(
            (s for s in sheet_names if s.strip().upper() == "COMPLETE ACCOUNTS"), None
        )
        if match is None:
            raise ValueError(
                f"No \"Complete Accounts\" sheet found in this file. "
                f"Sheets present: {', '.join(sheet_names)}"
            )
        return match

    def _build_ui(self):
        # ── Header ────────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  📊  Analytical Reports  —  Orico Auto Loan",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)
        tk.Label(left_hdr, text="Upload · View · Search · Filter",
                 font=("Segoe UI", 8), bg=HEADER_BG, fg=SUBTEXT).pack(side="left", padx=4)

        right_hdr = tk.Frame(header, bg=HEADER_BG)
        right_hdr.pack(side="right", padx=16)
        role_colors = {"admin": DANGER, "supervisor": WARNING, "user": SUCCESS}
        role_color  = role_colors.get(self.role, ACCENT)
        tk.Label(right_hdr, text=f"👤  {self.username}",
                 font=("Segoe UI", 9, "bold"), bg=HEADER_BG, fg=ACCENT).pack()
        tk.Label(right_hdr, text=f"● {self.role.upper()}",
                 font=("Segoe UI", 7, "bold"), bg=HEADER_BG, fg=role_color).pack()

        tk.Frame(self, bg="#0E9F6E", height=3).pack(fill="x")

        # ── Tabs ──────────────────────────────────────────────────────────────
        style = ttk.Style()
        style.configure("Analytics.TNotebook", background=BG, borderwidth=0)
        style.configure("Analytics.TNotebook.Tab", background=CARD, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), padding=(16, 8))
        style.map("Analytics.TNotebook.Tab",
                  background=[("selected", HEADER_BG)],
                  foreground=[("selected", "white")])

        notebook = ttk.Notebook(self, style="Analytics.TNotebook")
        notebook.pack(fill="both", expand=True)

        field_result_tab = _AnalyticsGridPanel(
            notebook,
            color="#0E9F6E",
            drop_text="Load any Excel file to view it",
            sheet_picker=self._field_result_sheet_picker,
            bank_filter_col="BANK",
            bank_filter_value="ORICO AUTO LOAN",
            export_filename="field_result.xlsx",
            empty_msg="No data loaded. Browse a file on the left to get started.",
        )
        masterfile_tab = _AnalyticsGridPanel(
            notebook,
            color="#7C3AED",
            drop_text="Load the Masterfile to import records",
            sheet_picker=self._masterfile_sheet_picker,
            default_headers=MASTERLIST_HEADERS,
            export_filename="masterfile.xlsx",
            empty_msg="No data loaded. Browse a Masterfile on the left to import.",
            usecols="D:R",
            force_date_columns=["ENDO DATE"],
            visit_count_col="No. of Visits",
            visit_count_key="CHCODE",
            visit_status_col="Visit Status",
            field_result_panel=field_result_tab,
            coverage_status_col="Coverage Status",
            coverage_status_source_col="STATE",
        )
        summary_tab = _SummaryPanel(notebook, masterfile_tab=masterfile_tab)

        def _on_field_result_changed(panel):
            # Field Result reloading affects the Masterfile's visit counts,
            # which don't feed the pivot directly, but keep everything in
            # lockstep anyway.
            masterfile_tab.recompute_visit_counts()
            summary_tab.refresh()

        def _on_masterfile_changed(panel):
            summary_tab.refresh()

        def _on_geo_reference_changed(panel):
            # Geo Reference changes ripple into Coverage Status, which is
            # one of the pivot's row fields — refresh both.
            masterfile_tab.recompute_coverage_status()
            summary_tab.refresh()

        field_result_tab.on_data_changed = _on_field_result_changed
        masterfile_tab.on_data_changed   = _on_masterfile_changed

        # When the Geo Reference tab's saved data changes (add/edit/delete/
        # import/clear), refresh the Masterfile tab's Coverage Status column
        # too — it's looked up from that saved data.
        geo_reference_tab = _GeoReferencePanel(
            notebook,
            on_data_changed=_on_geo_reference_changed,
        )

        notebook.add(field_result_tab, text="📋  Field Result")
        notebook.add(masterfile_tab,  text="🗂  Masterfile")
        notebook.add(summary_tab, text="📊  Summary")
        notebook.add(geo_reference_tab, text="🌍  Geo Reference")

        # Footer
        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)


# ══════════════════════════════════════════════════════════════════════════════
# BANK OF MAKATI PLACEHOLDER
# ══════════════════════════════════════════════════════════════════════════════
class BankOfMakatiScreen(tk.Frame):
    def __init__(self, master, username, on_back):
        super().__init__(master, bg=BG)
        self.username = username
        self.on_back  = on_back
        self._build()

    def _build(self):
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  🏦  Bank of Makati",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)

        tk.Frame(self, bg=ACCENT2, height=3).pack(fill="x")

        outer = tk.Frame(self, bg=BG)
        outer.pack(expand=True)

        card = tk.Frame(outer, bg=CARD, padx=60, pady=60,
                        highlightthickness=1, highlightbackground=BORDER)
        card.pack(padx=20, pady=20)

        tk.Label(card, text="🏦", font=("Segoe UI", 48), bg=CARD, fg=ACCENT2).pack()
        tk.Label(card, text="Bank of Makati", font=("Segoe UI", 20, "bold"),
                 bg=CARD, fg=TEXT).pack(pady=(12, 4))
        tk.Label(card, text="Tools for this account are coming soon.",
                 font=("Segoe UI", 10), bg=CARD, fg=SUBTEXT).pack()

        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNT MANAGEMENT SCREEN  (admin only)
# ══════════════════════════════════════════════════════════════════════════════
class AccountManagementScreen(tk.Frame):
    def __init__(self, master, username, on_back):
        super().__init__(master, bg=BG)
        self.username = username
        self.on_back  = on_back
        self._build()

    # ── layout ────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        header = tk.Frame(self, bg=HEADER_BG, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)

        left_hdr = tk.Frame(header, bg=HEADER_BG)
        left_hdr.pack(side="left", padx=12, pady=8)
        tk.Button(left_hdr, text="←  Back",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=TEXT, relief="flat",
                  padx=12, pady=6, cursor="hand2",
                  command=self.on_back).pack(side="left")
        tk.Label(left_hdr, text="  ⚙  Account Management",
                 font=("Segoe UI", 13, "bold"), bg=HEADER_BG, fg=TEXT).pack(side="left", padx=8)

        tk.Frame(self, bg=GOLD, height=3).pack(fill="x")

        # Two-column layout
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=24, pady=20)

        # LEFT — user table
        left = tk.Frame(body, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))

        tk.Label(left, text="👥  Registered Users",
                 font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT
                 ).pack(anchor="w", padx=16, pady=(14, 8))

        # Treeview
        tree_frame = tk.Frame(left, bg=CARD)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Mgmt.Treeview",
                        background=DARK, fieldbackground=DARK,
                        foreground=TEXT, rowheight=30,
                        bordercolor=BORDER, relief="flat")
        style.configure("Mgmt.Treeview.Heading",
                        background=HEADER_BG, foreground=ACCENT,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Mgmt.Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", "white")])

        cols = ("username", "role", "created_at")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                 style="Mgmt.Treeview", selectmode="browse")
        for col, w, label in [("username", 140, "Username"),
                               ("role",     80,  "Role"),
                               ("created_at", 160, "Created At")]:
            self.tree.heading(col, text=label)
            self.tree.column(col, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        # Action buttons below table
        btns = tk.Frame(left, bg=CARD)
        btns.pack(fill="x", padx=12, pady=(0, 14))
        tk.Button(btns, text="🗑  Delete Selected",
                  font=("Segoe UI", 9, "bold"),
                  bg=DANGER, fg="white", relief="flat",
                  padx=10, pady=6, cursor="hand2",
                  command=self._delete_selected).pack(side="left", padx=(0, 8))
        tk.Button(btns, text="↺  Refresh",
                  font=("Segoe UI", 9, "bold"),
                  bg=CARD, fg=SUBTEXT, relief="flat",
                  padx=10, pady=6, cursor="hand2",
                  command=self._refresh).pack(side="left")

        # RIGHT — add / edit panel
        right = tk.Frame(body, bg=CARD, highlightthickness=1,
                         highlightbackground=BORDER, width=300)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        tk.Label(right, text="➕  Add / Edit Account",
                 font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT
                 ).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Frame(right, bg=BORDER, height=1).pack(fill="x", padx=12, pady=(0, 12))

        # Form fields
        self._form_username = self._rfield(right, "Username")
        self._form_password = self._rfield(right, "Password", show="●")
        self._form_confirm  = self._rfield(right, "Confirm Password", show="●")

        tk.Label(right, text="ROLE", font=("Segoe UI", 8, "bold"),
                 bg=CARD, fg=SUBTEXT).pack(anchor="w", padx=16, pady=(8, 2))
        self._role_var = tk.StringVar(value="user")
        role_frame = tk.Frame(right, bg=CARD)
        role_frame.pack(anchor="w", padx=16)
        for val, label in [("user", "User"), ("supervisor", "Supervisor"), ("admin", "Admin")]:
            tk.Radiobutton(role_frame, text=label, variable=self._role_var, value=val,
                           font=("Segoe UI", 9), bg=CARD, fg=TEXT,
                           selectcolor=CARD, activebackground=CARD,
                           activeforeground=ACCENT).pack(side="left", padx=(0, 12))

        # Status / error
        self._form_err = tk.StringVar()
        tk.Label(right, textvariable=self._form_err,
                 font=("Segoe UI", 8), bg=CARD, fg=DANGER,
                 wraplength=260, justify="left").pack(anchor="w", padx=16, pady=(8, 0))

        # Buttons
        tk.Button(right, text="✚  Add New User",
                  font=("Segoe UI", 10, "bold"),
                  bg=SUCCESS, fg="white", relief="flat",
                  pady=9, cursor="hand2",
                  command=self._add_user).pack(fill="x", padx=16, pady=(12, 6))

        tk.Button(right, text="✎  Update Selected",
                  font=("Segoe UI", 10, "bold"),
                  bg=ACCENT, fg="white", relief="flat",
                  pady=9, cursor="hand2",
                  command=self._update_user).pack(fill="x", padx=16, pady=(0, 6))

        tk.Button(right, text="✕  Clear Form",
                  font=("Segoe UI", 9),
                  bg=CARD, fg=SUBTEXT, relief="flat",
                  pady=7, cursor="hand2",
                  command=self._clear_form).pack(fill="x", padx=16)

        # Footer
        footer = tk.Frame(self, bg=DARK)
        footer.pack(fill="x", side="bottom")
        tk.Label(footer, text="Created by  Vincent Corocoto  ·  09567796275",
                 font=("Segoe UI", 8, "bold"), bg=DARK, fg=ACCENT).pack(side="left", padx=16, pady=8)
        tk.Label(footer, text='"Kapag ang palay naging bigas, May bumayo."',
                 font=("Segoe UI", 8, "italic"), bg=DARK, fg="#4A5568").pack(side="right", padx=16, pady=8)

        self._refresh()

    # ── helpers ───────────────────────────────────────────────────────────────
    def _rfield(self, parent, label_text, show=None):
        tk.Label(parent, text=label_text.upper(),
                 font=("Segoe UI", 8, "bold"), bg=CARD, fg=SUBTEXT
                 ).pack(anchor="w", padx=16, pady=(8, 2))
        var = tk.StringVar()
        wrapper = tk.Frame(parent, bg=BORDER, padx=1, pady=1)
        wrapper.pack(fill="x", padx=12)
        inner = tk.Frame(wrapper, bg=CARD)
        inner.pack(fill="x")
        entry = tk.Entry(inner, textvariable=var,
                         font=("Segoe UI", 10), bg=CARD, fg=TEXT,
                         insertbackground=TEXT, relief="flat", bd=6, show=show or "")
        entry.pack(fill="x")
        def fi(e): wrapper.config(bg=ACCENT)
        def fo(e): wrapper.config(bg=BORDER)
        entry.bind("<FocusIn>", fi)
        entry.bind("<FocusOut>", fo)
        return var

    def _refresh(self):
        self.tree.delete(*self.tree.get_children())
        for row in db_list_users():
            tag = "admin" if row[1] == "admin" else ""
            self.tree.insert("", "end", values=row, tags=(tag,))
        self.tree.tag_configure("admin", foreground=GOLD)

    def _on_select(self, _=None):
        sel = self.tree.selection()
        if not sel:
            return
        vals = self.tree.item(sel[0], "values")
        self._form_username.set(vals[0])
        self._form_password.set("")
        self._form_confirm.set("")
        self._role_var.set(vals[1])
        self._form_err.set("")

    def _clear_form(self):
        self._form_username.set("")
        self._form_password.set("")
        self._form_confirm.set("")
        self._role_var.set("user")
        self._form_err.set("")

    def _add_user(self):
        u  = self._form_username.get().strip()
        p  = self._form_password.get().strip()
        p2 = self._form_confirm.get().strip()
        r  = self._role_var.get()
        if not u or not p:
            self._form_err.set("⚠ Username and password are required.")
            return
        if p != p2:
            self._form_err.set("⚠ Passwords do not match.")
            return
        err = db_add_user(u, p, r)
        if err:
            self._form_err.set(f"✗ {err}")
        else:
            self._form_err.set("")
            self._clear_form()
            self._refresh()
            show_toast(self.winfo_toplevel(), f"User '{u}' added.", color=SUCCESS)

    def _update_user(self):
        sel = self.tree.selection()
        if not sel:
            self._form_err.set("⚠ Select a user from the table first.")
            return
        original_username = self.tree.item(sel[0], "values")[0]
        p  = self._form_password.get().strip()
        p2 = self._form_confirm.get().strip()
        r  = self._role_var.get()
        if p and p != p2:
            self._form_err.set("⚠ Passwords do not match.")
            return
        # Protect last admin
        if r != "admin":
            rows = db_list_users()
            admins = [row for row in rows if row[1] == "admin"]
            if len(admins) == 1 and admins[0][0] == original_username:
                self._form_err.set("✗ Cannot remove admin role from the only admin.")
                return
        db_update_user(original_username, new_password=p, new_role=r)
        self._form_err.set("")
        self._clear_form()
        self._refresh()
        show_toast(self.winfo_toplevel(), f"User '{original_username}' updated.", color=SUCCESS)

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        username = self.tree.item(sel[0], "values")[0]
        if username == self.username:
            messagebox.showwarning("Cannot Delete", "You cannot delete your own account.", parent=self)
            return
        if not messagebox.askyesno("Confirm Delete",
                                   f"Delete user '{username}'? This cannot be undone.",
                                   parent=self):
            return
        err = db_delete_user(username)
        if err:
            messagebox.showerror("Error", err, parent=self)
        else:
            self._refresh()
            show_toast(self.winfo_toplevel(), f"User '{username}' deleted.", color=WARNING)


# ══════════════════════════════════════════════════════════════════════════════
# ROOT  — navigation controller
# ══════════════════════════════════════════════════════════════════════════════
_AppBase = TkinterDnD.Tk if _DND_AVAILABLE else tk.Tk

class App(_AppBase):
    def __init__(self):
        super().__init__()
        self.title("S.P. Madrid Philippines")
        self.geometry("1180x760")
        self.minsize(900, 620)
        self.configure(bg=BG)
        self.resizable(True, True)
        self._username = None
        self._role     = None
        self._current  = None
        self._show_login()

    # ── screen switcher ───────────────────────────────────────────────────────
    def _show(self, screen):
        if self._current:
            self._current.destroy()
        self._current = screen
        self._current.pack(fill="both", expand=True)

    def _show_login(self):
        self._show(LoginScreen(self, on_login=self._on_login))

    def _on_login(self, username, role):
        self._username = username
        self._role     = role
        self._show_dashboard()

    def _show_dashboard(self):
        self._show(DashboardScreen(
            self,
            username=self._username,
            role=self._role,
            on_select=self._on_account_select,
            on_logout=self._show_login,
            on_manage_accounts=self._show_account_management,
        ))

    def _show_account_management(self):
        self._show(AccountManagementScreen(
            self,
            username=self._username,
            on_back=self._show_dashboard,
        ))

    def _on_account_select(self, account):
        if account == "Orico Auto Loan":
            self._show(OricoToolsScreen(
                self,
                username=self._username,
                role=self._role,
                on_tool=self._on_tool_select,
                on_back=self._show_dashboard,
            ))
        elif account == "Bank of Makati":
            self._show(BankOfMakatiScreen(
                self,
                username=self._username,
                on_back=self._show_dashboard,
            ))

    def _on_tool_select(self, tool_tag):
        if tool_tag == "drr_cleaner":
            self._show(DRRCleanerScreen(
                self,
                username=self._username,
                role=self._role,
                on_back=lambda: self._on_account_select("Orico Auto Loan"),
            ))
        elif tool_tag == "autostat":
            self._show(AutostatScreen(
                self,
                username=self._username,
                role=self._role,
                on_back=lambda: self._on_account_select("Orico Auto Loan"),
            ))
        elif tool_tag == "analytical_reports":
            self._show(AnalyticalReportsScreen(
                self,
                username=self._username,
                role=self._role,
                on_back=lambda: self._on_account_select("Orico Auto Loan"),
            ))


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()