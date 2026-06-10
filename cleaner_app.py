import tkinter as tk
from tkinter import ttk, messagebox
import threading
import os
import io
import pandas as pd
from openpyxl.styles import PatternFill, Font
import openpyxl

# ── CONFIG ────────────────────────────────────────────────────────────────────
STATUS_COL_NAME    = "Status"
STATUS_COL_INDEX   = 9
DATE_COL_NAME      = "Date"
TIME_COL_NAME      = "Time"
COL_E_INDEX        = 4
PTP_AMOUNT_COL     = "PTP Amount"
CLAIM_PAID_COL     = "Claim Paid Amount"
REMARK_COL         = "Remark"
REMOVE_STATUSES    = {"BP", "REACTIVE", "SMS FAILED", "NEW", "ABORTED"}
COLS_TO_DROP_START = 27
COLS_TO_DROP_END   = 50

# ── COLORS ────────────────────────────────────────────────────────────────────
BG        = "#0F1117"
CARD      = "#1A1D27"
BORDER    = "#2A2D3E"
ACCENT    = "#4F8EF7"
ACCENT2   = "#7C3AED"
SUCCESS   = "#22C55E"
DANGER    = "#EF4444"
WARNING   = "#F59E0B"
TEXT      = "#F1F5F9"
SUBTEXT   = "#94A3B8"
HEADER_BG = "#1E2235"

# ── PROCESSING LOGIC ─────────────────────────────────────────────────────────
def process_file(filepath):
    df_headers = pd.read_excel(filepath, nrows=0)
    col_e_name = df_headers.columns[COL_E_INDEX] if COL_E_INDEX < len(df_headers.columns) else None
    dtype_override = {col_e_name: str} if col_e_name else {}
    df = pd.read_excel(filepath, dtype=dtype_override)

    if col_e_name and col_e_name in df.columns:
        def clean_account(x):
            if pd.isna(x) or str(x).strip() in ("", "nan"):
                return ""
            s = str(x).strip()
            if s.endswith(".0") and s[:-2].lstrip("-").isdigit() and not s[:-2].startswith("0"):
                return s[:-2]
            return s
        df[col_e_name] = df[col_e_name].apply(clean_account)

    STATUS_COL = STATUS_COL_NAME if STATUS_COL_NAME in df.columns else (
        df.columns[STATUS_COL_INDEX] if STATUS_COL_INDEX < len(df.columns) else None
    )
    if not STATUS_COL:
        raise ValueError("Status column not found.")

    status_normalized = df[STATUS_COL].astype(str).str.strip().str.upper()
    is_blank          = df[STATUS_COL].isna() | (df[STATUS_COL].astype(str).str.strip() == "")
    has_cease         = status_normalized.str.contains("CEASE", na=False)
    is_removable      = status_normalized.isin(REMOVE_STATUSES) | is_blank | has_cease

    if PTP_AMOUNT_COL in df.columns:
        has_ptp     = status_normalized.str.contains(r"\bPTP\b", regex=True, na=False)
        ptp_numeric = pd.to_numeric(df[PTP_AMOUNT_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        ptp_has_value = has_ptp & (ptp_numeric > 0)
        ptp_no_value  = has_ptp & (ptp_numeric <= 0)
        is_removable  = (is_removable | ptp_no_value) & ~ptp_has_value

    if CLAIM_PAID_COL in df.columns:
        has_kept      = status_normalized.str.contains(r"\bKEPT\b", regex=True, na=False)
        claim_numeric = pd.to_numeric(df[CLAIM_PAID_COL].astype(str).str.replace(",", "", regex=False), errors="coerce").fillna(0)
        kept_has_value = has_kept & (claim_numeric > 0)
        kept_no_value  = has_kept & (claim_numeric <= 0)
        is_removable   = (is_removable | kept_no_value) & ~kept_has_value

    cleaned_df = df[~is_removable].copy()
    removed_df = df[is_removable].copy()

    if REMARK_COL in cleaned_df.columns and PTP_AMOUNT_COL in cleaned_df.columns and CLAIM_PAID_COL in cleaned_df.columns:
        remark_norm    = cleaned_df[REMARK_COL].astype(str).str.strip().str.upper()
        status_clean   = cleaned_df[STATUS_COL].astype(str).str.strip().str.upper()
        has_action_ptp = remark_norm.str.contains(r"ACTION\s*:\s*PTP", regex=True, na=False)
        no_ptp_kept    = (~status_clean.str.contains(r"PTP", regex=True, na=False) &
                          ~status_clean.str.contains(r"KEPT", regex=True, na=False))
        ptp_amt   = pd.to_numeric(cleaned_df[PTP_AMOUNT_COL].astype(str).str.replace(",","",regex=False), errors="coerce").fillna(0)
        claim_amt = pd.to_numeric(cleaned_df[CLAIM_PAID_COL].astype(str).str.replace(",","",regex=False), errors="coerce").fillna(0)
        srp_mask  = has_action_ptp & no_ptp_kept & (ptp_amt > 0) & (claim_amt > 0)
        cleaned_df.loc[srp_mask, REMARK_COL] = cleaned_df.loc[srp_mask, REMARK_COL].astype(str).str.replace(
            r"(?i)Action\s*:\s*PTP", "Action: SRP", regex=True)

    cols_to_drop = [df.columns[i] for i in range(COLS_TO_DROP_START, min(COLS_TO_DROP_END + 1, len(df.columns)))]
    cleaned_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")
    removed_df.drop(columns=cols_to_drop, inplace=True, errors="ignore")

    if DATE_COL_NAME in cleaned_df.columns:
        cleaned_df[DATE_COL_NAME] = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m-%d-%Y")

    if DATE_COL_NAME in cleaned_df.columns and TIME_COL_NAME in cleaned_df.columns:
        date_str = pd.to_datetime(cleaned_df[DATE_COL_NAME], errors="coerce").dt.strftime("%m-%d-%Y")
        time_str = pd.to_datetime(cleaned_df[TIME_COL_NAME], errors="coerce").dt.strftime("%I:%M:%S %p")
        cleaned_df[TIME_COL_NAME] = date_str + " " + time_str

    stats = {
        "total": len(df),
        "retained": len(cleaned_df),
        "removed": len(removed_df),
        "srp_changed": int(srp_mask.sum()) if REMARK_COL in cleaned_df.columns else 0,
    }

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        cleaned_df.to_excel(writer, index=False, sheet_name="Cleaned")
        removed_df.to_excel(writer, index=False, sheet_name="Removed")
        for sheet_name, frame in [("Cleaned", cleaned_df), ("Removed", removed_df)]:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1E2235")
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    buf.seek(0)
    return cleaned_df, removed_df, stats, buf.read(), col_e_name


# ── APP ───────────────────────────────────────────────────────────────────────
class CleanerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Cleaner")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(bg=BG)
        self.resizable(True, True)

        self.file_path   = None
        self.output_bytes = None
        self.cleaned_df  = None
        self.removed_df  = None

        self._build_ui()

    def _build_ui(self):
        # ── Header bar ────────────────────────────────────────────────────────
        header = tk.Frame(self, bg=HEADER_BG, height=56)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="⬡  Excel Data Cleaner", font=("Segoe UI", 14, "bold"),
                 bg=HEADER_BG, fg=TEXT).pack(side="left", padx=20, pady=14)
        tk.Label(header, text="Drop · Preview · Verify · Download",
                 font=("Segoe UI", 9), bg=HEADER_BG, fg=SUBTEXT).pack(side="left", padx=4, pady=14)

        # ── Main content ──────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        # Left panel
        left = tk.Frame(body, bg=BG, width=300)
        left.pack(side="left", fill="y", padx=(0, 14))
        left.pack_propagate(False)

        # Drop zone
        self.drop_frame = tk.Frame(left, bg=CARD, bd=0, relief="flat",
                                   highlightthickness=2, highlightbackground=BORDER)
        self.drop_frame.pack(fill="x", pady=(0, 14))

        self.drop_inner = tk.Frame(self.drop_frame, bg=CARD)
        self.drop_inner.pack(fill="both", padx=20, pady=28)

        tk.Label(self.drop_inner, text="📂", font=("Segoe UI", 28),
                 bg=CARD, fg=ACCENT).pack()
        tk.Label(self.drop_inner, text="Drop Excel file here",
                 font=("Segoe UI", 11, "bold"), bg=CARD, fg=TEXT).pack(pady=(6, 2))
        tk.Label(self.drop_inner, text=".xlsx  ·  .xlsm  ·  .xls",
                 font=("Segoe UI", 8), bg=CARD, fg=SUBTEXT).pack()

        # Enable drag-and-drop via tkinterdnd2 if available, else browse button
        try:
            self.drop_frame.drop_target_register('DND_Files')
            self.drop_frame.dnd_bind('<<Drop>>', self._on_drop)
            tk.Label(self.drop_inner, text="or", font=("Segoe UI", 8),
                     bg=CARD, fg=SUBTEXT).pack(pady=(10, 4))
        except Exception:
            pass

        browse_btn = tk.Button(self.drop_inner, text="Browse File",
                               font=("Segoe UI", 9, "bold"),
                               bg=ACCENT, fg="white", relief="flat",
                               padx=16, pady=6, cursor="hand2",
                               command=self._browse)
        browse_btn.pack()

        self.file_label = tk.Label(left, text="No file selected",
                                   font=("Segoe UI", 8), bg=BG, fg=SUBTEXT,
                                   wraplength=270, justify="left")
        self.file_label.pack(fill="x", pady=(0, 14))

        # Stats cards
        stats_title = tk.Label(left, text="SUMMARY", font=("Segoe UI", 8, "bold"),
                                bg=BG, fg=SUBTEXT)
        stats_title.pack(anchor="w", pady=(0, 6))

        self.stat_vars = {}
        stat_defs = [
            ("total",    "Total Rows",     TEXT,    "📋"),
            ("retained", "Rows Retained",  SUCCESS, "✅"),
            ("removed",  "Rows Removed",   DANGER,  "🗑"),
            ("srp",      "Remarks Fixed",  WARNING, "✏️"),
        ]
        for key, label, color, icon in stat_defs:
            card = tk.Frame(left, bg=CARD, bd=0, relief="flat",
                            highlightthickness=1, highlightbackground=BORDER)
            card.pack(fill="x", pady=3)
            inner = tk.Frame(card, bg=CARD)
            inner.pack(fill="x", padx=12, pady=8)
            tk.Label(inner, text=icon, font=("Segoe UI", 11), bg=CARD, fg=color).pack(side="left")
            tk.Label(inner, text=label, font=("Segoe UI", 9), bg=CARD, fg=SUBTEXT).pack(side="left", padx=8)
            var = tk.StringVar(value="—")
            self.stat_vars[key] = var
            tk.Label(inner, textvariable=var, font=("Segoe UI", 11, "bold"),
                     bg=CARD, fg=color).pack(side="right")

        # Process + Download buttons
        self.process_btn = tk.Button(left, text="▶  Run Cleaner",
                                     font=("Segoe UI", 10, "bold"),
                                     bg=ACCENT, fg="white", relief="flat",
                                     padx=0, pady=10, cursor="hand2",
                                     state="disabled", command=self._run_process)
        self.process_btn.pack(fill="x", pady=(16, 6))

        self.download_btn = tk.Button(left, text="⬇  Download cleaned_output.xlsx",
                                      font=("Segoe UI", 9, "bold"),
                                      bg=SUCCESS, fg="white", relief="flat",
                                      padx=0, pady=10, cursor="hand2",
                                      state="disabled", command=self._download)
        self.download_btn.pack(fill="x")

        # Progress bar
        self.progress = ttk.Progressbar(left, mode="indeterminate")
        self.progress.pack(fill="x", pady=(10, 0))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TProgressbar", troughcolor=CARD, background=ACCENT, thickness=4)

        # ── Right panel: preview ───────────────────────────────────────────────
        right = tk.Frame(body, bg=BG)
        right.pack(side="left", fill="both", expand=True)

        # Tab bar
        tab_bar = tk.Frame(right, bg=BG)
        tab_bar.pack(fill="x", pady=(0, 8))

        self.active_tab = tk.StringVar(value="cleaned")
        self.tab_btns   = {}
        for tab_id, tab_label in [("cleaned", "Cleaned Rows"), ("removed", "Removed Rows")]:
            btn = tk.Button(tab_bar, text=tab_label,
                            font=("Segoe UI", 9, "bold"),
                            bg=ACCENT if tab_id == "cleaned" else CARD,
                            fg="white", relief="flat",
                            padx=16, pady=7, cursor="hand2",
                            command=lambda t=tab_id: self._switch_tab(t))
            btn.pack(side="left", padx=(0, 6))
            self.tab_btns[tab_id] = btn

        # Search bar
        search_frame = tk.Frame(right, bg=CARD, highlightthickness=1, highlightbackground=BORDER)
        search_frame.pack(fill="x", pady=(0, 8))
        tk.Label(search_frame, text="🔍", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 10)).pack(side="left", padx=8)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self._on_search)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var,
                                font=("Segoe UI", 9), bg=CARD, fg=TEXT,
                                insertbackground=TEXT, relief="flat", bd=0)
        search_entry.pack(side="left", fill="x", expand=True, pady=8, padx=4)
        tk.Label(search_frame, text="Filter rows…", bg=CARD, fg=SUBTEXT,
                 font=("Segoe UI", 8)).pack(side="right", padx=8)

        # Table frame
        table_frame = tk.Frame(right, bg=BG)
        table_frame.pack(fill="both", expand=True)

        # Scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        hsb = ttk.Scrollbar(table_frame, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        style.configure("Custom.Treeview",
                         background=CARD, foreground=TEXT,
                         fieldbackground=CARD, rowheight=28,
                         font=("Segoe UI", 9), borderwidth=0)
        style.configure("Custom.Treeview.Heading",
                         background=HEADER_BG, foreground=TEXT,
                         font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT2)],
                  foreground=[("selected", "white")])

        self.tree = ttk.Treeview(table_frame, style="Custom.Treeview",
                                  yscrollcommand=vsb.set, xscrollcommand=hsb.set,
                                  show="headings", selectmode="browse")
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        self.row_count_label = tk.Label(right, text="", font=("Segoe UI", 8),
                                         bg=BG, fg=SUBTEXT)
        self.row_count_label.pack(anchor="e", pady=(4, 0))

        # Status bar
        self.status_var = tk.StringVar(value="Ready — drop or browse an Excel file to begin.")
        status_bar = tk.Label(self, textvariable=self.status_var,
                              font=("Segoe UI", 8), bg=HEADER_BG, fg=SUBTEXT,
                              anchor="w", padx=16, pady=6)
        status_bar.pack(fill="x", side="bottom")

    # ── DRAG & DROP ───────────────────────────────────────────────────────────
    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        self._load_file(path)

    def _browse(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
        )
        if path:
            self._load_file(path)

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("File not found", f"Cannot find:\n{path}")
            return
        self.file_path = path
        name = os.path.basename(path)
        self.file_label.config(text=f"📄 {name}", fg=ACCENT)
        self.process_btn.config(state="normal")
        self.download_btn.config(state="disabled")
        self.status_var.set(f"File loaded: {name}  —  Click 'Run Cleaner' to process.")
        self._clear_table()
        for k in self.stat_vars:
            self.stat_vars[k].set("—")

    # ── PROCESSING ───────────────────────────────────────────────────────────
    def _run_process(self):
        if not self.file_path:
            return
        self.process_btn.config(state="disabled")
        self.download_btn.config(state="disabled")
        self.status_var.set("Processing…")
        self.progress.start(10)
        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        try:
            cleaned, removed, stats, out_bytes, col_e = process_file(self.file_path)
            self.after(0, lambda: self._on_done(cleaned, removed, stats, out_bytes))
        except Exception as e:
            self.after(0, lambda: self._on_error(str(e)))

    def _on_done(self, cleaned, removed, stats, out_bytes):
        self.progress.stop()
        self.cleaned_df  = cleaned
        self.removed_df  = removed
        self.output_bytes = out_bytes

        self.stat_vars["total"].set(str(stats["total"]))
        self.stat_vars["retained"].set(str(stats["retained"]))
        self.stat_vars["removed"].set(str(stats["removed"]))
        self.stat_vars["srp"].set(str(stats["srp_changed"]))

        self.download_btn.config(state="normal")
        self.process_btn.config(state="normal")
        self.status_var.set(
            f"Done  ·  {stats['retained']} rows retained  ·  {stats['removed']} rows removed  ·  "
            f"{stats['srp_changed']} remarks changed to SRP"
        )
        self._switch_tab(self.active_tab.get())

    def _on_error(self, msg):
        self.progress.stop()
        self.process_btn.config(state="normal")
        self.status_var.set(f"Error: {msg}")
        messagebox.showerror("Processing Error", msg)

    # ── TABLE ─────────────────────────────────────────────────────────────────
    def _switch_tab(self, tab_id):
        self.active_tab.set(tab_id)
        for tid, btn in self.tab_btns.items():
            btn.config(bg=ACCENT if tid == tab_id else CARD)
        df = self.cleaned_df if tab_id == "cleaned" else self.removed_df
        if df is not None:
            self._populate_table(df)

    def _populate_table(self, df, filter_text=""):
        self._clear_table()
        if df is None or df.empty:
            self.row_count_label.config(text="No rows to display.")
            return

        cols = list(df.columns[:20])  # show first 20 cols max for performance
        self.tree["columns"] = cols
        for col in cols:
            self.tree.heading(col, text=col, anchor="w")
            self.tree.column(col, width=120, minwidth=60, anchor="w")

        filt = filter_text.lower()
        count = 0
        for _, row in df.iterrows():
            vals = [str(v) if pd.notna(v) else "" for v in row[cols]]
            if filt and not any(filt in v.lower() for v in vals):
                continue
            tag = "even" if count % 2 == 0 else "odd"
            self.tree.insert("", "end", values=vals, tags=(tag,))
            count += 1

        self.tree.tag_configure("even", background=CARD)
        self.tree.tag_configure("odd",  background="#14172A")
        total = len(df)
        self.row_count_label.config(
            text=f"Showing {count} of {total} rows" + (f"  ·  {len(df.columns)} columns (first 20 shown)" if len(df.columns) > 20 else f"  ·  {len(df.columns)} columns")
        )

    def _clear_table(self):
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = []
        self.row_count_label.config(text="")

    def _on_search(self, *_):
        tab  = self.active_tab.get()
        df   = self.cleaned_df if tab == "cleaned" else self.removed_df
        if df is not None:
            self._populate_table(df, self.search_var.get())

    # ── DOWNLOAD ─────────────────────────────────────────────────────────────
    def _download(self):
        from tkinter import filedialog
        default_name = "cleaned_output.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Save cleaned file as",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            with open(save_path, "wb") as f:
                f.write(self.output_bytes)
            self.status_var.set(f"Saved → {save_path}")
            messagebox.showinfo("Saved", f"File saved to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Save Error", str(e))


if __name__ == "__main__":
    app = CleanerApp()
    app.mainloop()
